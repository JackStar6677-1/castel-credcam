from __future__ import annotations

import csv
import json
import logging
import subprocess
import sys
import threading
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
from openpyxl import load_workbook
from PySide6.QtCore import QSignalBlocker, QThread, QTimer, Qt, Signal
from PySide6.QtGui import QFont, QImage, QKeySequence, QPixmap, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QListWidget,
    QListWidgetItem,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
    QHeaderView,
)

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from castel_credcam import (  # noqa: E402
    BACKUP_PHOTOS_DIRNAME,
    CSV_FILENAME,
    PHOTOS_DIRNAME,
    TEST_FOLDER_NAME,
    PhotoRecord,
    append_csv_record,
    build_photo_filename,
    configure_capture,
    backend_key_from_id,
    camera_source_kind,
    ensure_photo_backup,
    get_logs_dir,
    list_available_cameras,
    load_camera_aliases,
    load_camera_resolution,
    load_existing_records,
    load_last_camera,
    open_camera,
    open_folder,
    rewrite_csv,
    sanitize_folder_name,
    save_last_camera,
    save_camera_resolution,
    setup_logging,
    silence_opencv_logs,
)


APP_TITLE = "CastelCredCam Studio"
WINDOW_BG = "#141317"
PANEL_BG = "#1C1B21"
CARD_BG = "#23222A"
CARD_EDGE = "#383545"
INFO_BG = "#17161B"
TEXT_PRIMARY = "#F2EEF7"
TEXT_MUTED = "#C7C1CF"
ACCENT_PURPLE = "#6A4A98"
ACCENT_GOLD = "#C4B06A"
ACCENT_GOLD_DARK = "#2D2816"
SUCCESS = "#6FBFA5"
PENDING = "#B8AEC9"
DANGER = "#C87182"
DONE_BG = "#1D2722"
CURRENT_BG = "#352C43"
PENDING_BG = "#25222E"
COMMON_CAMERA_RESOLUTIONS: list[tuple[int, int]] = [
    (1280, 720),
    (1920, 1080),
    (1600, 1200),
    (1280, 960),
    (960, 720),
    (640, 480),
    (640, 360),
    (2560, 1440),
]
RESOLUTION_AUTO_LABEL = "Automatico"
FACE_DETECT_MAX_WIDTH = 960
PREVIEW_MAX_WIDTH = 640
FACE_HOLD_FRAMES = 6
CROP_MIN_HEIGHT = 220
CROP_MANUAL_STEP = 0.035
CROP_MANUAL_ZOOM_STEP = 1.07
LOCAL_CONFIG_FILENAME = "local_config.json"
DEFAULT_EXTERNAL_DATA_ROOT = Path("D:/Colegio/Fotos_Perfil_Estudiantes_Castel")
DEFAULT_ROSTER_DIR = Path("D:/Colegio/Nominas_Castel_2026")
PREFERRED_ROSTER_NAMES = (
    "NÓMINA DE ESTUDIANTES 2026.xlsx",
    "NOMINA DE ESTUDIANTES 2026.xlsx",
    "CASTELGANDOLFO_lista_cursos_2026.xlsx",
)


def _local_config_path() -> Path:
    return APP_ROOT / LOCAL_CONFIG_FILENAME


def _read_local_config() -> dict[str, object]:
    path = _local_config_path()
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _write_local_config(payload: dict[str, object]) -> None:
    _local_config_path().write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _valid_data_root(path: Path) -> bool:
    return (path / PHOTOS_DIRNAME).is_dir() and (path / BACKUP_PHOTOS_DIRNAME).is_dir()


def _detect_data_root(config: dict[str, object]) -> Path:
    configured = config.get("data_root")
    candidates = []
    if isinstance(configured, str) and configured.strip():
        candidates.append(Path(configured))
    candidates.append(DEFAULT_EXTERNAL_DATA_ROOT)
    candidates.append(APP_ROOT)
    for candidate in candidates:
        try:
            if _valid_data_root(candidate):
                return candidate
        except OSError:
            continue
    return APP_ROOT


def _detect_default_roster(config: dict[str, object]) -> Optional[Path]:
    configured = config.get("default_roster")
    if isinstance(configured, str) and configured.strip():
        path = Path(configured)
        if path.is_file():
            return path
    if not DEFAULT_ROSTER_DIR.is_dir():
        return None
    for name in PREFERRED_ROSTER_NAMES:
        path = DEFAULT_ROSTER_DIR / name
        if path.is_file():
            return path
    candidates = [
        path
        for path in DEFAULT_ROSTER_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm", ".csv", ".tsv", ".txt"}
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip()).casefold()
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _normalize_rut(value: object) -> str:
    text = str(value or "").strip().upper()
    return "".join(ch for ch in text if ch.isdigit() or ch == "K")


def _display_name_from_parts(
    apellido_paterno: str = "",
    apellido_materno: str = "",
    segundo_nombre: str = "",
    primer_nombre: str = "",
    full_name: str = "",
) -> str:
    pieces = [apellido_paterno.strip(), apellido_materno.strip(), segundo_nombre.strip(), primer_nombre.strip()]
    name = " ".join(part for part in pieces if part)
    return name or full_name.strip()


def _sort_key(student: "RosterStudent") -> tuple[str, str, str, str, str]:
    return (
        _normalize_text(student.apellido_paterno),
        _normalize_text(student.apellido_materno),
        _normalize_text(student.segundo_nombre),
        _normalize_text(student.primer_nombre),
        _normalize_text(student.full_name),
    )


def _sanitize_filename_component(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("/", "-").replace("\\", "-")
    text = "".join(ch for ch in text if ch.isalnum() or ch in {" ", "-", "_"})
    text = " ".join(text.split())
    return text or "sin_nombre"


def _center_crop_to_aspect(frame: np.ndarray, target_aspect: float = 3 / 4) -> np.ndarray:
    height, width = frame.shape[:2]
    if height <= 0 or width <= 0:
        return frame

    current_aspect = width / height
    if abs(current_aspect - target_aspect) < 0.01:
        return frame

    if current_aspect > target_aspect:
        new_width = max(1, int(height * target_aspect))
        x0 = max(0, (width - new_width) // 2)
        return frame[:, x0 : x0 + new_width]

    new_height = max(1, int(width / target_aspect))
    y0 = max(0, (height - new_height) // 2)
    return frame[y0 : y0 + new_height, :]


def _rotate_frame(frame: np.ndarray, rotation_label: str) -> np.ndarray:
    if rotation_label == "90 deg":
        return cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)
    if rotation_label == "180 deg":
        return cv2.rotate(frame, cv2.ROTATE_180)
    if rotation_label == "270 deg":
        return cv2.rotate(frame, cv2.ROTATE_90_COUNTERCLOCKWISE)
    return frame


def _cv_to_qpixmap(frame: np.ndarray) -> QPixmap:
    height, width, channels = frame.shape
    if hasattr(QImage.Format, "Format_BGR888"):
        image = QImage(frame.data, width, height, channels * width, QImage.Format.Format_BGR888).copy()
    else:
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        height, width, channels = rgb.shape
        image = QImage(rgb.data, width, height, channels * width, QImage.Format.Format_RGB888).copy()
    return QPixmap.fromImage(image)


def _student_key(student_name: str, rut: str) -> tuple[str, str]:
    rut_key = _normalize_rut(rut)
    if rut_key:
        return ("rut", rut_key)
    return ("name", _normalize_text(student_name))


@dataclass
class RosterStudent:
    rut: str = ""
    apellido_paterno: str = ""
    apellido_materno: str = ""
    segundo_nombre: str = ""
    primer_nombre: str = ""
    full_name: str = ""

    @property
    def display_name(self) -> str:
        return _display_name_from_parts(
            self.apellido_paterno,
            self.apellido_materno,
            self.segundo_nombre,
            self.primer_nombre,
            self.full_name,
        )

    @property
    def key(self) -> tuple[str, str]:
        return _student_key(self.display_name, self.rut)


@dataclass
class SessionState:
    mode: str
    course_display: str
    course_slug: str
    photos_root: Path
    backup_root: Path
    session_dir: Path
    backup_dir: Path
    csv_path: Path
    records: list[PhotoRecord]
    started_at: datetime
    roster_students: list[RosterStudent] = field(default_factory=list)
    roster_index: int = 0

    @property
    def next_id(self) -> int:
        return len(self.records) + 1

    @property
    def has_roster(self) -> bool:
        return bool(self.roster_students)

    @property
    def roster_total(self) -> int:
        return len(self.roster_students)

    @property
    def roster_remaining(self) -> int:
        return max(0, len(self.roster_students) - min(self.roster_index, len(self.roster_students)))

    def current_student(self) -> Optional[RosterStudent]:
        if not self.roster_students:
            return None
        if self.roster_index < 0 or self.roster_index >= len(self.roster_students):
            return None
        return self.roster_students[self.roster_index]

    def advance(self) -> Optional[RosterStudent]:
        if not self.roster_students:
            return None
        self.roster_index = min(self.roster_index + 1, len(self.roster_students))
        return self.current_student()

    def retreat(self) -> Optional[RosterStudent]:
        if not self.roster_students:
            return None
        self.roster_index = max(0, self.roster_index - 1)
        return self.current_student()


class CameraThread(QThread):
    frame_ready = Signal(object)
    camera_message = Signal(str)
    camera_error = Signal(str)

    def __init__(
        self,
        camera_index: int,
        backend_id: int,
        preferred_resolution: Optional[tuple[int, int]] = None,
        parent: Optional[QWidget] = None,
    ) -> None:
        super().__init__(parent)
        self.camera_index = camera_index
        self.backend_id = backend_id
        self.preferred_resolution = preferred_resolution
        self._capture: Optional[cv2.VideoCapture] = None
        self._frame_lock = threading.Lock()
        self._latest_frame: Optional[np.ndarray] = None

    def run(self) -> None:
        try:
            self._capture = open_camera(self.camera_index, self.backend_id)
            if not self._capture.isOpened():
                self.camera_error.emit(f"No se pudo abrir la camara {self.camera_index}.")
                return

            requested_width, requested_height = configure_capture(self._capture, self.preferred_resolution)
            for _ in range(8):
                if self.isInterruptionRequested():
                    return
                self._capture.read()

            actual_width = int(self._capture.get(cv2.CAP_PROP_FRAME_WIDTH) or requested_width or 0)
            actual_height = int(self._capture.get(cv2.CAP_PROP_FRAME_HEIGHT) or requested_height or 0)
            self.camera_message.emit(
                f"Camara activa {self.camera_index} ({actual_width}x{actual_height})"
            )
            frame_count = 0
            fail_count = 0
            while not self.isInterruptionRequested():
                ok, frame = self._capture.read()
                if not ok or frame is None:
                    fail_count += 1
                    if fail_count == 1 or fail_count % 25 == 0:
                        self.camera_message.emit(
                            f"Esperando frames... intento {fail_count} en camara {self.camera_index}"
                        )
                    self.msleep(120)
                    continue
                fail_count = 0
                frame_count += 1
                with self._frame_lock:
                    self._latest_frame = frame
                if frame_count == 1 or frame_count % 60 == 0:
                    self.camera_message.emit(f"Frame recibido #{frame_count} en camara {self.camera_index}")
                self.msleep(33)
        except Exception as exc:
            self.camera_error.emit(str(exc))
        finally:
            if self._capture is not None:
                try:
                    self._capture.release()
                except Exception:
                    pass
                self._capture = None

    def latest_frame(self) -> Optional[np.ndarray]:
        with self._frame_lock:
            if self._latest_frame is None:
                return None
            return self._latest_frame.copy()

    def stop(self) -> None:
        self.requestInterruption()
        if self._capture is not None:
            try:
                self._capture.release()
            except Exception:
                pass


class CastelCredCamQt(QMainWindow):
    autoframe_finished = Signal(bool, str)
    post_progress_updated = Signal(int, str)
    camera_scan_ready = Signal(list)

    def __init__(self) -> None:
        super().__init__()
        silence_opencv_logs()
        self.logger, self.log_path = setup_logging(APP_ROOT, "gui_qt")
        self.logger.info("=== CastelCredCam Qt GUI start ===")
        self.logger.info("Log file: %s", self.log_path)
        self.logger.info("Logs dir: %s", get_logs_dir(APP_ROOT))
        self.logger.info("Python: %s", sys.version.replace("\n", " "))
        self.logger.info("Executable: %s", sys.executable)
        self.logger.info("CWD: %s", Path.cwd())
        self.logger.info("Args: %s", sys.argv[1:])

        self.local_config = _read_local_config()
        self.data_root = _detect_data_root(self.local_config)
        self.default_roster_path = _detect_default_roster(self.local_config)
        self.logger.info("Data root: %s", self.data_root)
        self.logger.info("Default roster: %s", self.default_roster_path or "not configured")

        self.aliases = load_camera_aliases(APP_ROOT)
        self.available_cameras: list[tuple[int, str, int, str, str]] = []
        self._camera_scan_running = False
        self.camera_index = 0
        self.backend_id = cv2.CAP_ANY
        self.backend_name = "Automatico"
        self.camera_alias = ""
        self.camera_resolution: Optional[tuple[int, int]] = None
        self.camera_thread: Optional[CameraThread] = None
        self.latest_frame: Optional[np.ndarray] = None
        self.frame_counter = 0
        self.current_face_box: Optional[tuple[int, int, int, int]] = None
        self.current_eye_centers: list[tuple[int, int]] = []
        self.current_crop_box: Optional[tuple[int, int, int, int]] = None
        self.stable_crop_box: Optional[tuple[int, int, int, int]] = None
        self.preview_stable_crop_box: Optional[tuple[int, int, int, int]] = None
        self.preview_face_box: Optional[tuple[int, int, int, int]] = None
        self.preview_eye_centers: list[tuple[int, int]] = []
        self.preview_detection_tick = 0
        self.last_face_detect_frame = -9999
        self.crop_manual_dx = 0.0
        self.crop_manual_dy = 0.0
        self.crop_manual_zoom = 1.0
        self.pending_capture_identity: Optional[tuple[str, str, Optional[int]]] = None
        self.face_cascades = [
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_alt2.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_alt.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_profileface.xml"),
        ]
        self.eye_cascades = [
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye_tree_eyeglasses.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye.xml"),
        ]
        self.session: Optional[SessionState] = None
        self.roster_map: dict[str, list[RosterStudent]] = {}
        self.roster_order: list[str] = []
        self.preview_index_by_course: dict[str, int] = {}
        self.autoframe_jobs: dict[Path, threading.Thread] = {}
        self.autoframe_jobs_lock = threading.Lock()
        self.course_table_syncing = False
        self.countdown_remaining = 0
        self.countdown_timer = QTimer(self)
        self.countdown_timer.setInterval(1000)
        self.countdown_timer.timeout.connect(self._countdown_tick)

        self.is_processing_photo = False
        self.autoframe_finished.connect(self._on_autoframe_finished)
        self.camera_scan_ready.connect(self._on_camera_scan_ready)
        self.post_progress_updated.connect(self._on_post_progress_updated)
        self.preview_render_timer = QTimer(self)
        self.preview_render_timer.setSingleShot(True)
        self.preview_render_timer.setInterval(33)
        self.preview_render_timer.timeout.connect(self._render_preview_safe)
        self.first_frame_timer = QTimer(self)
        self.first_frame_timer.setSingleShot(True)
        self.first_frame_timer.timeout.connect(self._on_first_frame_timeout)
        self._preview_frame_received = False
        self._auto_camera_failures: set[tuple[int, int]] = set()

        self.roster_source_path: Optional[Path] = None
        self.roster_status_text = "Carga una lista para avanzar por curso."

        self.mode = "test"
        self.course_text = ""

        self.setWindowTitle(APP_TITLE)
        self.resize(1600, 900)
        self.setMinimumSize(1280, 700)
        self.setStyleSheet(self._build_stylesheet())

        self._build_ui()
        self._update_crop_tuning_label()
        self._install_shortcuts()
        self._start_camera_scan()     # arranca antes que _load_camera_choices para que muestre "Detectando..."
        self._load_camera_choices()   # muestra "Detectando..." mientras el scan corre en background
        self._auto_load_default_roster()
        self._refresh_course_view()
        self._sync_session_ui()

    def _build_stylesheet(self) -> str:
        return f"""
            QMainWindow {{
                background: {WINDOW_BG};
            }}
            QWidget {{
                color: {TEXT_PRIMARY};
                font-family: "Segoe UI";
                font-size: 10pt;
            }}
            QFrame#Card {{
                background: {CARD_BG};
                border: 1px solid {CARD_EDGE};
                border-radius: 14px;
            }}
            QFrame#CapturePanel {{
                background: {INFO_BG};
            }}
            QScrollArea {{
                background: {PANEL_BG};
                border: none;
            }}
            QSplitter::handle {{
                background: {CARD_EDGE};
            }}
            QListWidget {{
                background: #19181E;
                border: 1px solid {CARD_EDGE};
                border-radius: 10px;
                padding: 3px;
            }}
            QListWidget::item {{
                padding: 6px 8px;
                border-radius: 8px;
            }}
            QListWidget::item:selected {{
                background: {ACCENT_PURPLE};
                color: {TEXT_PRIMARY};
            }}
            QLabel#AppTitle {{
                font-size: 16pt;
                font-weight: 800;
                color: {TEXT_PRIMARY};
            }}
            QLabel#AppSubtitle {{
                font-size: 10pt;
                font-weight: 700;
                color: {ACCENT_GOLD};
            }}
            QLabel#CardTitle {{
                font-size: 12pt;
                font-weight: 700;
                color: {ACCENT_GOLD};
            }}
            QLabel#Muted {{
                color: {TEXT_MUTED};
            }}
            QLabel#Status {{
                background: #1C1A22;
                color: {SUCCESS};
                border-radius: 10px;
                padding: 8px 10px;
                font-weight: 700;
            }}
            QLineEdit, QComboBox, QPlainTextEdit {{
                background: #222028;
                color: {TEXT_PRIMARY};
                border: 1px solid #49445A;
                border-radius: 8px;
                padding: 6px 7px;
            }}
            QLineEdit:focus, QComboBox:focus, QPlainTextEdit:focus {{
                border: 1px solid {ACCENT_PURPLE};
            }}
            QComboBox::drop-down {{
                border: none;
                width: 28px;
            }}
            QPushButton {{
                background: {ACCENT_PURPLE};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 7px 10px;
                font-weight: 700;
            }}
            QPushButton:hover {{
                background: #A97BFF;
            }}
            QPushButton:disabled {{
                background: #6A597E;
                color: #d9d1e6;
            }}
            QPushButton#GoldButton {{
                background: {ACCENT_GOLD};
                color: {ACCENT_GOLD_DARK};
            }}
            QPushButton#GoldButton:hover {{
                background: #FFD77A;
            }}
            QPushButton#DangerButton {{
                background: {DANGER};
                color: white;
            }}
            QPushButton#DangerButton:hover {{
                background: #FF99AA;
            }}
            QRadioButton, QCheckBox {{
                spacing: 8px;
            }}
            QTableWidget {{
                background: #19181F;
                alternate-background-color: #202028;
                gridline-color: #423D54;
                color: {TEXT_PRIMARY};
                selection-background-color: {ACCENT_PURPLE};
                selection-color: white;
                border: none;
                border-radius: 8px;
            }}
            QHeaderView::section {{
                background: #24222C;
                color: {ACCENT_GOLD};
                padding: 8px;
                border: none;
                font-weight: 700;
            }}
            QTabWidget::pane {{
                border: none;
                background: transparent;
            }}
            QTabBar::tab {{
                background: #23222A;
                color: #C7C1CF;
                border: 1px solid #383545;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                padding: 10px 18px;
                font-weight: bold;
            }}
            QTabBar::tab:selected {{
                background: #6A4A98;
                color: #F2EEF7;
                border-color: #6A4A98;
            }}
            QTabBar::tab:hover {{
                background: #352C43;
                color: #F2EEF7;
            }}
        """

    def _build_ui(self) -> None:
        central = QWidget(self)
        self.setCentralWidget(central)

        outer = QHBoxLayout(central)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setChildrenCollapsible(False)
        outer.addWidget(splitter)

        self.sidebar_scroll = QScrollArea()
        self.sidebar_scroll.setWidgetResizable(True)
        self.sidebar_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.sidebar_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.sidebar_scroll.setMinimumWidth(360)

        sidebar_host = QWidget()
        self.sidebar_scroll.setWidget(sidebar_host)
        sidebar_layout = QVBoxLayout(sidebar_host)
        sidebar_layout.setContentsMargins(12, 12, 12, 12)
        sidebar_layout.setSpacing(8)

        title_card = self._make_card()
        title_layout = QVBoxLayout(title_card)
        title_layout.setContentsMargins(14, 10, 14, 10)
        title_layout.setSpacing(2)
        title = QLabel(APP_TITLE)
        title.setObjectName("AppTitle")
        title.setWordWrap(True)
        subtitle = QLabel("Captura por curso, roster y respaldo espejo")
        subtitle.setObjectName("AppSubtitle")
        subtitle.setWordWrap(True)
        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)
        title_layout.addStretch(1)
        sidebar_layout.addWidget(title_card)

        self.session_card = self._make_card()
        session_layout = QVBoxLayout(self.session_card)
        session_layout.setContentsMargins(10, 10, 10, 10)
        session_layout.setSpacing(5)
        session_layout.addWidget(self._card_title("Sesion"))
        self.test_radio = QRadioButton("Modo prueba")
        self.course_radio = QRadioButton("Modo curso")
        self.test_radio.setChecked(True)
        self.mode_group = QButtonGroup(self)
        self.mode_group.addButton(self.test_radio)
        self.mode_group.addButton(self.course_radio)
        self.test_radio.toggled.connect(self._on_mode_changed)
        session_layout.addWidget(self.test_radio)
        session_layout.addWidget(self.course_radio)

        session_layout.addWidget(self._muted_label("Curso"))
        self.course_combo = QComboBox()
        self.course_combo.setEditable(True)
        self.course_combo.currentTextChanged.connect(self._on_course_changed)
        session_layout.addWidget(self.course_combo)

        roster_btn_row = QHBoxLayout()
        self.load_roster_button = QPushButton("Cargar lista")
        self.load_roster_button.setObjectName("GoldButton")
        self.load_roster_button.clicked.connect(self.import_roster_file)
        roster_btn_row.addWidget(self.load_roster_button)
        self.open_photos_button = QPushButton("Abrir fotos")
        self.open_photos_button.setObjectName("GoldButton")
        self.open_photos_button.clicked.connect(self.open_photos_root)
        roster_btn_row.addWidget(self.open_photos_button)
        session_layout.addLayout(roster_btn_row)

        self.roster_path_label = QLabel("Lista no cargada")
        self.roster_path_label.setObjectName("Muted")
        self.roster_path_label.setWordWrap(True)
        session_layout.addWidget(self.roster_path_label)

        self.data_root_label = QLabel(f"Datos: {self.data_root}")
        self.data_root_label.setObjectName("Muted")
        self.data_root_label.setWordWrap(True)
        session_layout.addWidget(self.data_root_label)

        self.start_button = QPushButton("Iniciar sesion")
        self.start_button.clicked.connect(self.start_session)
        session_layout.addWidget(self.start_button)


        self.roster_card = self._make_card()
        roster_layout = QVBoxLayout(self.roster_card)
        roster_layout.setContentsMargins(10, 10, 10, 10)
        roster_layout.setSpacing(5)
        roster_layout.addWidget(self._card_title("Lista de alumnos"))
        self.roster_status_label = QLabel(self.roster_status_text)
        self.roster_status_label.setObjectName("Muted")
        self.roster_status_label.setWordWrap(True)
        roster_layout.addWidget(self.roster_status_label)

        nav_row = QHBoxLayout()
        self.prev_button = QPushButton("Anterior")
        self.prev_button.setObjectName("GoldButton")
        self.prev_button.clicked.connect(self.prev_roster_student)
        nav_row.addWidget(self.prev_button)
        self.next_button = QPushButton("Siguiente")
        self.next_button.setObjectName("GoldButton")
        self.next_button.clicked.connect(self.next_roster_student)
        nav_row.addWidget(self.next_button)
        roster_layout.addLayout(nav_row)

        self.align_button = QPushButton("Alinear con lista")
        self.align_button.clicked.connect(self.sync_student_with_roster)
        roster_layout.addWidget(self.align_button)


        self.camera_card = self._make_card()
        camera_layout = QVBoxLayout(self.camera_card)
        camera_layout.setContentsMargins(10, 10, 10, 10)
        camera_layout.setSpacing(5)
        camera_layout.addWidget(self._card_title("Fuentes de video"))
        camera_layout.addWidget(self._muted_label("DroidCam, OBS Virtual Camera o webcam integrada"))
        self.source_status_label = self._muted_label("Refresca la lista si abres OBS o DroidCam despues de iniciar.")
        camera_layout.addWidget(self.source_status_label)

        self.source_list = QListWidget()
        self.source_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.source_list.currentRowChanged.connect(self._on_source_selected)
        self.source_list.setMinimumHeight(92)
        camera_layout.addWidget(self.source_list)

        self.camera_combo = QComboBox()
        self.camera_combo.currentIndexChanged.connect(self._on_camera_selected)
        self.camera_combo.setVisible(False)
        camera_layout.addWidget(self.camera_combo)

        refresh_row = QHBoxLayout()
        self.refresh_sources_button = QPushButton("Refrescar fuentes")
        self.refresh_sources_button.setObjectName("GoldButton")
        self.refresh_sources_button.clicked.connect(self.refresh_sources)
        refresh_row.addWidget(self.refresh_sources_button)
        camera_layout.addLayout(refresh_row)

        self.resolution_combo = QComboBox()
        self.resolution_combo.currentIndexChanged.connect(self._on_resolution_selected)
        camera_layout.addWidget(self._muted_label("Resolucion"))
        camera_layout.addWidget(self.resolution_combo)

        self.mirror_check = QCheckBox("Espejo horizontal")
        self.face_check = QCheckBox("Ayuda visual de rostro")
        self.guide_check = QCheckBox("Mostrar guia")
        self.crop_check = QCheckBox("Recortar tipo credencial 3:4")
        self.face_check.setChecked(True)
        self.guide_check.setChecked(True)
        self.crop_check.setChecked(True)
        camera_layout.addWidget(self.mirror_check)
        camera_layout.addWidget(self.face_check)
        camera_layout.addWidget(self.guide_check)
        camera_layout.addWidget(self.crop_check)

        self.crop_tuning_label = self._muted_label("Ajuste fino: auto")
        camera_layout.addWidget(self.crop_tuning_label)
        tuning_row = QGridLayout()
        tuning_row.setHorizontalSpacing(6)
        tuning_row.setVerticalSpacing(6)
        self.crop_left_button = QPushButton("Izq")
        self.crop_left_button.clicked.connect(lambda: self._nudge_crop(-1, 0))
        self.crop_right_button = QPushButton("Der")
        self.crop_right_button.clicked.connect(lambda: self._nudge_crop(1, 0))
        self.crop_up_button = QPushButton("Arr")
        self.crop_up_button.clicked.connect(lambda: self._nudge_crop(0, -1))
        self.crop_down_button = QPushButton("Abj")
        self.crop_down_button.clicked.connect(lambda: self._nudge_crop(0, 1))
        self.crop_zoom_out_button = QPushButton("Zoom -")
        self.crop_zoom_out_button.clicked.connect(lambda: self._zoom_crop(1 / CROP_MANUAL_ZOOM_STEP))
        self.crop_zoom_in_button = QPushButton("Zoom +")
        self.crop_zoom_in_button.clicked.connect(lambda: self._zoom_crop(CROP_MANUAL_ZOOM_STEP))
        self.crop_reset_button = QPushButton("Reset")
        self.crop_reset_button.setObjectName("GoldButton")
        self.crop_reset_button.clicked.connect(self._reset_crop_tuning)
        tuning_row.addWidget(self.crop_left_button, 0, 0)
        tuning_row.addWidget(self.crop_right_button, 0, 1)
        tuning_row.addWidget(self.crop_up_button, 1, 0)
        tuning_row.addWidget(self.crop_down_button, 1, 1)
        tuning_row.addWidget(self.crop_zoom_out_button, 2, 0)
        tuning_row.addWidget(self.crop_zoom_in_button, 2, 1)
        tuning_row.addWidget(self.crop_reset_button, 3, 0, 1, 2)
        camera_layout.addLayout(tuning_row)

        row = QGridLayout()
        row.setHorizontalSpacing(10)
        row.setVerticalSpacing(8)
        row.addWidget(self._muted_label("Rotacion"), 0, 0)
        self.rotation_combo = QComboBox()
        self.rotation_combo.addItems(["0 deg", "90 deg", "180 deg", "270 deg"])
        row.addWidget(self.rotation_combo, 0, 1)
        row.addWidget(self._muted_label("Temporizador"), 1, 0)
        self.countdown_combo = QComboBox()
        self.countdown_combo.addItems(["0 s", "3 s", "5 s"])
        row.addWidget(self.countdown_combo, 1, 1)
        camera_layout.addLayout(row)

        self.prev_camera_button = QPushButton("Fuente anterior")
        self.prev_camera_button.setObjectName("GoldButton")
        self.prev_camera_button.clicked.connect(self.prev_camera)
        self.next_camera_button = QPushButton("Fuente siguiente")
        self.next_camera_button.setObjectName("GoldButton")
        self.next_camera_button.clicked.connect(self.next_camera)
        camera_layout.addWidget(self.prev_camera_button)
        camera_layout.addWidget(self.next_camera_button)


        self.capture_card = self._make_card()
        capture_layout = QVBoxLayout(self.capture_card)
        capture_layout.setContentsMargins(10, 10, 10, 10)
        capture_layout.setSpacing(5)
        self.capture_card_title = self._card_title("Captura")
        capture_layout.addWidget(self.capture_card_title)
        self.manual_student_label = self._muted_label("Nombre actual")
        capture_layout.addWidget(self.manual_student_label)
        self.student_edit = QLineEdit()
        self.student_edit.returnPressed.connect(self.capture_photo)
        capture_layout.addWidget(self.student_edit)
        btn_row = QHBoxLayout()
        self.capture_button = QPushButton("Capturar")
        self.capture_button.clicked.connect(self.capture_photo)
        btn_row.addWidget(self.capture_button)
        self.clear_button = QPushButton("Limpiar")
        self.clear_button.setObjectName("GoldButton")
        self.clear_button.clicked.connect(self.student_edit.clear)
        btn_row.addWidget(self.clear_button)
        capture_layout.addLayout(btn_row)

        btn_row2 = QHBoxLayout()
        self.retake_button = QPushButton("Rehacer ultima")
        self.retake_button.setObjectName("GoldButton")
        self.retake_button.clicked.connect(self.retake_last)
        btn_row2.addWidget(self.retake_button)
        self.close_session_button = QPushButton("Cerrar sesion")
        self.close_session_button.setObjectName("DangerButton")
        self.close_session_button.clicked.connect(self.close_session)
        btn_row2.addWidget(self.close_session_button)
        capture_layout.addLayout(btn_row2)
        # Definición de recientes e info cards antes de agruparlas en pestañas
        self.recent_card = self._make_card()
        recent_layout = QVBoxLayout(self.recent_card)
        recent_layout.setContentsMargins(10, 10, 10, 10)
        recent_layout.setSpacing(5)
        recent_layout.addWidget(self._card_title("Recientes"))
        self.recent_text = QPlainTextEdit()
        self.recent_text.setReadOnly(True)
        self.recent_text.setMaximumHeight(80)
        recent_layout.addWidget(self.recent_text)

        self.info_card = self._make_card()
        info_layout = QVBoxLayout(self.info_card)
        info_layout.setContentsMargins(10, 10, 10, 10)
        info_layout.setSpacing(5)
        info_layout.addWidget(self._card_title("Info"))
        self.info_text = QPlainTextEdit()
        self.info_text.setReadOnly(True)
        self.info_text.setMaximumHeight(80)
        info_layout.addWidget(self.info_text)

        # Crear el Tab Widget para la barra lateral
        self.sidebar_tabs = QTabWidget()
        self.sidebar_tabs.setObjectName("SidebarTabs")

        # Pestaña 1: Capturar (Controles de flujo diario)
        self.tab_capture = QWidget()
        self.tab_capture_layout = QVBoxLayout(self.tab_capture)
        self.tab_capture_layout.setContentsMargins(0, 8, 0, 8)
        self.tab_capture_layout.setSpacing(12)
        self.tab_capture_layout.addWidget(self.roster_card)
        self.tab_capture_layout.addWidget(self.capture_card)
        self.tab_capture_layout.addWidget(self.recent_card)
        self.tab_capture_layout.addStretch(1)

        # Pestaña 2: Ajustes (Configuraciones de hardware e inicio)
        self.tab_settings = QWidget()
        self.tab_settings_layout = QVBoxLayout(self.tab_settings)
        self.tab_settings_layout.setContentsMargins(0, 8, 0, 8)
        self.tab_settings_layout.setSpacing(12)
        self.tab_settings_layout.addWidget(self.session_card)
        self.tab_settings_layout.addWidget(self.camera_card)
        self.tab_settings_layout.addWidget(self.info_card)
        self.tab_settings_layout.addStretch(1)

        self.sidebar_tabs.addTab(self.tab_capture, "Capturar")
        self.sidebar_tabs.addTab(self.tab_settings, "Configuración")

        # Añadir el Tab Widget a la barra lateral principal
        sidebar_layout.addWidget(self.sidebar_tabs)

        self.main_splitter = QSplitter(Qt.Orientation.Vertical)
        self.main_splitter.setChildrenCollapsible(False)

        self.capture_tab = QWidget()
        self.course_tab = QWidget()
        self._build_capture_tab()
        self._build_course_tab()
        self.main_splitter.addWidget(self.capture_tab)
        self.main_splitter.addWidget(self.course_tab)
        self.main_splitter.setStretchFactor(0, 3)
        self.main_splitter.setStretchFactor(1, 2)
        self.main_splitter.setSizes([800, 340])

        self.root_splitter = splitter
        splitter.addWidget(self.sidebar_scroll)
        splitter.addWidget(self.main_splitter)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        splitter.setChildrenCollapsible(False)
        self._apply_responsive_layout()

        self.status_label = QLabel("Listo para iniciar. Selecciona camara, carga lista y abre sesion.")
        self.status_label.setObjectName("Status")
        self.statusBar().addPermanentWidget(self.status_label, 1)

    def _build_capture_tab(self) -> None:
        layout = QVBoxLayout(self.capture_tab)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)

        self.session_label = QLabel("Sesion no iniciada")
        self.session_label.setStyleSheet("font-size: 13pt; font-weight: 700; color: #E8E4EF;")
        layout.addWidget(self.session_label)

        self.preview_frame = QFrame()
        self.preview_frame.setObjectName("CapturePanel")
        self.preview_frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.preview_frame.setMinimumHeight(500)
        preview_layout = QVBoxLayout(self.preview_frame)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        preview_layout.setSpacing(0)
        self.preview_label = QLabel("Sin señal de camara")
        self.preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Ignored)
        self.preview_label.setStyleSheet("background: #111016; color: #CFC9D8; border-radius: 10px;")
        self.preview_label.setMinimumSize(0, 0)
        preview_layout.addWidget(self.preview_label)
        layout.addWidget(self.preview_frame, 1)

        self.post_progress_bar = QProgressBar()
        self.post_progress_bar.setRange(0, 100)
        self.post_progress_bar.setValue(0)
        self.post_progress_bar.setTextVisible(True)
        self.post_progress_bar.setFormat("Guardando y recortando foto: %p%")
        self.post_progress_bar.setVisible(False)
        self.post_progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #4F2B74;
                border-radius: 6px;
                background-color: #1A1822;
                text-align: center;
                color: #ffffff;
                font-weight: bold;
                height: 22px;
            }
            QProgressBar::chunk {
                background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #9D4EDD, stop:1 #7B2CBF);
                border-radius: 5px;
            }
        """)
        layout.addWidget(self.post_progress_bar)

        bottom_row = QHBoxLayout()
        self.capture_button_tab = QPushButton("Capturar")
        self.capture_button_tab.clicked.connect(self.capture_photo)
        bottom_row.addWidget(self.capture_button_tab)
        self.retake_button_tab = QPushButton("Rehacer ultima")
        self.retake_button_tab.setObjectName("GoldButton")
        self.retake_button_tab.clicked.connect(self.retake_last)
        bottom_row.addWidget(self.retake_button_tab)
        self.refresh_course_button = QPushButton("Actualizar curso")
        self.refresh_course_button.setObjectName("GoldButton")
        self.refresh_course_button.clicked.connect(self._refresh_course_view)
        bottom_row.addWidget(self.refresh_course_button)
        self.open_folder_tab_button = QPushButton("Abrir fotos")
        self.open_folder_tab_button.setObjectName("GoldButton")
        self.open_folder_tab_button.clicked.connect(self.open_photos_root)
        bottom_row.addWidget(self.open_folder_tab_button)
        layout.addLayout(bottom_row)

    def _build_course_tab(self) -> None:
        layout = QVBoxLayout(self.course_tab)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)
        self.course_header_label = QLabel("Lista completa del curso")
        self.course_header_label.setStyleSheet("font-size: 18pt; font-weight: 800; color: #F0C85C;")
        layout.addWidget(self.course_header_label)
        self.course_summary_label = QLabel("Sin curso activo")
        self.course_summary_label.setStyleSheet("font-size: 10pt; font-weight: 700;")
        layout.addWidget(self.course_summary_label)
        self.course_progress_bar = QProgressBar()
        self.course_progress_bar.setRange(0, 100)
        layout.addWidget(self.course_progress_bar)
        self.course_count_label = QLabel("0 capturados de 0")
        self.course_count_label.setObjectName("Muted")
        layout.addWidget(self.course_count_label)

        self.course_table = QTableWidget(0, 3)
        self.course_table.setHorizontalHeaderLabels(["Estado", "Alumno", "RUT"])
        self.course_table.setAlternatingRowColors(True)
        self.course_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.course_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.course_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.course_table.cellClicked.connect(self._on_course_row_clicked)
        self.course_table.horizontalHeader().setStretchLastSection(False)
        self.course_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.course_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.course_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.course_table, 1)

        footer = QHBoxLayout()
        self.course_refresh_button = QPushButton("Actualizar lista")
        self.course_refresh_button.setObjectName("GoldButton")
        self.course_refresh_button.clicked.connect(self._refresh_course_view)
        footer.addWidget(self.course_refresh_button)
        self.course_align_button = QPushButton("Alinear con lista")
        self.course_align_button.clicked.connect(self.sync_student_with_roster)
        footer.addWidget(self.course_align_button)
        footer.addStretch(1)
        layout.addLayout(footer)

    def _make_card(self) -> QFrame:
        frame = QFrame()
        frame.setObjectName("Card")
        frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Maximum)
        return frame

    def _card_title(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setObjectName("CardTitle")
        return label

    def _muted_label(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setObjectName("Muted")
        return label

    def _install_shortcuts(self) -> None:
        shortcuts = {
            "Return": self.capture_photo,
            "Enter": self.capture_photo,
            "Ctrl+R": self.retake_last,
            "Ctrl+N": self.next_roster_student,
            "Ctrl+P": self.prev_roster_student,
            "Ctrl+L": self.import_roster_file,
            "Ctrl+O": self.open_photos_root,
            "Left": lambda: self._nudge_crop(-1, 0),
            "Right": lambda: self._nudge_crop(1, 0),
            "Up": lambda: self._nudge_crop(0, -1),
            "Down": lambda: self._nudge_crop(0, 1),
            "Ctrl+=": lambda: self._zoom_crop(CROP_MANUAL_ZOOM_STEP),
            "Ctrl+-": lambda: self._zoom_crop(1 / CROP_MANUAL_ZOOM_STEP),
            "Ctrl+0": self._reset_crop_tuning,
        }
        for sequence, callback in shortcuts.items():
            shortcut = QShortcut(QKeySequence(sequence), self)
            shortcut.activated.connect(callback)

    def _update_crop_tuning_label(self) -> None:
        self.crop_tuning_label.setText(
            f"Ajuste fino: x={self.crop_manual_dx:+.2f} y={self.crop_manual_dy:+.2f} zoom={self.crop_manual_zoom:.2f}"
        )

    def _apply_responsive_layout(self) -> None:
        if not hasattr(self, "root_splitter") or not hasattr(self, "main_splitter"):
            return

        window_width = max(1280, self.width())
        window_height = max(820, self.height())

        sidebar_width = int(window_width * 0.27)
        sidebar_width = max(360, min(480, sidebar_width))
        main_width = max(680, window_width - sidebar_width - 24)
        self.root_splitter.setSizes([sidebar_width, main_width])

        preview_height = int(window_height * 0.64)
        preview_height = max(520, min(840, preview_height))
        course_height = max(280, window_height - preview_height - 210)
        self.main_splitter.setSizes([preview_height, course_height])

    def _reset_crop_tuning(self) -> None:
        self.crop_manual_dx = 0.0
        self.crop_manual_dy = 0.0
        self.crop_manual_zoom = 1.0
        self._update_crop_tuning_label()

    def _nudge_crop(self, dx_steps: int, dy_steps: int) -> None:
        self.crop_manual_dx = max(-0.35, min(0.35, self.crop_manual_dx + (dx_steps * CROP_MANUAL_STEP)))
        self.crop_manual_dy = max(-0.35, min(0.35, self.crop_manual_dy + (dy_steps * CROP_MANUAL_STEP)))
        self._update_crop_tuning_label()

    def _zoom_crop(self, factor: float) -> None:
        self.crop_manual_zoom = max(0.82, min(1.35, self.crop_manual_zoom * factor))
        self._update_crop_tuning_label()

    def _load_camera_choices(self) -> None:
        self.camera_combo.blockSignals(True)
        self.source_list.blockSignals(True)
        self.camera_combo.clear()
        self.source_list.clear()
        self.camera_choices: list[tuple[int, str, int, str, str]] = self.available_cameras
        for index, label, backend_id, backend_name, alias in self.camera_choices:
            source_kind = camera_source_kind(alias, backend_name)
            text = f"{source_kind} | {label} | {backend_name}"
            self.camera_combo.addItem(text, (index, backend_id, backend_name, alias))
            self.source_list.addItem(QListWidgetItem(text))
        self.camera_combo.blockSignals(False)
        self.source_list.blockSignals(False)
        if not self.camera_choices:
            if self._camera_scan_running:
                self.camera_combo.addItem("Detectando camaras...", None)
                self.source_list.addItem(QListWidgetItem("Detectando fuentes..."))
            else:
                self.camera_combo.addItem("No se detectaron camaras", None)
                self.source_list.addItem(QListWidgetItem("No se detectaron fuentes"))
        if self.camera_choices:
            self.logger.info("Detected cameras: %s", len(self.camera_choices))
            self.source_status_label.setText(f"{len(self.camera_choices)} fuentes detectadas. Selecciona una y refresca si cambias el cliente.")
        elif self._camera_scan_running:
            self.source_status_label.setText("Detectando fuentes de video, por favor espera...")
        else:
            self.logger.warning("No cameras detected at startup.")
            self.source_status_label.setText("No se detectaron fuentes de video.")

        self.resolution_combo.blockSignals(True)
        self.resolution_combo.clear()
        self.resolution_combo.addItem(RESOLUTION_AUTO_LABEL, None)
        for width, height in COMMON_CAMERA_RESOLUTIONS:
            self.resolution_combo.addItem(f"{width} x {height}", (width, height))
        self.resolution_combo.setCurrentIndex(0)
        self.resolution_combo.blockSignals(False)

    def _resolution_label(self, resolution: Optional[tuple[int, int]]) -> str:
        if resolution is None:
            return RESOLUTION_AUTO_LABEL
        width, height = resolution
        return f"{width} x {height}"

    def _resolution_from_combo_index(self, combo_index: int) -> Optional[tuple[int, int]]:
        if combo_index < 0:
            return None
        data = self.resolution_combo.itemData(combo_index)
        if isinstance(data, tuple) and len(data) == 2:
            width, height = data
            if int(width) > 0 and int(height) > 0:
                return int(width), int(height)
        return None

    def _current_selected_resolution(self) -> Optional[tuple[int, int]]:
        return self._resolution_from_combo_index(self.resolution_combo.currentIndex())

    def _select_default_camera(self) -> None:
        if not self.camera_choices:
            return
        remembered_index, remembered_backend = load_last_camera(APP_ROOT)
        selected = 0
        if remembered_index is not None:
            for i, (index, _label, backend_id, _backend_name, _alias) in enumerate(self.camera_choices):
                if index != remembered_index:
                    continue
                if remembered_backend:
                    expected = remembered_backend.lower()
                    if expected != "any" and backend_key_from_id(backend_id) != expected:
                        continue
                selected = i
                break
        self.camera_combo.blockSignals(True)
        self.source_list.blockSignals(True)
        self.camera_combo.setCurrentIndex(selected)
        self.source_list.setCurrentRow(selected)
        self.camera_combo.blockSignals(False)
        self.source_list.blockSignals(False)
        self._apply_selected_camera(selected)
        self._start_camera_thread()

    def _sync_resolution_combo(self, block_signals: bool = False) -> None:
        if not self.camera_choices:
            return
        if block_signals:
            self.resolution_combo.blockSignals(True)
        try:
            preferred = load_camera_resolution(APP_ROOT, self.camera_index, backend_key_from_id(self.backend_id))
            target_index = 0
            if preferred is not None:
                for i in range(self.resolution_combo.count()):
                    if self._resolution_from_combo_index(i) == preferred:
                        target_index = i
                        break
            self.camera_resolution = preferred
            self.resolution_combo.setCurrentIndex(target_index)
            self.status_label.setText(
                f"Camara lista: {self.camera_alias} | {self.backend_name} | {self._resolution_label(preferred)}"
            )
        finally:
            if block_signals:
                self.resolution_combo.blockSignals(False)

    def _apply_selected_camera(self, combo_index: int) -> None:
        if combo_index < 0 or combo_index >= len(self.camera_choices):
            return
        index, _label, backend_id, backend_name, alias = self.camera_choices[combo_index]
        self.camera_index = index
        self.backend_id = backend_id
        self.backend_name = backend_name
        self.camera_alias = alias
        self.preview_face_box = None
        self.preview_eye_centers = []
        self.preview_stable_crop_box = None
        self.preview_detection_tick = 0
        self._auto_camera_failures.discard((index, backend_id))
        save_last_camera(APP_ROOT, index, backend_key_from_id(backend_id))
        self.logger.info("Camera selected index=%s backend=%s alias=%s", index, backend_name, alias)
        self.source_status_label.setText(
            f"Fuente activa: {camera_source_kind(alias, backend_name)} | {alias or f'Camara {index}'} | {backend_name}"
        )
        self.source_list.blockSignals(True)
        self.source_list.setCurrentRow(combo_index)
        self.source_list.blockSignals(False)
        self._sync_resolution_combo(block_signals=True)

    def _start_camera_thread(self) -> None:
        if not self.camera_choices:
            self._show_preview_message("No hay camara disponible.")
            return
        if not self._stop_camera_thread():
            self.logger.warning("Camera thread still shutting down; skip restart.")
            self.status_label.setText("La camara sigue cerrando. Intenta de nuevo en unos segundos.")
            return
        self._preview_frame_received = False
        self.preview_face_box = None
        self.preview_eye_centers = []
        self.preview_stable_crop_box = None
        self.preview_detection_tick = 0
        selected_resolution = self._current_selected_resolution()
        self.camera_resolution = selected_resolution
        save_camera_resolution(
            APP_ROOT,
            self.camera_index,
            backend_key_from_id(self.backend_id),
            selected_resolution,
        )
        self.camera_thread = CameraThread(self.camera_index, self.backend_id, selected_resolution, self)
        self.camera_thread.camera_message.connect(self._on_camera_message)
        self.camera_thread.camera_error.connect(self._on_camera_error)
        self.camera_thread.start()
        self.logger.info(
            "Camera thread started index=%s backend=%s resolution=%s",
            self.camera_index,
            self.backend_name,
            selected_resolution,
        )
        resolution_label = self._resolution_label(selected_resolution)
        self.status_label.setText(f"Esperando primer frame de camara... | {resolution_label}")
        self._show_preview_message("Esperando primer frame...")
        self.first_frame_timer.start(6000)
        # The capture thread writes frames independently; start the GUI poller now.
        # Previously it only started after opening a session, leaving a live camera blank.
        self._schedule_preview_render(immediate=True)

    def _stop_camera_thread(self) -> bool:
        if self.camera_thread is None:
            return True
        try:
            self.camera_thread.stop()
            if not self.camera_thread.wait(5000):
                self.logger.warning("Camera thread did not stop in time. index=%s backend=%s", self.camera_index, self.backend_name)
                return False
        except Exception:
            self.logger.exception("Error while stopping camera thread.")
            return False
        finally:
            if self.camera_thread is not None and not self.camera_thread.isRunning():
                self.camera_thread = None
                self.first_frame_timer.stop()
        self.camera_thread = None
        self.first_frame_timer.stop()
        return True

    def _on_camera_message(self, message: str) -> None:
        self.logger.info("Camera: %s", message)
        if message.startswith("Frame recibido"):
            return
        self.status_label.setText(message)

    def _on_camera_error(self, message: str) -> None:
        self.logger.error("Camera error: %s", message)
        self.status_label.setText(message)
        self._show_preview_message(message)
        self._try_next_camera_source("error de camara")

    def _show_preview_message(self, message: str) -> None:
        self.preview_label.clear()
        self.preview_label.setText(message)
        self.preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

    def _pull_latest_preview_frame(self) -> Optional[np.ndarray]:
        if self.camera_thread is None:
            return self.latest_frame
        frame = self.camera_thread.latest_frame()
        if frame is not None:
            self.latest_frame = frame
        return self.latest_frame

    def _on_mode_changed(self) -> None:
        self.mode = "course" if self.course_radio.isChecked() else "test"
        self.logger.info("Mode changed to %s", self.mode)
        self._sync_session_ui()
        self._refresh_course_view()

    def _on_course_changed(self, text: str) -> None:
        self.course_text = text.strip()
        if self.session and self.session.course_display != self.course_text:
            self._sync_session_ui()
        self._refresh_course_view()

    def _on_camera_selected(self, *_args) -> None:
        if not self.camera_choices:
            return
        combo_index = self.camera_combo.currentIndex()
        self._apply_selected_camera(combo_index)
        self._start_camera_thread()

    def _on_source_selected(self, row: int) -> None:
        if not self.camera_choices:
            return
        if row < 0 or row >= len(self.camera_choices):
            return
        if self.camera_combo.currentIndex() != row:
            self.camera_combo.setCurrentIndex(row)

    def _start_camera_scan(self) -> None:
        """Detecta camaras disponibles en un hilo de fondo para no bloquear la UI."""
        if self._camera_scan_running:
            return
        self._camera_scan_running = True
        aliases = self.aliases
        self.logger.info("Camera scan started in background thread.")

        def worker() -> None:
            try:
                cameras = list_available_cameras(aliases)
            except Exception as exc:
                self.logger.exception("Camera scan failed: %s", exc)
                cameras = []
            self.camera_scan_ready.emit(cameras)

        t = threading.Thread(target=worker, name="camera-scan", daemon=True)
        t.start()

    def _on_camera_scan_ready(self, cameras: list) -> None:
        """Llamado en el hilo principal cuando termina la deteccion de camaras."""
        self._camera_scan_running = False
        self.available_cameras = cameras
        self.logger.info("Camera scan finished. cameras=%s", len(cameras))
        self._load_camera_choices()
        self._select_default_camera()

    def refresh_sources(self) -> None:
        self.aliases = load_camera_aliases(APP_ROOT)
        self._auto_camera_failures.clear()
        self.available_cameras = []
        self._load_camera_choices()
        self.status_label.setText("Buscando fuentes de video...")
        self._start_camera_scan()

    def _try_next_camera_source(self, reason: str) -> bool:
        if len(self.camera_choices) <= 1:
            return False

        current_row = self.camera_combo.currentIndex()
        if current_row < 0 or current_row >= len(self.camera_choices):
            current_row = 0

        current_key = (self.camera_index, self.backend_id)
        self._auto_camera_failures.add(current_key)

        for offset in range(1, len(self.camera_choices) + 1):
            next_row = (current_row + offset) % len(self.camera_choices)
            index, _label, backend_id, backend_name, alias = self.camera_choices[next_row]
            if (index, backend_id) in self._auto_camera_failures:
                continue

            self.logger.warning(
                "Switching camera automatically after %s. from=%s/%s to=%s/%s alias=%s",
                reason,
                self.camera_index,
                self.backend_name,
                index,
                backend_name,
                alias,
            )
            self.status_label.setText(f"Cambiando fuente por {reason}: {alias} | {backend_name}")
            self.camera_combo.blockSignals(True)
            self.source_list.blockSignals(True)
            self.camera_combo.setCurrentIndex(next_row)
            self.source_list.setCurrentRow(next_row)
            self.camera_combo.blockSignals(False)
            self.source_list.blockSignals(False)
            self._apply_selected_camera(next_row)
            QTimer.singleShot(250, self._start_camera_thread)
            return True

        self.logger.error("All camera sources failed after %s.", reason)
        self.status_label.setText("No hay otra fuente de camara usable. Refresca fuentes o revisa el cable/app.")
        return False

    def _on_resolution_selected(self, *_args) -> None:
        if not self.camera_choices:
            return
        if self.resolution_combo.currentIndex() < 0:
            return
        self.camera_resolution = self._current_selected_resolution()
        save_camera_resolution(
            APP_ROOT,
            self.camera_index,
            backend_key_from_id(self.backend_id),
            self.camera_resolution,
        )
        self.status_label.setText(
            f"Resolucion seleccionada: {self._resolution_label(self.camera_resolution)}"
        )
        if self.camera_thread is not None:
            self._start_camera_thread()

    def _active_course_text(self) -> str:
        if self.session is not None:
            return self.session.course_display
        return self.course_combo.currentText().strip() or "SIN CURSO"

    def _active_students(self) -> list[RosterStudent]:
        course = self._active_course_text()
        return self.roster_map.get(course, [])

    def _active_index(self) -> int:
        if self.session and self.session.has_roster:
            return self.session.roster_index
        course = self._active_course_text()
        return self.preview_index_by_course.get(course, 0)

    def _set_active_index(self, index: int) -> None:
        if self.session and self.session.has_roster:
            self.session.roster_index = index
            return
        course = self._active_course_text()
        self.preview_index_by_course[course] = index

    def _current_student(self) -> Optional[RosterStudent]:
        students = self._active_students()
        if not students:
            return None
        idx = self._active_index()
        if idx < 0 or idx >= len(students):
            return None
        return students[idx]

    def _sync_course_table_selection(self, row_index: int) -> None:
        if self.course_table is None:
            return
        students = self._active_students()
        with QSignalBlocker(self.course_table):
            self.course_table.clearSelection()
            if row_index < 0 or row_index >= len(students):
                return
            self.course_table.selectRow(row_index)
            item = self.course_table.item(row_index, 1)
            if item is not None:
                self.course_table.scrollToItem(item)

    def _advance_roster_past_completed(self) -> None:
        if self.session is None or not self.session.has_roster:
            return
        completed_keys = self._session_completed_keys()
        while True:
            current = self.session.current_student()
            if current is None:
                return
            if current.key not in completed_keys:
                return
            if self.session.roster_index >= self.session.roster_total:
                return
            self.session.advance()

    def _on_course_row_clicked(self, row: int, _column: int) -> None:
        students = self._active_students()
        if row < 0 or row >= len(students):
            return
        self._set_active_index(row)
        current = students[row]
        if self.session and self.session.has_roster:
            self.student_edit.setText(current.display_name)
        self.status_label.setText(f"Seleccionado: {current.display_name}")
        self._sync_session_ui()
        self._refresh_course_view(force=True)
        self._render_preview()

    def _update_roster_summary(self) -> None:
        students = self._active_students()
        course = self._active_course_text()
        idx = self._active_index()
        
        is_test = (self.session.mode == "test") if self.session else (self.mode == "test")
        if is_test:
            self.roster_status_label.setText("Modo Prueba activo.\nIngresa el nombre del alumno abajo y captura libremente.")
            return

        if not students:
            if self.roster_map:
                self.roster_status_label.setText(f"{course}\nSin alumnos cargados para este curso.")
            else:
                self.roster_status_label.setText("Carga un Excel o CSV para operar por lista.")
            return

        current = self._current_student()
        remaining = max(0, len(students) - min(idx, len(students)))
        if current is None:
            current_text = "Lista completa"
            rut_text = "-"
        else:
            current_text = current.display_name
            rut_text = current.rut or "-"

        self.roster_status_label.setText(
            f"{course}\n"
            f"Actual: {current_text}\n"
            f"RUT: {rut_text}\n"
            f"Pendientes: {remaining}"
        )

    def _session_completed_keys(self) -> set[tuple[str, str]]:
        keys: set[tuple[str, str]] = set()
        if not self.session:
            return keys
        for record in self.session.records:
            keys.add(_student_key(record.student_name, record.rut))
        return keys

    def _reconcile_session_roster_progress(self) -> None:
        if self.session is None or not self.session.has_roster:
            return
        completed_keys = self._session_completed_keys()
        next_pending_index = len(self.session.roster_students)
        for index, student in enumerate(self.session.roster_students):
            if student.key not in completed_keys:
                next_pending_index = index
                break
        if self.session.roster_index != next_pending_index:
            self.session.roster_index = next_pending_index

    def _refresh_course_view(self, force: bool = False) -> None:
        students = self._active_students()
        course = self._active_course_text()
        current_index = self._active_index()
        completed_keys = self._session_completed_keys()
        if not completed_keys and not self.session and students:
            # Pre-session preview: no students are completed yet.
            completed_keys = set()

        is_test = (self.session.mode == "test") if self.session else (self.mode == "test")
        if is_test:
            self.course_summary_label.setText("Modo Prueba activo")
            self.course_progress_bar.setValue(0)
            self.course_count_label.setText(f"{len(self.session.records) if self.session else 0} fotos de prueba capturadas")
            self.course_table.setRowCount(0)
            return

        self.course_summary_label.setText(
            f"{course} | {len(completed_keys)} capturados | {max(0, len(students) - len(completed_keys))} pendientes"
            if students
            else f"{course} | sin roster cargado"
        )
        total = len(students)
        done = min(len(completed_keys), total)
        self.course_progress_bar.setValue(int((done / total) * 100) if total else 0)
        self.course_count_label.setText(f"{done} capturados de {total}")

        self.course_table.setUpdatesEnabled(False)
        try:
            self.course_table.setRowCount(0)
            for row_index, student in enumerate(students):
                if row_index == current_index and current_index < len(students):
                    state = "Actual"
                    background = CURRENT_BG
                elif student.key in completed_keys:
                    state = "Hecho"
                    background = DONE_BG
                else:
                    state = "Pendiente"
                    background = PENDING_BG

                row = self.course_table.rowCount()
                self.course_table.insertRow(row)
                for column, value in enumerate((state, student.display_name, student.rut or "-")):
                    item = QTableWidgetItem(value)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    item.setBackground(Qt.GlobalColor.transparent)
                    self.course_table.setItem(row, column, item)
                    item.setForeground(Qt.GlobalColor.white)
                    if column == 0:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                for column in range(3):
                    item = self.course_table.item(row, column)
                    if item is not None:
                        item.setBackground(Qt.GlobalColor.black)
                        # Apply custom colors through the row. Backgrounds are repeated intentionally.
                        if state == "Hecho":
                            item.setBackground(Qt.GlobalColor.darkGreen)
                        elif state == "Actual":
                            item.setBackground(Qt.GlobalColor.darkYellow)
                        else:
                            item.setBackground(Qt.GlobalColor.darkMagenta)
        finally:
            self.course_table.setUpdatesEnabled(True)

        self._sync_course_table_selection(current_index)
        self._update_roster_summary()
        self._update_info_tab()

    def _update_info_tab(self) -> None:
        if self.info_text is None:
            return
        lines = [
            f"Titulo: {APP_TITLE}",
            f"Log actual: {self.log_path}",
            f"Tiempo: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Modo: {self.mode}",
            f"Curso: {self._active_course_text()}",
            f"Camara: {self.camera_alias or '-'} | {self.backend_name}",
            f"Roster: {self.roster_source_path if self.roster_source_path else 'No cargado'}",
        ]
        if self.session is not None:
            lines.extend(
                [
                    "",
                    f"Sesion iniciada: {self.session.started_at.strftime('%Y-%m-%d %H:%M:%S')}",
                    f"Carpeta fotos: {self.session.session_dir}",
                    f"Carpeta respaldo: {self.session.backup_dir}",
                    f"CSV: {self.session.csv_path}",
                    f"Fotos guardadas: {len(self.session.records)}",
                ]
            )
        self.info_text.setPlainText("\n".join(lines))

    def _sync_session_ui(self) -> None:
        roster_available = bool(self._active_students())
        roster_mode = self.mode == "course" and roster_available
        self.capture_card_title.setText("Captura por lista" if roster_mode else "Captura")
        self.manual_student_label.setVisible(not roster_mode)
        self.student_edit.setVisible(not roster_mode)
        self.clear_button.setVisible(not roster_mode)
        current = self._current_student()
        if roster_mode and current is not None:
            self.student_edit.setText(current.display_name)
        self._update_session_labels()
        self._update_roster_summary()
        self._update_info_tab()

    def _update_session_labels(self) -> None:
        if self.session is None:
            self.session_label.setText("Sesion no iniciada")
            self.session_label.setStyleSheet("font-size: 13pt; font-weight: 700; color: #E8E4EF;")
            self.test_radio.setEnabled(True)
            self.course_radio.setEnabled(True)
            self.course_combo.setEnabled(True)
            self.load_roster_button.setEnabled(True)
            self.start_button.setEnabled(True)
            self.capture_button.setEnabled(False)
            self.capture_button_tab.setEnabled(False)
            self.retake_button.setEnabled(False)
            self.retake_button_tab.setEnabled(False)
            self.close_session_button.setEnabled(False)
            self.prev_button.setEnabled(False)
            self.next_button.setEnabled(False)
            self.align_button.setEnabled(False)
            self.course_align_button.setEnabled(False)
            self.session = None
            return

        current = self.session.current_student()
        current_name = current.display_name if current else "Captura libre"
        current_rut = current.rut if current and current.rut else "-"
        pending = self.session.roster_remaining if self.session.has_roster else 0
        self.session_label.setText(
            f"{self.session.course_display} | {len(self.session.records)} capturados" + (f" | {pending} pendientes" if self.session.has_roster else "")
        )
        self.test_radio.setEnabled(False)
        self.course_radio.setEnabled(False)
        self.course_combo.setEnabled(False)
        self.load_roster_button.setEnabled(False)
        self.start_button.setEnabled(False)
        self.capture_button.setEnabled(True)
        self.capture_button_tab.setEnabled(True)
        self.retake_button.setEnabled(True)
        self.retake_button_tab.setEnabled(True)
        self.close_session_button.setEnabled(True)
        
        has_roster = self.session.has_roster
        self.prev_button.setEnabled(has_roster)
        self.next_button.setEnabled(has_roster)
        self.align_button.setEnabled(has_roster)
        self.course_align_button.setEnabled(has_roster)
        
        if self.session.mode == "test":
            self.status_label.setText(f"Modo Prueba | {len(self.session.records)} capturadas")
        else:
            self.status_label.setText(f"Actual: {current_name} | RUT: {current_rut}")

    def _load_roster_map(self, path: Path) -> dict[str, list[RosterStudent]]:
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return self._load_roster_from_workbook(path)
        if suffix in {".csv", ".tsv", ".txt"}:
            return self._load_roster_from_csv(path)
        raise ValueError("Formato no soportado. Usa XLSX o CSV.")

    def _load_roster_from_workbook(self, path: Path) -> dict[str, list[RosterStudent]]:
        workbook = load_workbook(path, data_only=True, read_only=True)
        result: dict[str, list[RosterStudent]] = {}
        for sheet in workbook.worksheets:
            students = self._parse_table_rows(list(sheet.iter_rows(values_only=True)), sheet.title)
            if students:
                result[sheet.title.strip()] = students
        return result

    def _load_roster_from_csv(self, path: Path) -> dict[str, list[RosterStudent]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|") if sample.strip() else csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
            rows = list(reader)

        grouped: dict[str, list[RosterStudent]] = {}
        for row in rows:
            course = str(row.get("course") or row.get("curso") or path.stem).strip() or path.stem
            student = self._student_from_mapping(row)
            if student is None:
                continue
            grouped.setdefault(course, []).append(student)

        for course in grouped:
            grouped[course].sort(key=_sort_key)
        return grouped

    def _parse_table_rows(self, rows: list[tuple[object, ...]], default_course: str) -> list[RosterStudent]:
        header_index = None
        header_map: dict[str, int] = {}
        for idx, row in enumerate(rows[:12]):
            candidate = self._header_map_from_sequence(row)
            if candidate:
                header_index = idx
                header_map = candidate
                break

        if header_index is None:
            return []

        students: list[RosterStudent] = []
        for row in rows[header_index + 1 :]:
            student = self._student_from_sequence(row, header_map)
            if student is not None:
                students.append(student)

        students.sort(key=_sort_key)
        return students

    def _header_map_from_sequence(self, row: tuple[object, ...]) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for index, value in enumerate(row):
            text = _normalize_text(value)
            if not text:
                continue
            if "rut" in text and "curso" not in text:
                header_map["rut"] = index
            elif "apellido paterno" in text or text in {"apellidopaterno", "primerapellido"}:
                header_map["apellido_paterno"] = index
            elif "apellido materno" in text or text in {"apellidomaterno", "segundoapellido"}:
                header_map["apellido_materno"] = index
            elif "segundo nombre" in text or text in {"segundonombre"}:
                header_map["segundo_nombre"] = index
            elif "primer nombre" in text or text in {"primernombre"}:
                header_map["primer_nombre"] = index
            elif text in {"nombre", "nombres", "alumno", "estudiante"} or "nombre completo" in text:
                header_map.setdefault("full_name", index)
            elif "curso" in text:
                header_map.setdefault("course", index)

        if "rut" not in header_map and "full_name" not in header_map and "primer_nombre" not in header_map:
            return {}
        return header_map

    def _student_from_sequence(self, row: tuple[object, ...], header_map: dict[str, int]) -> Optional[RosterStudent]:
        def cell(key: str) -> str:
            index = header_map.get(key)
            if index is None or index >= len(row):
                return ""
            return str(row[index]).strip() if row[index] is not None else ""

        student = RosterStudent(
            rut=cell("rut"),
            apellido_paterno=cell("apellido_paterno"),
            apellido_materno=cell("apellido_materno"),
            segundo_nombre=cell("segundo_nombre"),
            primer_nombre=cell("primer_nombre"),
            full_name=cell("full_name"),
        )
        if student.display_name or student.rut:
            return student
        return None

    def _student_from_mapping(self, row: dict[str, object]) -> Optional[RosterStudent]:
        normalized = {_normalize_text(key): value for key, value in row.items()}

        def pick(*names: str) -> str:
            for name in names:
                value = normalized.get(_normalize_text(name))
                if value is not None and str(value).strip():
                    return str(value).strip()
            return ""

        student = RosterStudent(
            rut=pick("rut"),
            apellido_paterno=pick("apellido paterno", "apellido_paterno", "primer apellido"),
            apellido_materno=pick("apellido materno", "apellido_materno", "segundo apellido"),
            segundo_nombre=pick("segundo nombre", "segundo_nombre"),
            primer_nombre=pick("primer nombre", "primer_nombre"),
            full_name=pick("nombre completo", "nombres", "nombre", "alumno", "estudiante"),
        )
        if student.display_name or student.rut:
            return student
        return None

    def import_roster_file(self) -> None:
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar lista de alumnos",
            str(self._default_roster_dialog_dir()),
            "Listas (*.xlsx *.xlsm *.csv *.tsv *.txt)",
        )
        if not file_name:
            return

        path = Path(file_name)
        self._load_roster_from_path(path, persist=True, show_errors=True)

    def _default_roster_dialog_dir(self) -> Path:
        if self.roster_source_path and self.roster_source_path.parent.exists():
            return self.roster_source_path.parent
        if self.default_roster_path and self.default_roster_path.parent.exists():
            return self.default_roster_path.parent
        if DEFAULT_ROSTER_DIR.exists():
            return DEFAULT_ROSTER_DIR
        return Path.home()

    def _load_roster_from_path(self, path: Path, persist: bool, show_errors: bool) -> bool:
        self.logger.info("Import roster requested: %s", path)
        try:
            roster_map = self._load_roster_map(path)
        except Exception as exc:
            self.logger.exception("Failed to load roster: %s", exc)
            if show_errors:
                QMessageBox.critical(self, "Error al cargar lista", f"No se pudo leer la lista:\n{exc}")
            return False

        if not roster_map:
            if show_errors:
                QMessageBox.warning(self, "Lista vacia", "No se encontraron alumnos en el archivo seleccionado.")
            return False

        self._apply_roster(path, roster_map, persist=persist)
        return True

    def _apply_roster(self, path: Path, roster_map: dict[str, list[RosterStudent]], persist: bool) -> None:
        self.roster_map = roster_map
        self.roster_order = list(roster_map.keys())
        self.roster_source_path = path
        self.roster_status_text = f"Lista cargada: {path.name} ({len(roster_map)} cursos)"
        self.roster_path_label.setText(self.roster_status_text)
        self.logger.info("Roster loaded: %s courses=%s", path.name, len(roster_map))
        if persist:
            self.default_roster_path = path
            self.local_config["default_roster"] = str(path)
            self.local_config["data_root"] = str(self.data_root)
            try:
                _write_local_config(self.local_config)
            except Exception as exc:
                self.logger.warning("No se pudo guardar la configuracion local: %s", exc)

        existing_course = self.course_combo.currentText().strip()
        self.course_combo.blockSignals(True)
        self.course_combo.clear()
        self.course_combo.addItems(self.roster_order)
        if existing_course and existing_course in roster_map:
            self.course_combo.setCurrentText(existing_course)
        elif self.roster_order:
            self.course_combo.setCurrentIndex(0)
        self.course_combo.blockSignals(False)

        self.preview_index_by_course.setdefault(self.course_combo.currentText().strip(), 0)
        self._sync_session_ui()
        self._refresh_course_view(force=True)

    def _auto_load_default_roster(self) -> None:
        if not self.default_roster_path:
            return
        loaded = self._load_roster_from_path(self.default_roster_path, persist=False, show_errors=False)
        if loaded:
            self.status_label.setText(f"Lista restaurada: {self.default_roster_path.name}")
        else:
            self.logger.warning("No se pudo restaurar la lista predeterminada: %s", self.default_roster_path)

    def _photos_root(self) -> Path:
        return self.data_root / PHOTOS_DIRNAME

    def _backup_root(self) -> Path:
        return self.data_root / BACKUP_PHOTOS_DIRNAME

    def _selected_course_for_session(self) -> str:
        text = self.course_combo.currentText().strip()
        return text or "SIN CURSO"

    def start_session(self) -> None:
        course_display = self._selected_course_for_session()
        mode = "course" if self.course_radio.isChecked() else "test"
        photos_root = self._photos_root()
        backup_root = self._backup_root()
        photos_root.mkdir(parents=True, exist_ok=True)
        backup_root.mkdir(parents=True, exist_ok=True)

        if mode == "test":
            course_display = "PRUEBA"
            course_slug = "PRUEBA"
            session_dir = photos_root / TEST_FOLDER_NAME
            backup_dir = backup_root / TEST_FOLDER_NAME
        else:
            course_slug = sanitize_folder_name(course_display)
            session_dir = photos_root / course_slug
            backup_dir = backup_root / course_slug

        session_dir.mkdir(parents=True, exist_ok=True)
        backup_dir.mkdir(parents=True, exist_ok=True)
        csv_path = session_dir / CSV_FILENAME
        if not csv_path.exists():
            csv_path.write_text("id,filename,student_name,course,rut,timestamp\n", encoding="utf-8")
        try:
            rewrite_csv(csv_path, load_existing_records(csv_path))
            rewrite_csv(backup_dir / CSV_FILENAME, load_existing_records(csv_path))
        except Exception:
            pass

        records = load_existing_records(csv_path)
        roster_students: list[RosterStudent] = []
        if mode == "course":
            roster_students = self.roster_map.get(course_display, [])
        roster_index = min(len(records), len(roster_students)) if roster_students else 0

        self.session = SessionState(
            mode=mode,
            course_display=course_display,
            course_slug=course_slug,
            photos_root=photos_root,
            backup_root=backup_root,
            session_dir=session_dir,
            backup_dir=backup_dir,
            csv_path=csv_path,
            records=records,
            started_at=datetime.now(),
            roster_students=roster_students,
            roster_index=roster_index,
        )
        self._advance_roster_past_completed()

        self.logger.info(
            "Session started. mode=%s course=%s session_dir=%s backup_dir=%s roster_students=%s",
            mode,
            course_display,
            session_dir,
            backup_dir,
            len(roster_students),
        )
        self.status_label.setText(f"Sesion iniciada: {course_display}")
        self._sync_session_ui()
        self._refresh_course_view(force=True)
        self._render_preview()
        self._schedule_preview_render(immediate=True)

    def close_session(self) -> None:
        if self.session is None:
            return
        self.logger.info("Session closed. records=%s csv=%s", len(self.session.records), self.session.csv_path)
        self.status_label.setText("Sesion cerrada.")
        self.session = None
        self.pending_capture_identity = None
        self._update_session_labels()
        self._sync_session_ui()
        self._refresh_course_view(force=True)

    def open_photos_root(self) -> None:
        open_folder(self._photos_root())

    def _apply_settings_to_frame(self, frame: np.ndarray, for_preview: bool = True) -> np.ndarray:
        transformed = frame.copy()
        if self.mirror_check.isChecked():
            transformed = cv2.flip(transformed, 1)
        transformed = _rotate_frame(transformed, self.rotation_combo.currentText())
        detect_face = self.face_check.isChecked() or self.crop_check.isChecked()
        face_box = self._detect_primary_face(transformed, for_preview=for_preview) if detect_face else None
        self.current_face_box = face_box

        output = transformed
        crop_box: Optional[tuple[int, int, int, int]] = None
        if self.crop_check.isChecked():
            crop_box = self._compute_portrait_crop_box(
                transformed.shape[1],
                transformed.shape[0],
                face_box,
                self.current_eye_centers,
            )
            crop_box = self._smooth_crop_box(crop_box)
            crop_box = self._apply_manual_crop_tuning(crop_box, transformed.shape[1], transformed.shape[0])
            self.current_crop_box = crop_box
            target_size = (450, 600) if for_preview else (1500, 2000)
            output = self._crop_frame_with_box(transformed, crop_box, output_size=target_size)
        else:
            self.current_crop_box = None
            self.stable_crop_box = None
            self.current_eye_centers = []

        if for_preview:
            if self.guide_check.isChecked():
                self._draw_context_guides(output, transformed.shape[1], transformed.shape[0], crop_box, face_box)
            if self.face_check.isChecked() and face_box is not None:
                self._draw_face_anchor(output, crop_box, face_box)

        return output

    def _source_backup_frame(self, frame: np.ndarray) -> np.ndarray:
        """Return the camera source with only user orientation settings, before any portrait crop."""
        backup_frame = frame.copy()
        if self.mirror_check.isChecked():
            backup_frame = cv2.flip(backup_frame, 1)
        return _rotate_frame(backup_frame, self.rotation_combo.currentText())

    def _next_backup_path(self, backup_path: Path) -> Path:
        """Avoid overwriting older source backups when a student is photographed again."""
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        if not backup_path.exists():
            return backup_path
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return backup_path.with_name(f"{backup_path.stem}__reintento_{stamp}{backup_path.suffix}")

    def _compute_portrait_crop_box(
        self,
        width: int,
        height: int,
        face_box: Optional[tuple[int, int, int, int]],
        eye_centers: Optional[list[tuple[int, int]]] = None,
    ) -> tuple[int, int, int, int]:
        target_ratio = 3 / 4
        max_crop_h = min(height, int(width / target_ratio))
        max_crop_w = int(max_crop_h * target_ratio)

        if face_box is not None:
            fx, fy, fw, fh = face_box
            face_cx = fx + fw / 2
            face_cy = fy + fh / 2
            # Abrimos lo justo para sostener hombros y cara sin dejar tanto aire vacío.
            crop_h = max(int(fh * 1.85), int(height * 0.60))
            crop_h = min(crop_h, max(220, int(height * 0.88)))
            crop_h = min(crop_h, max_crop_h)
            crop_w = int(crop_h * target_ratio)
            if fx < width * 0.36:
                x1 = int(face_cx - crop_w * 0.62)
            elif fx > width * 0.64:
                x1 = int(face_cx - crop_w * 0.38)
            else:
                x1 = int(face_cx - crop_w / 2)
            if eye_centers:
                eye_y = sum(pt[1] for pt in eye_centers) / len(eye_centers)
                y1 = int(eye_y - crop_h * 0.24)
            else:
                y1 = int(face_cy - crop_h * 0.30)
            # Haar commonly starts inside the hairline; always reserve visible headroom.
            y1 = min(y1, int(fy - fh * 0.30))
        else:
            crop_h = int(max_crop_h * 0.82)
            crop_h = max(CROP_MIN_HEIGHT, crop_h)
            crop_w = int(crop_h * target_ratio)
            x1 = (width - crop_w) // 2
            y1 = (height - crop_h) // 2

        crop_w = min(crop_w, max_crop_w)
        crop_h = min(crop_h, height)
        x1 = max(0, min(x1, width - crop_w))
        y1 = max(0, min(y1, height - crop_h))
        return x1, y1, x1 + crop_w, y1 + crop_h

    def _apply_manual_crop_tuning(
        self,
        crop_box: tuple[int, int, int, int],
        width: int,
        height: int,
    ) -> tuple[int, int, int, int]:
        if (
            abs(self.crop_manual_dx) < 1e-6
            and abs(self.crop_manual_dy) < 1e-6
            and abs(self.crop_manual_zoom - 1.0) < 1e-6
        ):
            return crop_box

        x1, y1, x2, y2 = crop_box
        crop_w = x2 - x1
        crop_h = y2 - y1
        center_x = x1 + crop_w / 2 + (crop_w * self.crop_manual_dx)
        center_y = y1 + crop_h / 2 + (crop_h * self.crop_manual_dy)
        tuned_h = max(CROP_MIN_HEIGHT, int(crop_h * self.crop_manual_zoom))
        tuned_w = int(tuned_h * (3 / 4))
        tuned_w = min(tuned_w, width)
        tuned_h = min(tuned_h, height)
        tuned_w = min(tuned_w, int(tuned_h * (3 / 4)))
        x1 = int(center_x - tuned_w / 2)
        y1 = int(center_y - tuned_h / 2)
        x1 = max(0, min(x1, width - tuned_w))
        y1 = max(0, min(y1, height - tuned_h))
        return x1, y1, x1 + tuned_w, y1 + tuned_h

    def _smooth_crop_box(self, next_box: tuple[int, int, int, int]) -> tuple[int, int, int, int]:
        if self.stable_crop_box is None:
            self.stable_crop_box = next_box
            return next_box

        prev_x1, prev_y1, prev_x2, prev_y2 = self.stable_crop_box
        next_x1, next_y1, next_x2, next_y2 = next_box
        prev_w = prev_x2 - prev_x1
        prev_h = prev_y2 - prev_y1
        next_w = next_x2 - next_x1
        next_h = next_y2 - next_y1

        move_threshold_x = max(10, int(prev_w * 0.06))
        move_threshold_y = max(10, int(prev_h * 0.06))
        size_threshold_w = max(12, int(prev_w * 0.08))
        size_threshold_h = max(12, int(prev_h * 0.08))

        if (
            abs(next_x1 - prev_x1) < move_threshold_x
            and abs(next_y1 - prev_y1) < move_threshold_y
            and abs(next_w - prev_w) < size_threshold_w
            and abs(next_h - prev_h) < size_threshold_h
        ):
            return self.stable_crop_box

        alpha = 0.24
        prev_cx = (prev_x1 + prev_x2) / 2
        prev_cy = (prev_y1 + prev_y2) / 2
        next_cx = (next_x1 + next_x2) / 2
        next_cy = (next_y1 + next_y2) / 2

        blended_cx = prev_cx + (next_cx - prev_cx) * alpha
        blended_cy = prev_cy + (next_cy - prev_cy) * alpha
        blended_h = max(260, prev_h + (next_h - prev_h) * alpha)
        blended_w = blended_h * (3 / 4)

        x1 = int(blended_cx - blended_w / 2)
        y1 = int(blended_cy - blended_h / 2)
        x2 = int(x1 + blended_w)
        y2 = int(y1 + blended_h)
        self.stable_crop_box = (x1, y1, x2, y2)
        return self.stable_crop_box

    def _crop_frame_with_box(
        self,
        frame: np.ndarray,
        crop_box: tuple[int, int, int, int],
        output_size: Optional[tuple[int, int]] = (900, 1200),
    ) -> np.ndarray:
        x1, y1, x2, y2 = crop_box
        crop = frame[y1:y2, x1:x2]
        if crop.size == 0:
            return frame
        if output_size is None:
            return crop.copy()
        crop_h, crop_w = crop.shape[:2]
        target_w, target_h = output_size
        if target_w < crop_w or target_h < crop_h:
            interpolation = cv2.INTER_AREA
        else:
            interpolation = cv2.INTER_CUBIC
        return cv2.resize(crop, output_size, interpolation=interpolation)

    def _draw_context_guides(
        self,
        frame: np.ndarray,
        source_width: int,
        source_height: int,
        crop_box: Optional[tuple[int, int, int, int]],
        face_box: Optional[tuple[int, int, int, int]],
    ) -> None:
        height, width = frame.shape[:2]
        if crop_box is not None:
            x1, y1, x2, y2 = crop_box
            scale_x = width / max(1, x2 - x1)
            scale_y = height / max(1, y2 - y1)
            cv2.rectangle(frame, (8, 8), (width - 9, height - 9), (93, 201, 244), 2)
            cv2.line(frame, (width // 2, 8), (width // 2, height - 9), (93, 201, 244), 1)
            cv2.line(frame, (8, int(height * 0.38)), (width - 9, int(height * 0.38)), (93, 201, 244), 1)
            if face_box is not None:
                fx, fy, fw, fh = face_box
                rel_x1 = int((fx - x1) * scale_x)
                rel_y1 = int((fy - y1) * scale_y)
                rel_x2 = int((fx + fw - x1) * scale_x)
                rel_y2 = int((fy + fh - y1) * scale_y)
                if rel_x2 > 0 and rel_y2 > 0 and rel_x1 < width and rel_y1 < height:
                    cv2.rectangle(
                        frame,
                        (max(0, rel_x1), max(0, rel_y1)),
                        (min(width - 1, rel_x2), min(height - 1, rel_y2)),
                        (244, 201, 93),
                        2,
                    )
        else:
            self._draw_guides(frame)

    def _draw_face_anchor(self, frame: np.ndarray, crop_box: Optional[tuple[int, int, int, int]], face_box: tuple[int, int, int, int]) -> None:
        if crop_box is None:
            x, y, w, h = face_box
            cv2.rectangle(frame, (x, y), (x + w, y + h), (244, 201, 93), 2)
            cv2.putText(frame, "Rostro", (x, max(24, y - 8)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (244, 201, 93), 1, cv2.LINE_AA)
            return

        x1, y1, x2, y2 = crop_box
        out_w = frame.shape[1]
        out_h = frame.shape[0]
        scale_x = out_w / max(1, x2 - x1)
        scale_y = out_h / max(1, y2 - y1)
        fx, fy, fw, fh = face_box
        rel_x1 = int((fx - x1) * scale_x)
        rel_y1 = int((fy - y1) * scale_y)
        rel_x2 = int((fx + fw - x1) * scale_x)
        rel_y2 = int((fy + fh - y1) * scale_y)
        cv2.rectangle(frame, (max(0, rel_x1), max(0, rel_y1)), (min(out_w - 1, rel_x2), min(out_h - 1, rel_y2)), (244, 201, 93), 2)
        cv2.putText(
            frame,
            "Rostro",
            (max(8, rel_x1), max(28, rel_y1 - 8)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.5,
            (244, 201, 93),
            1,
            cv2.LINE_AA,
        )

    def _preview_status_lines(self) -> list[str]:
        course = self._active_course_text()
        current = self._current_student()
        student = current.display_name if current is not None else "-"
        rut = current.rut if current is not None and current.rut else "-"
        mode_text = "Curso" if self.course_radio.isChecked() else "Prueba"
        status = "Vista previa"
        if self.session is not None:
            status = f"{len(self.session.records)} capturados | {self.session.roster_remaining} pendientes"
        elif self.roster_map:
            status = f"{len(self._active_students())} alumnos cargados"
        return [
            f"{course} | {mode_text}",
            f"{student} | RUT: {rut} | {status}",
        ]

    def _draw_guides(self, frame: np.ndarray) -> None:
        height, width = frame.shape[:2]
        center_x = width // 2
        center_y = height // 2
        guide_w = int(width * 0.26)
        guide_h = int(height * 0.46)
        left = max(20, center_x - guide_w // 2)
        right = min(width - 20, center_x + guide_w // 2)
        top = max(40, center_y - guide_h // 2)
        bottom = min(height - 30, center_y + guide_h // 2)
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 210, 255), 2)
        cv2.line(frame, (center_x, top), (center_x, bottom), (0, 210, 255), 1)
        cv2.line(frame, (left, center_y), (right, center_y), (0, 210, 255), 1)
        cv2.putText(frame, "Guia", (left, max(20, top - 8)), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 210, 255), 1, cv2.LINE_AA)

    def _expand_box(
        self,
        box: tuple[int, int, int, int],
        width: int,
        height: int,
        scale: float = 1.0,
        pad_x: int = 0,
        pad_y: int = 0,
    ) -> tuple[int, int, int, int]:
        x, y, w, h = box
        cx = x + w / 2
        cy = y + h / 2
        new_w = max(1, int(w * scale) + pad_x * 2)
        new_h = max(1, int(h * scale) + pad_y * 2)
        x1 = max(0, int(cx - new_w / 2))
        y1 = max(0, int(cy - new_h / 2))
        x2 = min(width, x1 + new_w)
        y2 = min(height, y1 + new_h)
        return x1, y1, x2, y2

    def _clip_box(
        self,
        box: tuple[int, int, int, int],
        width: int,
        height: int,
    ) -> tuple[int, int, int, int]:
        x1, y1, x2, y2 = box
        x1 = max(0, min(x1, width - 1))
        y1 = max(0, min(y1, height - 1))
        x2 = max(x1 + 1, min(x2, width))
        y2 = max(y1 + 1, min(y2, height))
        return x1, y1, x2, y2

    def _detect_eyes_in_face(self, gray: np.ndarray, face_box: tuple[int, int, int, int]) -> list[tuple[int, int]]:
        x, y, w, h = face_box
        roi = gray[y : y + h, x : x + w]
        if roi.size == 0:
            return []
        eye_min_w = max(12, w // 10)
        eye_min_h = max(8, h // 12)
        eye_centers: list[tuple[int, int]] = []
        for cascade in self.eye_cascades:
            if cascade.empty():
                continue
            eyes = cascade.detectMultiScale(
                roi,
                scaleFactor=1.06,
                minNeighbors=4,
                minSize=(eye_min_w, eye_min_h),
            )
            for ex, ey, ew, eh in eyes:
                eye_centers.append((x + ex + ew // 2, y + ey + eh // 2))
            if len(eye_centers) >= 2:
                break
        eye_centers.sort(key=lambda pt: pt[0])
        unique: list[tuple[int, int]] = []
        for pt in eye_centers:
            if not unique or abs(unique[-1][0] - pt[0]) > 12 or abs(unique[-1][1] - pt[1]) > 12:
                unique.append(pt)
            if len(unique) >= 2:
                break
        return unique

    def _score_face_candidate(
        self,
        box: tuple[int, int, int, int],
        width: int,
        height: int,
        eye_count: int,
    ) -> float:
        x, y, w, h = box
        area = float(w * h)
        cx = x + w / 2
        cy = y + h / 2
        center_bias = 1.0 - min(1.0, (abs(cx - width / 2) / max(1, width)) * 1.15 + (abs(cy - height * 0.42) / max(1, height)) * 0.55)
        aspect = w / max(1, h)
        aspect_penalty = 1.0 - min(0.38, abs(aspect - 0.78) * 0.16)
        eye_bonus = 1.0 + (0.22 * min(2, eye_count))
        return area * max(0.2, center_bias) * aspect_penalty * eye_bonus

    def _detect_face_candidates(
        self,
        frame: np.ndarray,
        mirrored: bool = False,
        offset: tuple[int, int] = (0, 0),
        for_preview: bool = False,
    ) -> list[tuple[int, int, int, int]]:
        if frame.size == 0:
            return []
        height, width = frame.shape[:2]
        detect_frame = frame
        scale = 1.0
        if width > FACE_DETECT_MAX_WIDTH:
            scale = FACE_DETECT_MAX_WIDTH / width
            detect_frame = cv2.resize(frame, (FACE_DETECT_MAX_WIDTH, max(1, int(height * scale))), interpolation=cv2.INTER_AREA)

        gray = cv2.cvtColor(detect_frame, cv2.COLOR_BGR2GRAY)
        gray = cv2.equalizeHist(gray)
        candidates: list[tuple[int, int, int, int]] = []
        cascades_to_use = [self.face_cascades[0]] if for_preview else self.face_cascades
        profiles_to_use = ((1.06, 4, (45, 45)),) if for_preview else (
            (1.03, 3, (28, 28)),
            (1.05, 4, (40, 40)),
            (1.08, 5, (60, 60)),
        )
        for cascade in cascades_to_use:
            if cascade.empty():
                continue
            for scale_factor, min_neighbors, min_size in profiles_to_use:
                faces = cascade.detectMultiScale(
                    gray,
                    scaleFactor=scale_factor,
                    minNeighbors=min_neighbors,
                    minSize=min_size,
                )
                for fx, fy, fw, fh in faces:
                    if scale != 1.0:
                        fx = int(fx / scale)
                        fy = int(fy / scale)
                        fw = int(fw / scale)
                        fh = int(fh / scale)
                    if mirrored:
                        fx = width - fx - fw
                    fx += offset[0]
                    fy += offset[1]
                    candidates.append((fx, fy, fw, fh))
        return candidates

    def _detect_primary_face(
        self,
        frame: np.ndarray,
        update_state: bool = True,
        for_preview: bool = False,
    ) -> Optional[tuple[int, int, int, int]]:
        if update_state and for_preview and self.frame_counter % 5 != 0:
            return self.current_face_box
        if not self.face_cascades or all(cascade.empty() for cascade in self.face_cascades):
            return None
        height, width = frame.shape[:2]
        search_frames: list[tuple[np.ndarray, bool, tuple[int, int]]] = [(frame, False, (0, 0))]
        if not for_preview:
            search_frames.append((cv2.flip(frame, 1), True, (0, 0)))
        if update_state and self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
            roi = self._expand_box(self.current_face_box, width, height, scale=2.15, pad_x=32, pad_y=32)
            x1, y1, x2, y2 = roi
            search_frames.insert(0, (frame[y1:y2, x1:x2], False, (x1, y1)))

        candidates: list[tuple[int, int, int, int]] = []
        for search_frame, mirrored, offset in search_frames:
            if search_frame.size == 0:
                continue
            candidates.extend(self._detect_face_candidates(search_frame, mirrored=mirrored, offset=offset, for_preview=for_preview))
            if candidates:
                break

        if not candidates:
            if update_state and self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
                return self.current_face_box
            if update_state:
                self.current_face_box = None
                self.current_eye_centers = []
            return None

        detection_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        detection_gray = cv2.equalizeHist(detection_gray)
        evaluated_candidates: list[tuple[tuple[int, int, int, int], list[tuple[int, int]]]] = []
        for box in candidates:
            x, y, w, h = self._clip_box(box, width, height)
            if w < 36 or h < 36:
                continue

            cx = x + w / 2
            cy = y + h / 2
            # 1. Filtro de tamaño relativo (ancho de la cabeza entre 7% y 28% de la pantalla)
            if w < width * 0.07 or w > width * 0.28:
                continue
            # 2. Filtro de centrado horizontal (el estudiante debe estar en el rango [30%, 70%] de la pantalla)
            if cx < width * 0.30 or cx > width * 0.70:
                continue
            # 3. Filtro de centrado vertical (el rostro debe estar en la zona superior/media [15%, 65%] de la pantalla)
            if cy < height * 0.15 or cy > height * 0.65:
                continue

            eyes = self._detect_eyes_in_face(detection_gray, (x, y, w, h))
            evaluated_candidates.append(((x, y, w, h), eyes))

        verified = [candidate for candidate in evaluated_candidates if len(candidate[1]) >= 2]
        if verified:
            selection_pool = verified
        else:
            # Avoid anchoring the portrait to a large false positive on clothing or furniture.
            selection_pool = [
                candidate
                for candidate in evaluated_candidates
                if candidate[0][1] + candidate[0][3] / 2 <= height * 0.62
            ]

        best_face: Optional[tuple[int, int, int, int]] = None
        best_eyes: list[tuple[int, int]] = []
        best_score = -1.0
        for (x, y, w, h), eyes in selection_pool:
            score = self._score_face_candidate((x, y, w, h), width, height, len(eyes))
            if update_state and self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
                prev_x, prev_y, prev_w, prev_h = self.current_face_box
                prev_cx = prev_x + prev_w / 2
                prev_cy = prev_y + prev_h / 2
                cur_cx = x + w / 2
                cur_cy = y + h / 2
                distance = abs(cur_cx - prev_cx) + abs(cur_cy - prev_cy)
                score -= distance * 3.0
            if score > best_score:
                best_score = score
                best_face = (x, y, w, h)
                best_eyes = eyes

        if best_face is None:
            if update_state and self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
                return self.current_face_box
            if update_state:
                self.current_face_box = None
                self.current_eye_centers = []
            return None

        if update_state:
            self.current_face_box = best_face
            self.current_eye_centers = best_eyes
            self.last_face_detect_frame = self.frame_counter
            return self.current_face_box
        return best_face

    def _render_preview(self) -> None:
        frame_source = self._pull_latest_preview_frame()
        if frame_source is None:
            self._show_preview_message("Esperando señal de camara...")
            return
        self.frame_counter += 1
        if not self._preview_frame_received:
            self._preview_frame_received = True
            self.first_frame_timer.stop()
            self.logger.info("First preview frame received. shape=%s", getattr(frame_source, "shape", None))
        
        # Obtener el frame de vista previa completo (rotado/volteado) sin recortar
        frame = self._render_preview_frame(frame_source)
        height, width = frame.shape[:2]

        # Calcular la caja de recorte para el preview basada en fallback centrado
        target_ratio = 3 / 4
        max_crop_h = min(height, int(width / target_ratio))
        max_crop_w = int(max_crop_h * target_ratio)
        crop_h = int(max_crop_h * 0.82)
        crop_h = max(CROP_MIN_HEIGHT, crop_h)
        crop_w = int(crop_h * target_ratio)
        x1 = (width - crop_w) // 2
        y1 = (height - crop_h) // 2
        crop_box = (x1, y1, x1 + crop_w, y1 + crop_h)

        # Aplicar el ajuste manual si existe
        crop_box = self._apply_manual_crop_tuning(crop_box, width, height)

        # Dibujar la caja de recorte (rectángulo azul/cian) si está activada
        if self.crop_check.isChecked() and self.guide_check.isChecked():
            cv2.rectangle(frame, (crop_box[0], crop_box[1]), (crop_box[2], crop_box[3]), (244, 201, 93), 2)
            # Dibujar líneas del centro del recorte
            cx = (crop_box[0] + crop_box[2]) // 2
            cy_eye = crop_box[1] + int((crop_box[3] - crop_box[1]) * 0.38)
            cv2.line(frame, (cx, crop_box[1]), (cx, crop_box[3]), (244, 201, 93), 1)
            cv2.line(frame, (crop_box[0], cy_eye), (crop_box[2], cy_eye), (244, 201, 93), 1)

        # Dibujar la guía estática de encuadre si está activada
        if self.guide_check.isChecked():
            self._draw_guides(frame)

        if self.guide_check.isChecked():
            info_lines = self._preview_status_lines()
            self._draw_preview_banner(frame, info_lines, 0)

        pixmap = _cv_to_qpixmap(frame)
        scaled = pixmap.scaled(
            self.preview_label.size(),
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.FastTransformation,
        )
        self.preview_label.setPixmap(scaled)
        self.preview_label.setText("")

    def _render_preview_frame(self, frame: np.ndarray) -> np.ndarray:
        preview_frame = frame
        width = frame.shape[1]
        if width > PREVIEW_MAX_WIDTH:
            scale = PREVIEW_MAX_WIDTH / width
            preview_frame = cv2.resize(
                frame,
                (PREVIEW_MAX_WIDTH, max(1, int(frame.shape[0] * scale))),
                interpolation=cv2.INTER_AREA,
            )

        transformed = preview_frame.copy()
        if self.mirror_check.isChecked():
            transformed = cv2.flip(transformed, 1)
        transformed = _rotate_frame(transformed, self.rotation_combo.currentText())

        # Preview stays intentionally lightweight; heavy detection runs on capture.

        return transformed

    def _smooth_preview_crop_box(self, next_box: tuple[int, int, int, int]) -> tuple[int, int, int, int]:
        if self.preview_stable_crop_box is None:
            self.preview_stable_crop_box = next_box
            return next_box

        prev_x1, prev_y1, prev_x2, prev_y2 = self.preview_stable_crop_box
        next_x1, next_y1, next_x2, next_y2 = next_box
        prev_w = prev_x2 - prev_x1
        prev_h = prev_y2 - prev_y1
        next_w = next_x2 - next_x1
        next_h = next_y2 - next_y1

        move_threshold_x = max(10, int(prev_w * 0.06))
        move_threshold_y = max(10, int(prev_h * 0.06))
        size_threshold_w = max(12, int(prev_w * 0.08))
        size_threshold_h = max(12, int(prev_h * 0.08))

        if (
            abs(next_x1 - prev_x1) < move_threshold_x
            and abs(next_y1 - prev_y1) < move_threshold_y
            and abs(next_w - prev_w) < size_threshold_w
            and abs(next_h - prev_h) < size_threshold_h
        ):
            return self.preview_stable_crop_box

        alpha = 0.24
        prev_cx = (prev_x1 + prev_x2) / 2
        prev_cy = (prev_y1 + prev_y2) / 2
        next_cx = (next_x1 + next_x2) / 2
        next_cy = (next_y1 + next_y2) / 2

        blended_cx = prev_cx + (next_cx - prev_cx) * alpha
        blended_cy = prev_cy + (next_cy - prev_cy) * alpha
        blended_h = max(260, prev_h + (next_h - prev_h) * alpha)
        blended_w = blended_h * (3 / 4)

        x1 = int(blended_cx - blended_w / 2)
        y1 = int(blended_cy - blended_h / 2)
        x2 = int(x1 + blended_w)
        y2 = int(y1 + blended_h)
        self.preview_stable_crop_box = (x1, y1, x2, y2)
        return self.preview_stable_crop_box

    def _draw_preview_banner(self, frame: np.ndarray, lines: list[str], banner_height: int) -> None:
        if not lines:
            return

        height, width = frame.shape[:2]
        lines = lines[:2]
        text_scales = [0.45, 0.39]
        text_thickness = [1, 1]
        padding_x = 12
        padding_y = 8
        line_gap = 6
        text_sizes = [
            cv2.getTextSize(line, cv2.FONT_HERSHEY_SIMPLEX, text_scales[idx], text_thickness[idx])[0]
            for idx, line in enumerate(lines)
        ]
        chip_w = max(size[0] for size in text_sizes) + padding_x * 2
        chip_h = sum(size[1] for size in text_sizes) + padding_y * 2 + line_gap * max(0, len(lines) - 1)
        chip_w = min(chip_w, max(220, width - 24))
        chip_h = max(32, chip_h)
        x1 = 12
        y1 = max(12, height - chip_h - 12)
        x2 = min(width - 12, x1 + chip_w)
        y2 = min(height - 12, y1 + chip_h)

        overlay = frame.copy()
        cv2.rectangle(overlay, (x1, y1), (x2, y2), (14, 12, 18), -1)
        cv2.addWeighted(overlay, 0.52, frame, 0.48, 0, frame)
        cv2.rectangle(frame, (x1, y1), (x2, y2), (92, 77, 126), 1)

        text_x = x1 + padding_x
        text_y = y1 + padding_y + text_sizes[0][1]
        for idx, line in enumerate(lines):
            cv2.putText(
                frame,
                line,
                (text_x, text_y),
                cv2.FONT_HERSHEY_SIMPLEX,
                text_scales[idx],
                (244, 240, 247),
                text_thickness[idx],
                cv2.LINE_AA,
            )
            if idx + 1 < len(lines):
                text_y += text_sizes[idx + 1][1] + line_gap

    def _schedule_preview_render(self, immediate: bool = False) -> None:
        if self.preview_render_timer.isActive():
            if immediate:
                self.preview_render_timer.stop()
            else:
                return
        delay_ms = 1 if immediate else 33
        self.preview_render_timer.start(delay_ms)

    def _render_preview_safe(self) -> None:
        try:
            self._render_preview()
        except Exception as exc:
            self.logger.exception("Preview render failed: %s", exc)
            self._show_preview_message(f"Error al renderizar vista: {exc}")
        finally:
            if self.camera_thread is not None or self.countdown_remaining > 0:
                self._schedule_preview_render()

    def _on_first_frame_timeout(self) -> None:
        if self._preview_frame_received:
            return
        self.logger.warning(
            "No preview frames received after timeout. camera_index=%s backend=%s alias=%s",
            self.camera_index,
            self.backend_name,
            self.camera_alias,
        )
        self.status_label.setText("La camara abre, pero aun no entrega imagen.")
        self._show_preview_message("La camara abre, pero no entrega imagen")
        self._try_next_camera_source("no entregar imagen")

    def _current_capture_student(self) -> tuple[str, str]:
        name, rut, _roster_index = self._capture_identity_snapshot()
        return name, rut

    def _capture_identity_snapshot(self) -> tuple[str, str, Optional[int]]:
        if self.session and self.session.has_roster:
            current = self.session.current_student()
            if current is not None:
                return current.display_name, current.rut, self.session.roster_index
        return self.student_edit.text().strip(), "", None

    def _save_capture(self, frame: np.ndarray, capture_identity: Optional[tuple[str, str, Optional[int]]] = None) -> None:
        if self.session is None:
            self.status_label.setText("Primero inicia una sesion.")
            self.logger.warning("Capture requested without active session.")
            return

        student_name, rut, roster_index = capture_identity or self._capture_identity_snapshot()
        if not student_name:
            self.status_label.setText("Escribe o selecciona un estudiante antes de capturar.")
            return

        if self.session.has_roster and self.session.current_student() is None:
            self.status_label.setText("La nomina ya termino.")
            return

        if self.session.has_roster and roster_index is not None and 0 <= roster_index < self.session.roster_total:
            self.session.roster_index = roster_index

        self.frame_counter += 1
        source_backup = self._source_backup_frame(frame)
        record = PhotoRecord(
            id=self.session.next_id,
            filename=build_photo_filename(student_name, self.session.course_display, rut),
            student_name=student_name,
            course=self.session.course_display,
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            rut=rut,
        )

        image_path = self.session.session_dir / record.filename
        backup_path = self.session.backup_dir / record.filename
        try:
            # Guardar la imagen original completa (orientada) tanto en la ruta principal como en el respaldo
            if not cv2.imwrite(str(image_path), source_backup):
                raise RuntimeError(f"No se pudo guardar la foto en {image_path}")
            if not cv2.imwrite(str(backup_path), source_backup):
                raise RuntimeError(f"No se pudo guardar el respaldo en {backup_path}")
            append_csv_record(self.session.csv_path, record)
            ensure_photo_backup(self.session.csv_path, self.session.backup_dir / CSV_FILENAME)
            self._launch_photo_autoframe(
                image_path=image_path,
                backup_path=backup_path,
                dx=self.crop_manual_dx,
                dy=self.crop_manual_dy,
                zoom=self.crop_manual_zoom,
                crop_enabled=self.crop_check.isChecked(),
            )
            self._reset_crop_tuning()
        except Exception as exc:
            self.logger.exception("Failed to save capture: %s", exc)
            QMessageBox.critical(self, "Error al guardar", f"No se pudo guardar la captura:\n{exc}")
            return

        self.session.records.append(record)
        if self.session.has_roster:
            self.session.advance()
            self._advance_roster_past_completed()
            current = self.session.current_student()
            if current is not None:
                self.student_edit.setText(current.display_name)
            else:
                self.student_edit.clear()
        else:
            self.student_edit.clear()

        self.logger.info(
            "Capture saved. file=%s student=%s rut=%s course=%s records=%s backup=respaldo OK",
            record.filename,
            record.student_name,
            record.rut,
            record.course,
            len(self.session.records),
        )
        self.status_label.setText(f"Guardada: {record.filename}")
        self._append_recent_record(record)
        self._sync_session_ui()
        self._refresh_course_view(force=True)

    def capture_photo(self) -> None:
        if self.is_processing_photo:
            return
        if self.countdown_timer.isActive():
            return
        if self.session is None:
            self.status_label.setText("Primero inicia una sesion.")
            self.logger.warning("Capture requested without active session.")
            return
        frame = self._pull_latest_preview_frame()
        if frame is None:
            self.status_label.setText("No hay frame de camara disponible.")
            return

        capture_identity = self._capture_identity_snapshot()
        if not capture_identity[0]:
            self.status_label.setText("Escribe o selecciona un estudiante antes de capturar.")
            return
        self.pending_capture_identity = capture_identity
        countdown = int(self.countdown_combo.currentText().split()[0])
        if countdown > 0:
            self.countdown_remaining = countdown
            self.status_label.setText(f"Captura en {self.countdown_remaining}s | {capture_identity[0]}")
            self.countdown_timer.start()
            self._schedule_preview_render(immediate=True)
            return

        self._save_capture(frame.copy(), capture_identity)
        self.pending_capture_identity = None

    def _countdown_tick(self) -> None:
        if self.countdown_remaining <= 1:
            self.countdown_timer.stop()
            self.countdown_remaining = 0
            frame = self._pull_latest_preview_frame()
            if frame is not None:
                self._save_capture(frame.copy(), self.pending_capture_identity)
            self.pending_capture_identity = None
            return
        self.countdown_remaining -= 1
        target_name = self.pending_capture_identity[0] if self.pending_capture_identity else ""
        suffix = f" | {target_name}" if target_name else ""
        self.status_label.setText(f"Captura en {self.countdown_remaining}s{suffix}")
        self._schedule_preview_render(immediate=True)

    def retake_last(self) -> None:
        if self.is_processing_photo:
            return
        if self.session is None or not self.session.records:
            self.status_label.setText("No hay foto para rehacer.")
            return
        self.pending_capture_identity = None

        last_record = self.session.records.pop()
        image_path = self.session.session_dir / last_record.filename
        backup_path = self.session.backup_dir / last_record.filename
        if not self._wait_for_autoframe_job(image_path, timeout=10.0):
            self.status_label.setText("Todavia estoy reencuadrando la foto anterior. Espera un momento.")
            self.session.records.append(last_record)
            return
        try:
            if image_path.exists():
                image_path.unlink()
            if backup_path.exists():
                backup_path.unlink()
            rewrite_csv(self.session.csv_path, self.session.records)
            ensure_photo_backup(self.session.csv_path, self.session.backup_dir / CSV_FILENAME)
        except Exception as exc:
            self.logger.exception("Failed to retake last photo: %s", exc)
            QMessageBox.critical(self, "Error", f"No se pudo rehacer la ultima foto:\n{exc}")
            return

        if self.session.has_roster:
            self._reconcile_session_roster_progress()

        self.logger.info("Retake last. removed=%s records=%s", last_record.filename, len(self.session.records))
        self.status_label.setText(f"Foto eliminada: {last_record.filename}")
        self._append_recent_text(f"Rehecha: {last_record.filename}")
        self._sync_session_ui()
        self._refresh_course_view(force=True)

    def _append_recent_text(self, line: str) -> None:
        lines = self.recent_text.toPlainText().splitlines()
        lines.append(line)
        self.recent_text.setPlainText("\n".join(lines[-12:]))

    def _append_recent_record(self, record: PhotoRecord) -> None:
        text = f"{record.id:03d} | {record.student_name} | {record.rut or '-'} | {record.filename}"
        self._append_recent_text(text)

    def _autoframe_script_path(self) -> Path:
        return APP_ROOT / "photo_autoframe.py"

    def _launch_photo_autoframe(
        self,
        image_path: Path,
        backup_path: Path,
        dx: float,
        dy: float,
        zoom: float,
        crop_enabled: bool,
    ) -> None:
        self.is_processing_photo = True
        self.post_progress_bar.setValue(0)
        self.post_progress_bar.setFormat("Iniciando procesamiento: 0%")
        self.post_progress_bar.setVisible(True)
        self.capture_button.setEnabled(False)
        self.capture_button_tab.setEnabled(False)
        self.retake_button.setEnabled(False)
        self.retake_button_tab.setEnabled(False)

        def worker() -> None:
            try:
                self.post_progress_updated.emit(10, "Cargando foto original")
                image = cv2.imread(str(backup_path))
                if image is None:
                    self.logger.error("No se pudo leer la imagen original para postprocesado en %s", backup_path)
                    self.autoframe_finished.emit(False, image_path.name)
                    return

                height, width = image.shape[:2]

                if not crop_enabled:
                    self.post_progress_updated.emit(50, "Guardando imagen completa")
                    cv2.imwrite(str(image_path), image)
                    self.post_progress_updated.emit(100, "Completado")
                    self.autoframe_finished.emit(True, image_path.name)
                    return

                has_manual_tuning = (abs(dx) > 1e-6 or abs(dy) > 1e-6 or abs(zoom - 1.0) > 1e-6)
                if has_manual_tuning:
                    self.post_progress_updated.emit(40, "Aplicando recorte manual")
                    target_ratio = 3 / 4
                    max_crop_h = min(height, int(width / target_ratio))
                    max_crop_w = int(max_crop_h * target_ratio)
                    crop_h = int(max_crop_h * 0.82)
                    crop_h = max(CROP_MIN_HEIGHT, crop_h)
                    crop_w = int(crop_h * target_ratio)
                    x1 = (width - crop_w) // 2
                    y1 = (height - crop_h) // 2
                    crop_box = (x1, y1, x1 + crop_w, y1 + crop_h)

                    crop_box = self._apply_manual_crop_tuning_values(crop_box, width, height, dx, dy, zoom)
                    self.post_progress_updated.emit(70, "Procesando recorte")
                    cropped = self._crop_frame_with_box(image, crop_box, output_size=(1500, 2000))
                    self.post_progress_updated.emit(90, "Guardando imagen recortada")
                    cv2.imwrite(str(image_path), cropped)
                    self.post_progress_updated.emit(100, "Completado")
                    self.autoframe_finished.emit(True, image_path.name)
                    return

                self.post_progress_updated.emit(30, "Buscando rostro en la foto")
                cascades = self.face_cascades
                eye_cascades = self.eye_cascades
                face_box = self._detect_primary_face_for_postprocessing(image, cascades, eye_cascades)

                if face_box is None:
                    self.logger.warning("No se detecto rostro para %s. Aplicando recorte centrado fallback.", image_path.name)
                    self.post_progress_updated.emit(60, "Rostro no encontrado. Usando centrado")
                    target_ratio = 3 / 4
                    max_crop_h = min(height, int(width / target_ratio))
                    max_crop_w = int(max_crop_h * target_ratio)
                    crop_h = int(max_crop_h * 0.82)
                    crop_h = max(CROP_MIN_HEIGHT, crop_h)
                    crop_w = int(crop_h * target_ratio)
                    x1 = (width - crop_w) // 2
                    y1 = (height - crop_h) // 2
                    crop_box = (x1, y1, x1 + crop_w, y1 + crop_h)
                    cropped = self._crop_frame_with_box(image, crop_box, output_size=(1500, 2000))
                    cv2.imwrite(str(image_path), cropped)
                    self.post_progress_updated.emit(100, "Rostro no encontrado (recorte centrado)")
                    self.autoframe_finished.emit(True, f"{image_path.name} (Rostro no detectado, se aplico recorte centrado)")
                    return

                self.post_progress_updated.emit(60, "Rostro detectado. Buscando ojos")
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                gray = cv2.equalizeHist(gray)
                eye_centers = self._detect_eyes_in_face(gray, face_box)

                self.post_progress_updated.emit(80, "Calculando encuadre final")
                crop_box = self._compute_portrait_crop_box(width, height, face_box, eye_centers)
                cropped = self._crop_frame_with_box(image, crop_box, output_size=(1500, 2000))

                self.post_progress_updated.emit(90, "Guardando imagen final")
                cv2.imwrite(str(image_path), cropped)
                self.post_progress_updated.emit(100, "Completado")
                self.autoframe_finished.emit(True, image_path.name)

            except Exception as exc:
                self.logger.exception("Error en el postprocesamiento asincrono: %s", exc)
                self.autoframe_finished.emit(False, image_path.name)
            finally:
                with self.autoframe_jobs_lock:
                    self.autoframe_jobs.pop(image_path, None)

        thread = threading.Thread(target=worker, name=f"autoframe-{image_path.stem}", daemon=True)
        with self.autoframe_jobs_lock:
            self.autoframe_jobs[image_path] = thread
        thread.start()

    def _on_post_progress_updated(self, value: int, message: str) -> None:
        self.post_progress_bar.setValue(value)
        self.post_progress_bar.setFormat(f"{message}: %p%")

    def _apply_manual_crop_tuning_values(
        self,
        crop_box: tuple[int, int, int, int],
        width: int,
        height: int,
        dx: float,
        dy: float,
        zoom: float,
    ) -> tuple[int, int, int, int]:
        if abs(dx) < 1e-6 and abs(dy) < 1e-6 and abs(zoom - 1.0) < 1e-6:
            return crop_box

        x1, y1, x2, y2 = crop_box
        crop_w = x2 - x1
        crop_h = y2 - y1
        center_x = x1 + crop_w / 2 + (crop_w * dx)
        center_y = y1 + crop_h / 2 + (crop_h * dy)
        tuned_h = max(CROP_MIN_HEIGHT, int(crop_h * zoom))
        tuned_w = int(tuned_h * (3 / 4))
        tuned_w = min(tuned_w, width)
        tuned_h = min(tuned_h, height)
        tuned_w = min(tuned_w, int(tuned_h * (3 / 4)))
        x1 = int(center_x - tuned_w / 2)
        y1 = int(center_y - tuned_h / 2)
        x1 = max(0, min(x1, width - tuned_w))
        y1 = max(0, min(y1, height - tuned_h))
        return x1, y1, x1 + tuned_w, y1 + tuned_h

    def _detect_primary_face_for_postprocessing(
        self,
        frame: np.ndarray,
        face_cascades,
        eye_cascades,
    ) -> Optional[tuple[int, int, int, int]]:
        if not face_cascades or all(cascade.empty() for cascade in face_cascades):
            return None
        height, width = frame.shape[:2]

        search_frames = [(frame, False, (0, 0)), (cv2.flip(frame, 1), True, (0, 0))]
        candidates: list[tuple[int, int, int, int]] = []
        for search_frame, mirrored, offset in search_frames:
            if search_frame.size == 0:
                continue
            candidates.extend(self._detect_face_candidates(search_frame, mirrored=mirrored, offset=offset, for_preview=False))
            if candidates:
                break

        if not candidates:
            return None

        detection_gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        detection_gray = cv2.equalizeHist(detection_gray)
        evaluated_candidates: list[tuple[tuple[int, int, int, int], list[tuple[int, int]]]] = []
        for box in candidates:
            x, y, w, h = self._clip_box(box, width, height)
            if w < 36 or h < 36:
                continue

            cx = x + w / 2
            cy = y + h / 2
            if w < width * 0.07 or w > width * 0.35:
                continue
            if cx < width * 0.30 or cx > width * 0.70:
                continue
            if cy < height * 0.15 or cy > height * 0.65:
                continue

            eyes = self._detect_eyes_in_face(detection_gray, (x, y, w, h))
            evaluated_candidates.append(((x, y, w, h), eyes))

        verified = [candidate for candidate in evaluated_candidates if len(candidate[1]) >= 2]
        if verified:
            selection_pool = verified
        else:
            selection_pool = [
                candidate
                for candidate in evaluated_candidates
                if candidate[0][1] + candidate[0][3] / 2 <= height * 0.65
            ]

        best_face: Optional[tuple[int, int, int, int]] = None
        best_score = -1.0
        for (x, y, w, h), eyes in selection_pool:
            score = self._score_face_candidate((x, y, w, h), width, height, len(eyes))
            if score > best_score:
                best_score = score
                best_face = (x, y, w, h)

        return best_face

    def _on_autoframe_finished(self, success: bool, filename: str) -> None:
        self.post_progress_bar.setValue(100)
        if success:
            if "Rostro no detectado" in filename or "Rostro no encontrado" in filename:
                self.status_label.setText(f"Listo con advertencia: {filename}.")
            else:
                self.status_label.setText(f"Listo: {filename} postprocesado.")
        else:
            self.status_label.setText(f"Error en autoframe: {filename}")
        QTimer.singleShot(500, self._cleanup_post_processing)

    def _cleanup_post_processing(self) -> None:
        self.post_progress_bar.setVisible(False)
        self.post_progress_bar.setValue(0)
        self.capture_button.setEnabled(True)
        self.capture_button_tab.setEnabled(True)
        self.retake_button.setEnabled(True)
        self.retake_button_tab.setEnabled(True)
        self.is_processing_photo = False

    def _wait_for_autoframe_job(self, image_path: Path, timeout: float = 10.0) -> bool:
        with self.autoframe_jobs_lock:
            thread = self.autoframe_jobs.get(image_path)
        if thread is None:
            return True
        thread.join(timeout)
        if thread.is_alive():
            return False
        with self.autoframe_jobs_lock:
            self.autoframe_jobs.pop(image_path, None)
        return True

    def next_roster_student(self) -> None:
        students = self._active_students()
        if not students:
            return
        current = min(self._active_index() + 1, len(students))
        self._set_active_index(current)
        self._sync_session_ui()
        self._refresh_course_view(force=True)
        self._render_preview()

    def prev_roster_student(self) -> None:
        students = self._active_students()
        if not students:
            return
        current = max(0, self._active_index() - 1)
        self._set_active_index(current)
        self._sync_session_ui()
        self._refresh_course_view(force=True)
        self._render_preview()

    def sync_student_with_roster(self) -> None:
        students = self._active_students()
        if not students:
            return
        text = self.student_edit.text().strip()
        target = _normalize_text(text)
        if not target and self.session and self.session.has_roster and self.session.current_student() is not None:
            target = _normalize_text(self.session.current_student().display_name)

        index = 0
        if target:
            for i, student in enumerate(students):
                if _normalize_text(student.display_name) == target or _normalize_rut(student.rut) == _normalize_rut(text):
                    index = i
                    break
        self._set_active_index(index)
        if self.session and self.session.has_roster:
            self.student_edit.setText(students[index].display_name)
        self.status_label.setText(f"Alineado con {students[index].display_name}")
        self._sync_session_ui()
        self._refresh_course_view(force=True)
        self._render_preview()

    def _on_resize(self) -> None:
        self._schedule_preview_render(immediate=True)

    def prev_camera(self) -> None:
        if not self.camera_choices:
            return
        index = (self.camera_combo.currentIndex() - 1) % len(self.camera_choices)
        self.camera_combo.setCurrentIndex(index)

    def next_camera(self) -> None:
        if not self.camera_choices:
            return
        index = (self.camera_combo.currentIndex() + 1) % len(self.camera_choices)
        self.camera_combo.setCurrentIndex(index)

    def resizeEvent(self, event) -> None:  # noqa: N802
        super().resizeEvent(event)
        self._apply_responsive_layout()
        self._on_resize()

    def closeEvent(self, event) -> None:  # noqa: N802
        self.logger.info("GUI closing.")
        self._stop_camera_thread()
        event.accept()


def main() -> int:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    window = CastelCredCamQt()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
