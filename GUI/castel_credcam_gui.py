from __future__ import annotations

import csv
import logging
import sys
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import cv2
import tkinter as tk
from openpyxl import load_workbook
from PIL import Image, ImageTk
from tkinter import filedialog, messagebox, ttk

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from castel_credcam import (  # noqa: E402
    BACKUP_PHOTOS_DIRNAME,
    CSV_FILENAME,
    PHOTOS_DIRNAME,
    TEST_FOLDER_NAME,
    PhotoRecord,
    append_csv_record,
    backend_key_from_id,
    backup_course_dir,
    build_photo_filename,
    configure_capture,
    find_similar_record,
    has_record_for_student,
    ensure_photo_backup,
    append_retake_audit,
    get_logs_dir,
    list_available_cameras,
    load_camera_aliases,
    load_existing_records,
    load_last_camera,
    open_camera,
    open_folder,
    rewrite_csv,
    sanitize_folder_name,
    setup_logging,
    save_last_camera,
    silence_opencv_logs,
)


def _normalize_key(value: str) -> str:
    import unicodedata

    text = str(value).strip().casefold()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _display_name_from_parts(apellido_paterno: str, apellido_materno: str, nombres: str) -> str:
    pieces = [apellido_paterno.strip(), apellido_materno.strip(), nombres.strip()]
    return " ".join(part for part in pieces if part)


def _normalize_rut_key(value: str) -> str:
    text = str(value).strip().upper()
    return "".join(ch for ch in text if ch.isdigit() or ch == "K")


@dataclass
class RosterStudent:
    rut: str
    apellido_paterno: str
    apellido_materno: str
    nombres: str

    @property
    def display_name(self) -> str:
        return _display_name_from_parts(self.apellido_paterno, self.apellido_materno, self.nombres)


APP_TITLE = "CastelCredCam Studio"
WINDOW_BG = "#14061E"
PANEL_BG = "#241033"
CARD_BG = "#31164A"
INFO_BG = "#1A0D28"
ACCENT_PURPLE = "#8B4DFF"
ACCENT_GOLD = "#F4C95D"
TEXT_PRIMARY = "#F7F1FF"
TEXT_MUTED = "#D8C8F2"
SUCCESS = "#6EE7B7"
DANGER = "#FF7A90"
CROP_MIN_HEIGHT = 220
FACE_DETECT_MAX_WIDTH = 1600
FACE_HOLD_FRAMES = 6


@dataclass
class GuiSession:
    mode: str
    course_display: str
    course_slug: str
    photos_root: Path
    backup_root: Path
    session_dir: Path
    backup_dir: Path
    csv_path: Path
    records: list[PhotoRecord]
    started_at: datetime
    roster_students: list[RosterStudent] = field(default_factory=list)
    roster_index: int = 0

    @property
    def next_id(self) -> int:
        return len(self.records) + 1

    @property
    def has_roster(self) -> bool:
        return bool(self.roster_students)

    @property
    def roster_total(self) -> int:
        return len(self.roster_students)

    @property
    def roster_remaining(self) -> int:
        return max(0, len(self.roster_students) - self.roster_index)

    def current_roster_student(self) -> Optional[RosterStudent]:
        if not self.roster_students or self.roster_index < 0 or self.roster_index >= len(self.roster_students):
            return None
        return self.roster_students[self.roster_index]

    def advance_roster(self) -> Optional[RosterStudent]:
        if not self.roster_students:
            return None
        self.roster_index = min(self.roster_index + 1, len(self.roster_students))
        return self.current_roster_student()

    def retreat_roster(self) -> Optional[RosterStudent]:
        if not self.roster_students:
            return None
        self.roster_index = max(0, self.roster_index - 1)
        return self.current_roster_student()

    def filename_for(self, student_name: str, rut: str = "") -> str:
        if student_name:
            course_label = "PRUEBA" if self.mode == "test" else self.course_display
            return build_photo_filename(student_name, course_label, rut)
        return f"{self.course_slug}_{len(self.records) + 1:03d}.jpg"


class CastelCredCamGUI:
    def __init__(self) -> None:
        silence_opencv_logs()
        self.logger, self.log_path = setup_logging(APP_ROOT, "gui")
        self.logger.info("=== CastelCredCam GUI start ===")
        self.logger.info("Log file: %s", self.log_path)
        self.logger.info("Logs dir: %s", get_logs_dir(APP_ROOT))
        self.logger.info("Python: %s", sys.version.replace("\n", " "))
        self.logger.info("Executable: %s", sys.executable)
        self.logger.info("CWD: %s", Path.cwd())
        self.logger.info("Args: %s", sys.argv[1:])
        try:
            self.root = tk.Tk()
        except Exception as exc:
            self.logger.exception("Failed to create Tk root window: %s", exc)
            raise
        self.root.title(APP_TITLE)
        self.root.geometry("1500x920")
        self.root.minsize(1200, 760)
        self.root.configure(bg=WINDOW_BG)

        self.aliases = load_camera_aliases(APP_ROOT)
        self.available_cameras = list_available_cameras(self.aliases)
        self.current_camera_index: Optional[int] = None
        self.current_backend_id: Optional[int] = None
        self.current_backend_name = ""
        self.current_camera_alias = ""
        self.capture: Optional[cv2.VideoCapture] = None
        self.preview_job: Optional[str] = None
        self.current_frame = None
        self.tk_image = None
        self.session: Optional[GuiSession] = None
        self.student_entry: Optional[ttk.Entry] = None
        self.student_manual_frame: Optional[tk.Frame] = None
        self.student_clear_button: Optional[ttk.Button] = None
        self.student_card_title_var: Optional[tk.StringVar] = None
        self.course_tree: Optional[tk.Text] = None
        self.course_tree_scroll: Optional[ttk.Scrollbar] = None
        self.course_progress_var: Optional[tk.IntVar] = None
        self.course_progress_text_var: Optional[tk.StringVar] = None
        self.course_total_var: Optional[tk.StringVar] = None
        self.course_current_var: Optional[tk.StringVar] = None
        self.current_face_box: Optional[tuple[int, int, int, int]] = None
        self.current_eye_centers: list[tuple[int, int]] = []
        self.current_crop_box: Optional[tuple[int, int, int, int]] = None
        self.stable_crop_box: Optional[tuple[int, int, int, int]] = None
        self.frame_counter = 0
        self.last_face_detect_frame = -9999
        self.sidebar_canvas: Optional[tk.Canvas] = None
        self.sidebar_content: Optional[ttk.Frame] = None
        self.sidebar_window_id: Optional[int] = None
        self.left_shell: Optional[ttk.Frame] = None
        self.capture_page: Optional[tk.Frame] = None
        self.course_page: Optional[tk.Frame] = None
        self.info_page: Optional[tk.Frame] = None
        self.preview_card: Optional[tk.Frame] = None
        self.preview_toolbar: Optional[tk.Frame] = None
        self.preview_tool_row: Optional[tk.Frame] = None
        self.info_text: Optional[tk.Text] = None
        self.session_title_label: Optional[ttk.Label] = None
        self.sidebar_subtitle_label: Optional[tk.Label] = None
        self.sidebar_title_label: Optional[ttk.Label] = None
        self.notebook: Optional[ttk.Notebook] = None
        self._last_layout_profile: Optional[tuple[int, int]] = None
        self._responsive_job: Optional[str] = None
        self.face_cascades = [
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_alt2.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_alt.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_profileface.xml"),
        ]
        self.eye_cascades = [
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye_tree_eyeglasses.xml"),
            cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye.xml"),
        ]

        self.mode_var = tk.StringVar(value="test")
        self.course_var = tk.StringVar(value="")
        self.student_var = tk.StringVar(value="")
        self.roster_path_var = tk.StringVar(value="Lista no cargada")
        self.roster_status_var = tk.StringVar(value="Carga un Excel o CSV para capturar sin escribir nombres.")
        self.camera_var = tk.StringVar(value="")
        self.preview_camera_var = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="Listo para iniciar. Selecciona camara y sesion.")
        self.session_var = tk.StringVar(value="Sesion no iniciada")
        self.recent_var = tk.StringVar(value="Sin capturas aun.")
        self.roster_map: dict[str, list[RosterStudent]] = {}
        self.roster_lookup: dict[str, str] = {}
        self.roster_preview_index: dict[str, int] = {}

        self.face_guide_var = tk.BooleanVar(value=True)
        self.frame_guide_var = tk.BooleanVar(value=True)
        self.mirror_var = tk.BooleanVar(value=False)
        self.crop_portrait_var = tk.BooleanVar(value=True)
        self.zoom_var = tk.DoubleVar(value=1.0)
        self.rotation_var = tk.StringVar(value="0 deg")
        self.countdown_var = tk.StringVar(value="0 s")

        self._configure_style()
        self._build_layout()
        self._load_camera_choices()
        self.root.bind("<Configure>", self._schedule_responsive_layout)
        self.root.after_idle(self._sync_responsive_layout)
        self.root.report_callback_exception = self._report_tk_exception
        sys.excepthook = self._sys_excepthook
        self._bind_shortcuts()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _configure_style(self) -> None:
        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")
        self.style.configure("Panel.TFrame", background=PANEL_BG)
        self.style.configure("Card.TFrame", background=CARD_BG)
        self.style.configure("Title.TLabel", background=PANEL_BG, foreground=TEXT_PRIMARY, font=("Segoe UI", 22, "bold"))
        self.style.configure("Muted.TLabel", background=CARD_BG, foreground=TEXT_MUTED, font=("Segoe UI", 9))
        self.style.configure("Accent.TButton", background=ACCENT_PURPLE, foreground=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"), padding=8)
        self.style.map("Accent.TButton", background=[("active", "#A56CFF")])
        self.style.configure("Gold.TButton", background=ACCENT_GOLD, foreground="#291600", font=("Segoe UI", 10, "bold"), padding=8)
        self.style.map("Gold.TButton", background=[("active", "#FFD97F")])
        self.style.configure("Danger.TButton", background=DANGER, foreground=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"), padding=8)
        self.style.configure("TEntry", fieldbackground="#FFF9FE", foreground="#180E24", padding=6)
        self.style.configure("TCombobox", fieldbackground="#FFF9FE", foreground="#180E24", padding=4)
        self.style.configure("TRadiobutton", background=CARD_BG, foreground=TEXT_PRIMARY, font=("Segoe UI", 10))
        self.style.map("TRadiobutton", background=[("active", CARD_BG)])
        self.style.configure("TCheckbutton", background=CARD_BG, foreground=TEXT_PRIMARY, font=("Segoe UI", 10))
        self.style.configure("Treeview", background="#1D102A", fieldbackground="#1D102A", foreground=TEXT_PRIMARY, rowheight=28)
        self.style.configure("Treeview.Heading", background="#2F1847", foreground=ACCENT_GOLD, font=("Segoe UI", 10, "bold"))

    def _build_layout(self) -> None:
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        left_shell = ttk.Frame(self.root, style="Panel.TFrame", width=340)
        self.left_shell = left_shell
        left_shell.grid(row=0, column=0, sticky="nsew")
        left_shell.grid_propagate(False)
        left_shell.grid_rowconfigure(0, weight=1)
        left_shell.grid_columnconfigure(0, weight=1)

        sidebar_canvas = tk.Canvas(left_shell, bg=PANEL_BG, highlightthickness=0, bd=0)
        sidebar_scrollbar = ttk.Scrollbar(left_shell, orient="vertical", command=sidebar_canvas.yview)
        sidebar_canvas.configure(yscrollcommand=sidebar_scrollbar.set)
        sidebar_canvas.grid(row=0, column=0, sticky="nsew")
        sidebar_scrollbar.grid(row=0, column=1, sticky="ns")

        sidebar_content = ttk.Frame(sidebar_canvas, style="Panel.TFrame")
        self.sidebar_canvas = sidebar_canvas
        self.sidebar_content = sidebar_content
        self.sidebar_window_id = sidebar_canvas.create_window((0, 0), window=sidebar_content, anchor="nw")

        def _sync_scrollregion(_event: tk.Event) -> None:
            sidebar_canvas.configure(scrollregion=sidebar_canvas.bbox("all"))

        def _sync_content_width(event: tk.Event) -> None:
            if self.sidebar_window_id is not None:
                sidebar_canvas.itemconfigure(self.sidebar_window_id, width=event.width)

        sidebar_content.bind("<Configure>", _sync_scrollregion)
        sidebar_canvas.bind("<Configure>", _sync_content_width)
        self._bind_sidebar_mousewheel(sidebar_canvas)

        right = ttk.Frame(self.root, style="Panel.TFrame")
        right.grid(row=0, column=1, sticky="nsew", padx=(0, 16), pady=16)
        right.grid_rowconfigure(0, weight=1)
        right.grid_columnconfigure(0, weight=1)

        self.sidebar_title_label = ttk.Label(sidebar_content, text=APP_TITLE, style="Title.TLabel")
        self.sidebar_title_label.pack(anchor="w", padx=18, pady=(18, 4))
        self.sidebar_subtitle_label = tk.Label(
            sidebar_content,
            text="Captura por curso con estilo morado y dorado",
            bg=PANEL_BG,
            fg=ACCENT_GOLD,
            font=("Segoe UI", 10, "bold"),
        )
        self.sidebar_subtitle_label.pack(anchor="w", padx=20, pady=(0, 14))

        self._make_session_card(sidebar_content)
        self._make_roster_card(sidebar_content)
        self._make_camera_card(sidebar_content)
        self._make_student_card(sidebar_content)
        self._make_recent_card(sidebar_content)

        notebook = ttk.Notebook(right)
        self.notebook = notebook
        notebook.grid(row=0, column=0, sticky="nsew")
        notebook.bind("<<NotebookTabChanged>>", self._on_notebook_tab_changed)

        capture_page = tk.Frame(notebook, bg=WINDOW_BG)
        self.capture_page = capture_page
        capture_page.grid_rowconfigure(1, weight=1)
        capture_page.grid_columnconfigure(0, weight=1)
        notebook.add(capture_page, text="Captura")

        info_page = tk.Frame(notebook, bg=INFO_BG)
        self.info_page = info_page
        notebook.add(info_page, text="Info")

        course_page = tk.Frame(notebook, bg=INFO_BG)
        self.course_page = course_page
        notebook.add(course_page, text="Curso")

        self.session_title_label = tk.Label(
            capture_page,
            textvariable=self.session_var,
            bg=WINDOW_BG,
            fg=TEXT_PRIMARY,
            anchor="w",
            font=("Segoe UI", 13, "bold"),
            padx=8,
            pady=6,
        )
        self.session_title_label.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        preview_card = tk.Frame(capture_page, bg=CARD_BG, highlightbackground="#4F2B74", highlightthickness=1)
        self.preview_card = preview_card
        preview_card.grid(row=1, column=0, sticky="nsew")
        preview_card.grid_rowconfigure(1, weight=1)
        preview_card.grid_columnconfigure(0, weight=1)

        toolbar = tk.Frame(preview_card, bg="#20102F", padx=10, pady=8)
        self.preview_toolbar = toolbar
        toolbar.grid(row=0, column=0, sticky="ew")
        toolbar.grid_columnconfigure(1, weight=1)

        tk.Label(toolbar, text="Camara", bg="#20102F", fg=ACCENT_GOLD, font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, sticky="w", padx=(0, 8)
        )
        self.preview_camera_combo = ttk.Combobox(toolbar, textvariable=self.preview_camera_var, state="readonly", width=34)
        self.preview_camera_combo.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        self.preview_camera_combo.bind("<<ComboboxSelected>>", lambda _event: self.change_camera(from_preview=True))

        ttk.Checkbutton(toolbar, text="Voltear", variable=self.mirror_var).grid(row=0, column=2, sticky="w", padx=(0, 6))
        ttk.Checkbutton(toolbar, text="Rostro", variable=self.face_guide_var).grid(row=0, column=3, sticky="w", padx=(0, 6))
        ttk.Checkbutton(toolbar, text="Guia", variable=self.frame_guide_var).grid(row=0, column=4, sticky="w", padx=(0, 6))
        ttk.Checkbutton(toolbar, text="Recortar", variable=self.crop_portrait_var).grid(row=0, column=5, sticky="w", padx=(0, 6))
        ttk.Button(toolbar, text="Sig. cam", style="Gold.TButton", command=self.cycle_camera).grid(row=0, column=6, sticky="e")

        tool_row = tk.Frame(preview_card, bg="#180B25", padx=10, pady=6)
        self.preview_tool_row = tool_row
        tool_row.grid(row=2, column=0, sticky="ew")

        tk.Label(tool_row, text="Zoom", bg="#180B25", fg=TEXT_MUTED, font=("Segoe UI", 9, "bold")).pack(side="left")
        tk.Scale(
            tool_row,
            from_=1.0,
            to=2.5,
            resolution=0.1,
            orient="horizontal",
            variable=self.zoom_var,
            bg="#180B25",
            fg=TEXT_PRIMARY,
            troughcolor="#5F34A8",
            highlightthickness=0,
            length=120,
        ).pack(side="left", padx=(8, 12))

        tk.Label(tool_row, text="Rotacion", bg="#180B25", fg=TEXT_MUTED, font=("Segoe UI", 9, "bold")).pack(side="left")
        rotation_combo = ttk.Combobox(tool_row, textvariable=self.rotation_var, state="readonly", width=7)
        rotation_combo["values"] = ("0 deg", "90 deg", "180 deg", "270 deg")
        rotation_combo.pack(side="left", padx=(8, 12))

        tk.Label(tool_row, text="Temporizador", bg="#180B25", fg=TEXT_MUTED, font=("Segoe UI", 9, "bold")).pack(side="left")
        countdown_combo = ttk.Combobox(tool_row, textvariable=self.countdown_var, state="readonly", width=6)
        countdown_combo["values"] = ("0 s", "3 s", "5 s")
        countdown_combo.pack(side="left", padx=(8, 12))

        ttk.Button(tool_row, text="Abrir fotos", style="Gold.TButton", command=self.open_photos_root).pack(side="right")

        self.preview_canvas = tk.Canvas(preview_card, bg="#0D0914", highlightthickness=0, bd=0)
        self.preview_canvas.grid(row=1, column=0, sticky="nsew")
        self.preview_canvas.bind("<Button-1>", lambda _event: self._focus_student())

        tk.Label(
            capture_page,
            textvariable=self.status_var,
            bg="#1B0F2A",
            fg=SUCCESS,
            anchor="w",
            padx=10,
            pady=8,
            font=("Segoe UI", 10, "bold"),
        ).grid(row=2, column=0, sticky="ew", pady=(10, 0))

        self._build_course_page(course_page)
        self._build_info_page(info_page)

    def _make_card(self, parent: tk.Widget, title: str) -> ttk.Frame:
        card = ttk.Frame(parent, style="Card.TFrame")
        card.pack(fill="x", padx=16, pady=8)
        title_label = tk.Label(card, text=title, bg=CARD_BG, fg=ACCENT_GOLD, font=("Segoe UI", 11, "bold"))
        title_label.pack(
            anchor="w", padx=14, pady=(12, 8)
        )
        card.title_label = title_label  # type: ignore[attr-defined]
        return card

    def _make_session_card(self, parent: tk.Widget) -> None:
        card = self._make_card(parent, "Sesion")
        ttk.Radiobutton(card, text="Modo prueba", variable=self.mode_var, value="test").pack(anchor="w", padx=14, pady=2)
        ttk.Radiobutton(card, text="Modo curso", variable=self.mode_var, value="course").pack(anchor="w", padx=14, pady=2)
        ttk.Label(card, text="Curso", style="Muted.TLabel").pack(anchor="w", padx=14, pady=(10, 2))
        self.course_combo = ttk.Combobox(card, textvariable=self.course_var, state="normal")
        self.course_combo.pack(fill="x", padx=14, pady=(0, 8))
        self.course_combo.bind("<<ComboboxSelected>>", lambda _event: self._handle_course_return())
        self.course_combo.bind("<Return>", self._handle_course_return)
        self.course_combo.bind("<KP_Enter>", self._handle_course_return)
        ttk.Button(card, text="Cargar lista", style="Gold.TButton", command=self.import_roster_file).pack(fill="x", padx=14, pady=(0, 8))
        self.roster_path_label = ttk.Label(card, textvariable=self.roster_path_var, style="Muted.TLabel", wraplength=280, justify="left")
        self.roster_path_label.pack(anchor="w", padx=14, pady=(0, 8))
        ttk.Button(card, text="Iniciar sesion", style="Accent.TButton", command=self.start_session).pack(fill="x", padx=14, pady=(0, 14))

    def _make_roster_card(self, parent: tk.Widget) -> None:
        card = self._make_card(parent, "Lista de alumnos")
        self.roster_status_label = tk.Label(
            card,
            textvariable=self.roster_status_var,
            justify="left",
            anchor="w",
            bg=CARD_BG,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9),
            wraplength=300,
        )
        self.roster_status_label.pack(fill="x", padx=14, pady=(0, 10))

        buttons = tk.Frame(card, bg=CARD_BG)
        buttons.pack(fill="x", padx=14, pady=(0, 6))
        ttk.Button(buttons, text="Anterior", style="Gold.TButton", command=self.prev_roster_student).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(buttons, text="Siguiente", style="Gold.TButton", command=self.next_roster_student).pack(
            side="left", fill="x", expand=True, padx=(8, 0)
        )
        ttk.Button(card, text="Alinear con lista", style="Accent.TButton", command=self.sync_student_with_roster).pack(
            fill="x", padx=14, pady=(0, 14)
        )

    def _make_camera_card(self, parent: tk.Widget) -> None:
        card = self._make_card(parent, "Camara")
        ttk.Label(card, text="Fuente detectada", style="Muted.TLabel").pack(anchor="w", padx=14, pady=(0, 2))
        self.camera_combo = ttk.Combobox(card, textvariable=self.camera_var, state="readonly")
        self.camera_combo.pack(fill="x", padx=14, pady=(0, 10))
        self.camera_combo.bind("<<ComboboxSelected>>", lambda _event: self.change_camera())
        ttk.Checkbutton(card, text="Ayuda visual de rostro", variable=self.face_guide_var).pack(anchor="w", padx=14, pady=2)
        ttk.Checkbutton(card, text="Espejo horizontal", variable=self.mirror_var).pack(anchor="w", padx=14, pady=2)
        ttk.Checkbutton(card, text="Mostrar guia", variable=self.frame_guide_var).pack(anchor="w", padx=14, pady=2)
        ttk.Checkbutton(card, text="Recortar tipo credencial 3:4", variable=self.crop_portrait_var).pack(anchor="w", padx=14, pady=2)
        ttk.Button(card, text="Abrir carpeta fotos", style="Gold.TButton", command=self.open_photos_root).pack(fill="x", padx=14, pady=(10, 14))

    def _make_student_card(self, parent: tk.Widget) -> None:
        card = self._make_card(parent, "Estudiante y captura")
        self.student_card_title_var = tk.StringVar(value="Estudiante y captura")
        card.title_label.configure(textvariable=self.student_card_title_var)  # type: ignore[attr-defined]

        self.student_manual_frame = tk.Frame(card, bg=CARD_BG)
        self.student_manual_frame.pack(fill="x", padx=14, pady=(0, 10))
        ttk.Label(self.student_manual_frame, text="Nombre actual", style="Muted.TLabel").pack(anchor="w", pady=(0, 2))
        self.student_entry = ttk.Entry(self.student_manual_frame, textvariable=self.student_var)
        self.student_entry.pack(fill="x")
        self.student_entry.bind("<Return>", self._handle_student_return)
        self.student_entry.bind("<KP_Enter>", self._handle_student_return)

        buttons = tk.Frame(card, bg=CARD_BG)
        buttons.pack(fill="x", padx=14, pady=(0, 6))
        ttk.Button(buttons, text="Capturar", style="Accent.TButton", command=self.capture_photo).pack(side="left", fill="x", expand=True)
        self.student_clear_button = ttk.Button(buttons, text="Limpiar", style="Gold.TButton", command=lambda: self.student_var.set(""))
        self.student_clear_button.pack(side="left", fill="x", expand=True, padx=(8, 0))

        buttons2 = tk.Frame(card, bg=CARD_BG)
        buttons2.pack(fill="x", padx=14, pady=(0, 14))
        ttk.Button(buttons2, text="Volver atras y reintentar", style="Gold.TButton", command=self.retake_last).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(buttons2, text="Cerrar sesion", style="Danger.TButton", command=self.close_session).pack(side="left", fill="x", expand=True, padx=(8, 0))

    def _make_recent_card(self, parent: tk.Widget) -> None:
        card = self._make_card(parent, "Recientes")
        self.recent_label = tk.Label(
            card,
            textvariable=self.recent_var,
            justify="left",
            anchor="w",
            bg=CARD_BG,
            fg=TEXT_MUTED,
            font=("Consolas", 9),
            wraplength=300,
        )
        self.recent_label.pack(fill="x", padx=14, pady=(0, 14))

    def _build_course_page(self, parent: tk.Widget) -> None:
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        header = tk.Frame(parent, bg=INFO_BG, padx=18, pady=16)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        self.course_total_var = tk.StringVar(value="0 alumnos")
        self.course_current_var = tk.StringVar(value="Sin curso activo")
        self.course_progress_var = tk.IntVar(value=0)
        self.course_progress_text_var = tk.StringVar(value="0 capturados")

        tk.Label(
            header,
            text="Lista completa del curso",
            bg=INFO_BG,
            fg=ACCENT_GOLD,
            font=("Segoe UI", 18, "bold"),
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            header,
            textvariable=self.course_current_var,
            bg=INFO_BG,
            fg=TEXT_PRIMARY,
            font=("Segoe UI", 10, "bold"),
        ).grid(row=1, column=0, sticky="w", pady=(6, 2))

        progress_row = tk.Frame(header, bg=INFO_BG)
        progress_row.grid(row=2, column=0, sticky="ew", pady=(10, 4))
        progress_row.grid_columnconfigure(0, weight=1)
        ttk.Progressbar(progress_row, maximum=100, variable=self.course_progress_var).grid(row=0, column=0, sticky="ew")
        tk.Label(
            progress_row,
            textvariable=self.course_progress_text_var,
            bg=INFO_BG,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9),
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        tk.Label(
            header,
            textvariable=self.course_total_var,
            bg=INFO_BG,
            fg=TEXT_MUTED,
            font=("Segoe UI", 10),
        ).grid(row=3, column=0, sticky="w", pady=(2, 0))

        table_frame = tk.Frame(parent, bg=INFO_BG, padx=18, pady=0)
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.course_tree = tk.Text(
            table_frame,
            wrap="none",
            bg="#1D102A",
            fg=TEXT_PRIMARY,
            insertbackground=ACCENT_GOLD,
            relief="flat",
            borderwidth=0,
            padx=12,
            pady=10,
            font=("Consolas", 10),
            height=16,
        )
        self.course_tree_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.course_tree.yview)
        self.course_tree.configure(yscrollcommand=self.course_tree_scroll.set)
        self.course_tree.grid(row=0, column=0, sticky="nsew")
        self.course_tree_scroll.grid(row=0, column=1, sticky="ns")

        self.course_tree.tag_configure("done", background="#163022", foreground="#CFF9E0")
        self.course_tree.tag_configure("current", background="#4E390B", foreground="#FFF0B0")
        self.course_tree.tag_configure("pending", background="#241033", foreground=TEXT_MUTED)
        self.course_tree.tag_configure("empty", background="#241033", foreground="#A89BBB")
        self.course_tree.tag_configure("header", foreground=ACCENT_GOLD)

        footer = tk.Frame(parent, bg=INFO_BG, padx=18, pady=14)
        footer.grid(row=2, column=0, sticky="ew")
        footer.grid_columnconfigure(0, weight=1)
        ttk.Button(footer, text="Actualizar lista", style="Gold.TButton", command=self._refresh_course_view).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Button(footer, text="Alinear con lista", style="Accent.TButton", command=self.sync_student_with_roster).grid(
            row=0, column=1, sticky="e", padx=(8, 0)
        )

        self._refresh_course_view(force=True)

    def _bind_sidebar_mousewheel(self, canvas: tk.Canvas) -> None:
        def _is_within_sidebar(widget: Optional[tk.Misc]) -> bool:
            current = widget
            while current is not None:
                if current == self.sidebar_content:
                    return True
                try:
                    current = current.nametowidget(current.winfo_parent())
                except Exception:
                    return False
            return False

        def _on_mousewheel(event: tk.Event) -> str:
            if not _is_within_sidebar(getattr(event, "widget", None)):
                return ""
            delta = getattr(event, "delta", 0)
            if delta:
                canvas.yview_scroll(int(-1 * (delta / 120)), "units")
                return "break"
            return ""

        self.root.bind_all("<MouseWheel>", _on_mousewheel, add="+")
        self.root.bind_all("<Shift-MouseWheel>", _on_mousewheel, add="+")

    def _on_notebook_tab_changed(self, _event: Optional[tk.Event] = None) -> None:
        if self.notebook is None:
            return
        try:
            tab_name = self.notebook.tab(self.notebook.select(), "text")
        except Exception:
            tab_name = ""
        self.logger.info("Notebook tab changed: %s", tab_name)
        if tab_name == "Curso":
            self.root.after_idle(lambda: self._refresh_course_view(force=True))

    def _course_tab_active(self) -> bool:
        if self.notebook is None:
            return False
        try:
            return self.notebook.tab(self.notebook.select(), "text") == "Curso"
        except Exception:
            return False

    def _schedule_responsive_layout(self, _event: Optional[tk.Event] = None) -> None:
        if self._course_tab_active():
            return
        if self._responsive_job is not None:
            try:
                self.root.after_cancel(self._responsive_job)
            except Exception:
                pass
        self._responsive_job = self.root.after(120, self._sync_responsive_layout)

    def _sync_responsive_layout(self, _event: Optional[tk.Event] = None) -> None:
        self._responsive_job = None
        if self.root is None:
            return

        width = max(0, self.root.winfo_width())
        height = max(0, self.root.winfo_height())
        if width <= 1 or height <= 1:
            return

        if self.notebook is not None:
            try:
                current_tab = self.notebook.tab(self.notebook.select(), "text")
            except Exception:
                current_tab = ""
            if current_tab == "Curso":
                return

        width_profile = 0 if width < 1300 else 1 if width < 1600 else 2
        height_profile = 0 if height < 820 else 1 if height < 950 else 2
        layout_profile = (width_profile, height_profile)
        if self._last_layout_profile == layout_profile:
            return
        self._last_layout_profile = layout_profile

        sidebar_width = 300 if width_profile == 0 else 340 if width_profile == 1 else 380
        if self.left_shell is not None:
            self.left_shell.configure(width=sidebar_width)

        title_size = 18 if width_profile == 0 else 20 if width_profile == 1 else 22
        subtitle_size = 9 if width_profile == 0 else 10
        session_size = 11 if width_profile == 0 else 13
        info_size = 10 if width_profile == 0 else 11
        tree_height = 11 if height_profile == 0 else 14 if height_profile == 1 else 16

        self.style.configure("Title.TLabel", font=("Segoe UI", title_size, "bold"))
        self.style.configure("Muted.TLabel", font=("Segoe UI", 8 if width < 1300 else 9))
        self.style.configure("TRadiobutton", font=("Segoe UI", 9 if width < 1300 else 10))
        self.style.configure("TCheckbutton", font=("Segoe UI", 9 if width < 1300 else 10))
        self.style.configure("Treeview", rowheight=24 if width < 1300 else 28)
        self.style.configure("Treeview.Heading", font=("Segoe UI", 9 if width < 1300 else 10, "bold"))

        if self.sidebar_subtitle_label is not None:
            self.sidebar_subtitle_label.configure(font=("Segoe UI", subtitle_size, "bold"))
            self.sidebar_subtitle_label.configure(wraplength=max(220, sidebar_width - 40))
        if hasattr(self, "roster_path_label"):
            try:
                self.roster_path_label.configure(wraplength=max(220, sidebar_width - 50))
            except Exception:
                pass
        if hasattr(self, "roster_status_label"):
            try:
                self.roster_status_label.configure(wraplength=max(220, sidebar_width - 50))
            except Exception:
                pass
        if hasattr(self, "recent_label"):
            try:
                self.recent_label.configure(wraplength=max(220, sidebar_width - 50))
            except Exception:
                pass

        if self.session_title_label is not None:
            self.session_title_label.configure(font=("Segoe UI", session_size, "bold"))

        if self.info_text is not None:
            self.info_text.configure(font=("Segoe UI", info_size))

        if self.course_tree is not None:
            course_font_size = 9 if width < 1300 else 10
            self.course_tree.configure(height=tree_height, font=("Consolas", course_font_size))

        if hasattr(self, "preview_camera_combo"):
            try:
                self.preview_camera_combo.configure(width=26 if width < 1300 else 34)
            except Exception:
                pass
        if hasattr(self, "camera_combo"):
            try:
                self.camera_combo.configure(width=28 if width < 1300 else 34)
            except Exception:
                pass

        # Keep the preview toolbar usable when the window becomes narrow.
        if self.preview_toolbar is not None:
            for child in self.preview_toolbar.winfo_children():
                try:
                    child.configure(font=("Segoe UI", 9 if width < 1300 else 10))
                except Exception:
                    pass

        if self.preview_tool_row is not None:
            for child in self.preview_tool_row.winfo_children():
                try:
                    child.configure(font=("Segoe UI", 8 if width < 1300 else 9))
                except Exception:
                    pass

    def _build_info_page(self, parent: tk.Widget) -> None:
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        info = tk.Text(
            parent,
            wrap="word",
            bg=INFO_BG,
            fg=TEXT_PRIMARY,
            font=("Segoe UI", 11),
            relief="flat",
            padx=18,
            pady=18,
            insertbackground=ACCENT_GOLD,
        )
        self.info_text = info
        info.grid(row=0, column=0, sticky="nsew")
        info.insert(
            "1.0",
            (
                "CastelCredCam Studio\n\n"
                "Flujo recomendado\n"
                "1. Carga una lista de alumnos en Excel o CSV.\n"
                "2. Elige el curso en el campo Curso.\n"
                "3. Inicia la sesion.\n"
                "4. La app te va mostrando el siguiente alumno y su RUT.\n"
                "5. Usa Enter o el boton Capturar para avanzar automaticamente.\n"
                "6. Revisa la carpeta fotos mientras avanzas.\n\n"
                "Atajos\n"
                "- Enter: capturar foto\n"
                "- Ctrl+Izquierda / Ctrl+Derecha: mover alumno en la lista\n"
                "- C: siguiente camara\n"
                "- V: voltear horizontalmente\n"
                "- R: activar o desactivar ayuda de rostro\n"
                "- G: activar o desactivar guia de encuadre\n"
                "- X: activar o desactivar recorte automatico\n"
                "- O: abrir carpeta de fotos\n"
                "- F: enfocar nombre del estudiante\n"
                "- Escape: limpiar nombre actual\n\n"
                "Funciones utiles de camara\n"
                "- Zoom digital\n"
                "- Rotacion 0, 90, 180 y 270 grados\n"
                "- Temporizador de captura 0, 3 o 5 segundos\n"
                "- Guia de encuadre\n"
                "- Ayuda visual de rostro\n"
                "- Recorte automatico tipo credencial 3:4\n"
                "- Preview del recorte siguiendo el rostro cuando esta disponible\n"
                "- Selector rapido de camara dentro del preview\n\n"
                "Consejos\n"
                "- Usa buena luz frontal.\n"
                "- Manten la camara fija en tripode.\n"
                "- Si usas Recortar, procura que el rostro quede visible y centrado.\n"
                "- Haz una sesion de prueba antes de un curso real.\n"
                "- Verifica que el alumno actual sea el correcto antes de capturar.\n"
            ),
        )
        info.configure(state="disabled")

    def _extract_course_label(self, sheet) -> str:
        raw = sheet["A1"].value if sheet.max_row else None
        if isinstance(raw, str) and raw.strip().upper().startswith("LISTA ALUMNOS POR CURSO"):
            return raw.strip().replace("LISTA ALUMNOS POR CURSO", "", 1).strip()
        return str(sheet.title).strip()

    def _extract_students_from_rows(self, rows) -> list[RosterStudent]:
        header_index = None
        header_map: dict[str, int] = {}
        wanted = {
            "rut",
            "apellido paterno",
            "apellido materno",
            "primer nombre",
            "segundo nombre",
            "nombres",
        }

        for row_index, row in enumerate(rows[:20]):
            values = ["" if value is None else str(value).strip() for value in row]
            normalized = [_normalize_key(value) for value in values]
            if "rut" not in normalized:
                continue
            if not any(key in normalized for key in ("apellido paterno", "apellido materno", "nombres")):
                continue
            header_index = row_index
            for col_index, value in enumerate(normalized):
                if value in wanted:
                    header_map[value] = col_index
            break

        if header_index is None or "rut" not in header_map:
            return []

        students: list[RosterStudent] = []
        for row in rows[header_index + 1 :]:
            values = ["" if value is None else str(value).strip() for value in row]
            if not any(values):
                continue

            rut = values[header_map["rut"]] if "rut" in header_map and header_map["rut"] < len(values) else ""
            apellido_paterno = values[header_map.get("apellido paterno", -1)] if header_map.get("apellido paterno", -1) >= 0 and header_map.get("apellido paterno", -1) < len(values) else ""
            apellido_materno = values[header_map.get("apellido materno", -1)] if header_map.get("apellido materno", -1) >= 0 and header_map.get("apellido materno", -1) < len(values) else ""
            primer_nombre = values[header_map.get("primer nombre", -1)] if header_map.get("primer nombre", -1) >= 0 and header_map.get("primer nombre", -1) < len(values) else ""
            segundo_nombre = values[header_map.get("segundo nombre", -1)] if header_map.get("segundo nombre", -1) >= 0 and header_map.get("segundo nombre", -1) < len(values) else ""
            nombres = values[header_map.get("nombres", -1)] if header_map.get("nombres", -1) >= 0 and header_map.get("nombres", -1) < len(values) else ""

            if not nombres:
                names_parts = [part for part in (primer_nombre, segundo_nombre) if part]
                nombres = " ".join(names_parts)

            if not rut or not (apellido_paterno or apellido_materno or nombres):
                continue

            students.append(
                RosterStudent(
                    rut=rut,
                    apellido_paterno=apellido_paterno,
                    apellido_materno=apellido_materno,
                    nombres=nombres,
                )
            )

        return students

    def _load_roster_csv(self, path: Path) -> dict[str, list[RosterStudent]]:
        roster: dict[str, list[RosterStudent]] = {}
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if not row:
                    continue
                course = (row.get("course") or row.get("curso") or "").strip()
                rut = (row.get("rut") or row.get("RUT") or "").strip()
                apellido_paterno = (row.get("apellido paterno") or row.get("apellido_paterno") or row.get("paterno") or "").strip()
                apellido_materno = (row.get("apellido materno") or row.get("apellido_materno") or row.get("materno") or "").strip()
                nombres = (row.get("nombres") or row.get("nombre") or "").strip()
                if not nombres:
                    primer_nombre = (row.get("primer nombre") or row.get("primer_nombre") or "").strip()
                    segundo_nombre = (row.get("segundo nombre") or row.get("segundo_nombre") or "").strip()
                    nombres = " ".join(part for part in (primer_nombre, segundo_nombre) if part)
                if not course or not rut or not apellido_paterno:
                    continue
                roster.setdefault(course, []).append(
                    RosterStudent(
                        rut=rut,
                        apellido_paterno=apellido_paterno,
                        apellido_materno=apellido_materno,
                        nombres=nombres,
                    )
                )
        for students in roster.values():
            students.sort(key=lambda item: _normalize_key(item.display_name))
        return roster

    def _load_roster_excel(self, path: Path) -> dict[str, list[RosterStudent]]:
        workbook = load_workbook(path, data_only=True, read_only=True)
        roster: dict[str, list[RosterStudent]] = {}
        for sheet in workbook.worksheets:
            if _normalize_key(sheet.title) == "resumen":
                continue
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            course_label = self._extract_course_label(sheet)
            students = self._extract_students_from_rows(rows)
            if students:
                roster[course_label] = students
        for students in roster.values():
            students.sort(key=lambda item: _normalize_key(item.display_name))
        return roster

    def _load_roster_source(self, path: Path) -> dict[str, list[RosterStudent]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._load_roster_csv(path)
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return self._load_roster_excel(path)
        raise ValueError("Solo se aceptan archivos CSV o Excel.")

    def import_roster_file(self) -> None:
        path_str = filedialog.askopenfilename(
            title="Selecciona la lista de alumnos",
            filetypes=[
                ("Excel o CSV", "*.xlsx *.xlsm *.xltx *.xltm *.csv"),
                ("Excel", "*.xlsx *.xlsm *.xltx *.xltm"),
                ("CSV", "*.csv"),
                ("Todos los archivos", "*.*"),
            ],
        )
        if not path_str:
            return

        path = Path(path_str)
        self.logger.info("Import roster requested: %s", path)
        try:
            roster_map = self._load_roster_source(path)
        except Exception as exc:
            self.logger.exception("Roster import failed: %s", exc)
            messagebox.showerror(APP_TITLE, f"No se pudo cargar la lista.\n\n{exc}")
            return

        if not roster_map:
            self.logger.warning("Roster file contained no students: %s", path)
            messagebox.showwarning(APP_TITLE, "No encontré cursos con alumnos dentro del archivo seleccionado.")
            return

        self.roster_map = roster_map
        self.roster_lookup = {_normalize_key(course): course for course in roster_map}
        for course_name in roster_map:
            self.roster_preview_index.setdefault(course_name, 0)
        course_names = sorted(roster_map.keys(), key=_normalize_key)
        self.course_combo["values"] = course_names
        self.roster_path_var.set(f"Lista cargada: {path.name}")
        self.logger.info("Roster loaded: %s courses=%s", path.name, len(roster_map))

        current_course = self.course_var.get().strip()
        resolved = self._resolve_roster_course(current_course) if current_course else ""
        if not resolved:
            resolved = course_names[0]
            self.course_var.set(resolved)

        self._update_roster_preview(resolved)
        self._refresh_course_view(force=True)
        self.status_var.set(f"Lista lista: {path.name}. Elige curso y arranca la sesion.")

    def _resolve_roster_course(self, course_name: str) -> str:
        if not course_name:
            return ""
        normalized = _normalize_key(course_name)
        if normalized in self.roster_lookup:
            return self.roster_lookup[normalized]
        for key in self.roster_map:
            if _normalize_key(key) == normalized:
                return key
        return ""

    def _update_roster_preview(self, course_name: str = "") -> None:
        if not self.roster_map:
            self.roster_status_var.set("Carga una lista para habilitar captura secuencial.")
            return

        resolved = self._resolve_roster_course(course_name) if course_name else ""
        if not resolved and len(self.roster_map) == 1:
            resolved = next(iter(self.roster_map))

        if not resolved:
            total = sum(len(items) for items in self.roster_map.values())
            self.roster_status_var.set(f"Lista cargada con {len(self.roster_map)} cursos y {total} alumnos.\nElige un curso.")
            return

        students = self.roster_map.get(resolved, [])
        if not students:
            self.roster_status_var.set(f"{resolved}\nSin alumnos cargados.")
            self._refresh_course_view()
            return

        preview_index = self.roster_preview_index.get(resolved, 0)
        preview_index = max(0, min(preview_index, len(students) - 1))
        self.roster_preview_index[resolved] = preview_index
        next_student = students[preview_index]

        self.roster_status_var.set(
            f"{resolved}\nAlumno {preview_index + 1} de {len(students)}\nSiguiente: {next_student.display_name}\nRUT: {next_student.rut}"
        )
        self._refresh_course_view(force=True)

    def _sync_session_student_from_roster(self) -> None:
        if self.session is None or not self.session.has_roster:
            return
        student = self.session.current_roster_student()
        if student is None:
            self.student_var.set("")
            self._update_roster_session_label()
            self._refresh_student_card_mode()
            return
        self.student_var.set(student.display_name)
        self._update_roster_session_label()
        self._refresh_student_card_mode()
        self._refresh_course_view()

    def _advance_roster_past_completed(self) -> None:
        if self.session is None or not self.session.has_roster:
            return
        while True:
            student = self.session.current_roster_student()
            if student is None:
                return
            if not self._student_is_completed(student):
                return
            if self.session.roster_index >= self.session.roster_total:
                return
            self.session.advance_roster()

    def _refresh_student_card_mode(self) -> None:
        manual_mode = self.session is None or not self.session.has_roster
        if self.student_card_title_var is not None:
            self.student_card_title_var.set("Estudiante y captura" if manual_mode else "Captura")
        if self.student_manual_frame is not None:
            if manual_mode:
                if not self.student_manual_frame.winfo_ismapped():
                    self.student_manual_frame.pack(fill="x", padx=14, pady=(0, 10))
            else:
                self.student_manual_frame.pack_forget()
        if self.student_clear_button is not None:
            if manual_mode:
                if not self.student_clear_button.winfo_ismapped():
                    self.student_clear_button.pack(side="left", fill="x", expand=True, padx=(8, 0))
            else:
                self.student_clear_button.pack_forget()
        if self.student_entry is not None and not manual_mode:
            self.student_entry.state(["disabled"])
        elif self.student_entry is not None:
            self.student_entry.state(["!disabled"])

    def _update_roster_session_label(self) -> None:
        if self.session is None or not self.session.has_roster:
            self.roster_status_var.set("Lista cargada.\nLa captura secuencial queda lista al iniciar la sesión.")
            self._refresh_course_view()
            return
        student = self.session.current_roster_student()
        if student is None:
            self.roster_status_var.set(f"{self.session.course_display}\nLista completa.\nRevisa las capturas finales.")
            return
        self.roster_status_var.set(
            f"{self.session.course_display}\nActual: {student.display_name}\nRUT: {student.rut}\nPendientes: {self.session.roster_remaining}"
        )
        self._refresh_course_view()

    def _active_roster_course(self) -> str:
        if self.session is not None and self.session.has_roster:
            return self.session.course_display
        resolved = self._resolve_roster_course(self.course_var.get().strip())
        if resolved:
            return resolved
        if len(self.roster_map) == 1:
            return next(iter(self.roster_map))
        return ""

    def _course_csv_path(self, course_name: str) -> Path:
        course_slug = sanitize_folder_name(course_name)
        return APP_ROOT / PHOTOS_DIRNAME / course_slug / CSV_FILENAME

    def _records_for_course(self, course_name: str) -> list[PhotoRecord]:
        if self.session is not None and self.session.course_display == course_name:
            return list(self.session.records)
        csv_path = self._course_csv_path(course_name)
        if not csv_path.exists():
            return []
        return load_existing_records(csv_path)

    def _student_is_completed(self, student: RosterStudent, records: Optional[list[PhotoRecord]] = None) -> bool:
        if records is None:
            if self.session is not None and self.session.has_roster:
                records = self.session.records
            else:
                return False
        student_name_key = _normalize_key(student.display_name)
        student_rut_key = _normalize_rut_key(student.rut)
        for record in records:
            if student_rut_key and _normalize_rut_key(getattr(record, "rut", "")) == student_rut_key:
                return True
            if _normalize_key(record.student_name) == student_name_key:
                return True
        return False

    def _refresh_course_view(self, force: bool = False) -> None:
        if self.course_tree is None:
            return
        if not force and not self._course_tab_active():
            return

        start = datetime.now()
        self.logger.debug("Refreshing course view. force=%s", force)

        self.course_tree.configure(state="normal")
        self.course_tree.delete("1.0", "end")

        course_name = self._active_roster_course()
        if not self.roster_map:
            self.course_current_var.set("Carga una lista para ver el curso completo.")
            self.course_total_var.set("0 alumnos")
            self.course_progress_text_var.set("0 capturados")
            self.course_progress_var.set(0)
            self.course_tree.insert("end", "Vacío\nCarga una lista primero\n", ("empty",))
            self.course_tree.configure(state="disabled")
            return

        if not course_name:
            total = sum(len(items) for items in self.roster_map.values())
            self.course_current_var.set(f"{len(self.roster_map)} cursos cargados | {total} alumnos")
            self.course_total_var.set("Selecciona un curso para ver el detalle.")
            self.course_progress_text_var.set("Sin curso activo")
            self.course_progress_var.set(0)
            self.course_tree.insert("end", "Selecciona un curso en el campo Curso\n", ("empty",))
            self.course_tree.configure(state="disabled")
            return

        students = self.roster_map.get(course_name, [])
        total = len(students)
        course_records = self._records_for_course(course_name)
        completed = sum(1 for student in students if self._student_is_completed(student, course_records))
        remaining = max(0, total - completed)
        current_student = None
        if self.session is not None and self.session.has_roster:
            current_student = self.session.current_roster_student()
        elif students:
            preview_index = self.roster_preview_index.get(course_name, 0)
            preview_index = max(0, min(preview_index, len(students) - 1))
            current_student = students[preview_index]

        self.course_current_var.set(f"{course_name} | {completed} capturados | {remaining} pendientes")
        self.course_total_var.set(f"{total} alumnos en la nómina")
        self.course_progress_text_var.set(f"{completed} capturados de {total}")
        self.course_progress_var.set(0 if total == 0 else int((completed / total) * 100))

        if not students:
            self.course_tree.insert("end", "Vacío\nSin alumnos cargados\n", ("empty",))
            self.course_tree.configure(state="disabled")
            return

        current_key = _normalize_rut_key(current_student.rut) if current_student is not None else ""
        self.course_tree.insert("end", f"Estado    Alumno{' ' * 41}RUT\n", ("header",))
        for index, student in enumerate(students, start=1):
            completed_flag = self._student_is_completed(student)
            if completed_flag:
                status = "Hecho"
                tag = "done"
            elif current_student is not None and _normalize_rut_key(student.rut) == current_key:
                status = "Actual"
                tag = "current"
            else:
                status = "Pendiente"
                tag = "pending"
            line = f"{status:<9} {index:03d}. {student.display_name} | {student.rut}\n"
            self.course_tree.insert("end", line, (tag,))
        elapsed = (datetime.now() - start).total_seconds()
        self.logger.debug("Course view refreshed in %.3fs with %s students.", elapsed, len(students))
        self.course_tree.configure(state="disabled")

    def next_roster_student(self) -> None:
        if self.session is not None and self.session.has_roster:
            self.session.advance_roster()
            self._sync_session_student_from_roster()
            self.status_var.set("Siguiente alumno.")
            self._refresh_course_view()
            return

        course_name = self._resolve_roster_course(self.course_var.get().strip())
        if not course_name and len(self.roster_map) == 1:
            course_name = next(iter(self.roster_map))
        if not course_name:
            return
        students = self.roster_map.get(course_name, [])
        if not students:
            return
        current_index = self.roster_preview_index.get(course_name, 0)
        current_index = min(current_index + 1, len(students) - 1)
        self.roster_preview_index[course_name] = current_index
        self._update_roster_preview(course_name)
        self.status_var.set(f"Vista previa en {current_index + 1} de {len(students)}.")
        self._refresh_course_view()

    def prev_roster_student(self) -> None:
        if self.session is not None and self.session.has_roster:
            self.session.retreat_roster()
            self._sync_session_student_from_roster()
            self.status_var.set("Alumno anterior.")
            self._refresh_course_view()
            return

        course_name = self._resolve_roster_course(self.course_var.get().strip())
        if not course_name and len(self.roster_map) == 1:
            course_name = next(iter(self.roster_map))
        if not course_name:
            return
        students = self.roster_map.get(course_name, [])
        if not students:
            return
        current_index = self.roster_preview_index.get(course_name, 0)
        current_index = max(0, current_index - 1)
        self.roster_preview_index[course_name] = current_index
        self._update_roster_preview(course_name)
        self.status_var.set(f"Vista previa en {current_index + 1} de {len(students)}.")
        self._refresh_course_view()

    def sync_student_with_roster(self) -> None:
        if self.session is not None and self.session.has_roster:
            self._sync_session_student_from_roster()
            self.status_var.set("Alumno alineado con la lista.")
            self._refresh_course_view()
            return

        course_name = self._resolve_roster_course(self.course_var.get().strip())
        if not course_name and len(self.roster_map) == 1:
            course_name = next(iter(self.roster_map))
        if not course_name:
            messagebox.showinfo(APP_TITLE, "Carga una lista y elige un curso primero.")
            return
        self.roster_preview_index[course_name] = 0
        self._update_roster_preview(course_name)
        self.status_var.set("Vista previa alineada con el primer alumno.")
        self._refresh_course_view()

    def _handle_student_return(self, _event=None):
        self.capture_photo()
        return "break"

    def _handle_course_return(self, _event=None):
        self._update_roster_preview(self.course_var.get())
        self._refresh_course_view()
        return "break"

    def _handle_global_return(self, _event=None):
        widget = self.root.focus_get()
        if widget is self.student_entry:
            self.capture_photo()
            return "break"
        if widget is self.course_combo:
            self._update_roster_preview(self.course_var.get())
            return "break"
        if widget is not None:
            widget_class = widget.winfo_class()
            if widget_class in {"TButton", "Button"}:
                return None
        self.capture_photo()
        return "break"

    def _load_camera_choices(self) -> None:
        values = []
        for index, label, backend_id, backend_name, alias in self.available_cameras:
            values.append(f"{alias} | idx {index} | {backend_name} | {backend_key_from_id(backend_id)}")

        self.camera_combo["values"] = values
        self.preview_camera_combo["values"] = values
        if not values:
            self.status_var.set("No se detectaron camaras compatibles.")
            return

        preferred = 0
        remembered_index, remembered_backend = load_last_camera(APP_ROOT)
        if remembered_index is not None and remembered_backend is not None:
            for pos, (index, _label, backend_id, _backend_name, _alias) in enumerate(self.available_cameras):
                if index == remembered_index and backend_key_from_id(backend_id) == remembered_backend:
                    preferred = pos
                    break

        self.camera_combo.current(preferred)
        self.preview_camera_combo.current(preferred)
        self.change_camera()

    def change_camera(self, from_preview: bool = False) -> None:
        selection = self.preview_camera_combo.current() if from_preview else self.camera_combo.current()
        if selection < 0 or selection >= len(self.available_cameras):
            return

        index, _label, backend_id, backend_name, alias = self.available_cameras[selection]
        self.camera_combo.current(selection)
        self.preview_camera_combo.current(selection)
        self.current_camera_index = index
        self.current_backend_id = backend_id
        self.current_backend_name = backend_name
        self.current_camera_alias = alias
        save_last_camera(APP_ROOT, index, backend_key_from_id(backend_id))
        self._open_selected_camera()

    def cycle_camera(self) -> None:
        if not self.available_cameras:
            return
        current = self.preview_camera_combo.current()
        next_index = 0 if current < 0 else (current + 1) % len(self.available_cameras)
        self.preview_camera_combo.current(next_index)
        self.change_camera(from_preview=True)

    def _open_selected_camera(self) -> None:
        self._release_capture()
        if self.current_camera_index is None or self.current_backend_id is None:
            return

        cap = open_camera(self.current_camera_index, self.current_backend_id)
        if not cap.isOpened():
            self.status_var.set("No se pudo abrir la camara seleccionada.")
            return

        requested_width, requested_height = configure_capture(cap)
        self.capture = cap
        actual_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH) or requested_width or 0)
        actual_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT) or requested_height or 0)
        self.status_var.set(f"Camara activa: {self.current_camera_alias} ({actual_width}x{actual_height})")
        self._schedule_preview()

    def _schedule_preview(self, delay_ms: int = 30) -> None:
        if self.preview_job is not None:
            self.root.after_cancel(self.preview_job)
        self.preview_job = self.root.after(delay_ms, self._update_preview)

    def _update_preview(self) -> None:
        self.preview_job = None
        if self.capture is None:
            self._show_placeholder("Selecciona una camara para empezar.")
            self._schedule_preview(250)
            return

        if self.notebook is not None:
            try:
                current_tab = self.notebook.tab(self.notebook.select(), "text")
            except Exception:
                current_tab = ""
            if current_tab != "Captura":
                self._schedule_preview(500)
                return

        ok, frame = self.capture.read()
        if not ok or frame is None:
            self._show_placeholder("No se pudo leer la camara.\nRevisa la conexion o cambia de fuente.")
            self._schedule_preview(250)
            return

        if self.mirror_var.get():
            frame = cv2.flip(frame, 1)
        transformed = self._apply_transformations(frame)
        self.frame_counter += 1
        self.current_face_box = self._detect_primary_face(transformed)
        if self.crop_portrait_var.get():
            next_crop_box = self._compute_portrait_crop_box(
                transformed.shape[1], transformed.shape[0], self.current_face_box, self.current_eye_centers
            )
            smoothed_box = self._smooth_crop_box(next_crop_box)
            self.current_crop_box = self._constrain_crop_box(
                smoothed_box,
                transformed.shape[1],
                transformed.shape[0],
            )
            portrait_frame = self._crop_frame_with_box(transformed, self.current_crop_box)
            self.current_frame = portrait_frame.copy()
            preview_frame = self._decorate_frame(portrait_frame.copy())
        else:
            self.current_crop_box = None
            self.stable_crop_box = None
            self.current_frame = transformed.copy()
            preview_frame = self._decorate_frame(transformed.copy())
        display_frame = self._fit_frame_to_preview(preview_frame)

        rgb = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
        image = Image.fromarray(rgb)
        self.tk_image = ImageTk.PhotoImage(image=image)
        self.preview_canvas.delete("all")
        canvas_w = max(320, self.preview_canvas.winfo_width())
        canvas_h = max(240, self.preview_canvas.winfo_height())
        self.preview_canvas.create_image(canvas_w // 2, canvas_h // 2, image=self.tk_image, anchor="center")
        self._schedule_preview()

    def _fit_frame_to_preview(self, frame):
        target_w = max(320, self.preview_canvas.winfo_width() - 20)
        target_h = max(240, self.preview_canvas.winfo_height() - 20)
        src_h, src_w = frame.shape[:2]
        scale = min(target_w / src_w, target_h / src_h)
        new_w = max(1, int(src_w * scale))
        new_h = max(1, int(src_h * scale))
        return cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_AREA)

    def _show_placeholder(self, text: str) -> None:
        self.preview_canvas.delete("all")
        canvas_w = max(320, self.preview_canvas.winfo_width())
        canvas_h = max(240, self.preview_canvas.winfo_height())
        self.preview_canvas.create_text(
            canvas_w // 2,
            canvas_h // 2,
            text=text,
            fill=TEXT_MUTED,
            font=("Segoe UI", 16, "bold"),
            justify="center",
        )

    def _expand_box(
        self,
        box: tuple[int, int, int, int],
        width: int,
        height: int,
        scale: float = 1.0,
        pad_x: int = 0,
        pad_y: int = 0,
    ) -> tuple[int, int, int, int]:
        x, y, w, h = box
        cx = x + w / 2
        cy = y + h / 2
        new_w = max(1, int(w * scale) + pad_x * 2)
        new_h = max(1, int(h * scale) + pad_y * 2)
        x1 = max(0, int(cx - new_w / 2))
        y1 = max(0, int(cy - new_h / 2))
        x2 = min(width, x1 + new_w)
        y2 = min(height, y1 + new_h)
        return x1, y1, x2, y2

    def _clip_box(self, box: tuple[int, int, int, int], width: int, height: int) -> tuple[int, int, int, int]:
        x1, y1, x2, y2 = box
        x1 = max(0, min(x1, width - 1))
        y1 = max(0, min(y1, height - 1))
        x2 = max(x1 + 1, min(x2, width))
        y2 = max(y1 + 1, min(y2, height))
        return x1, y1, x2, y2

    def _detect_eyes_in_face(self, gray: np.ndarray, face_box: tuple[int, int, int, int]) -> list[tuple[int, int]]:
        x, y, w, h = face_box
        roi = gray[y : y + h, x : x + w]
        if roi.size == 0:
            return []
        eye_min_w = max(12, w // 10)
        eye_min_h = max(8, h // 12)
        eye_centers: list[tuple[int, int]] = []
        for cascade in self.eye_cascades:
            if cascade.empty():
                continue
            eyes = cascade.detectMultiScale(
                roi,
                scaleFactor=1.06,
                minNeighbors=4,
                minSize=(eye_min_w, eye_min_h),
            )
            for ex, ey, ew, eh in eyes:
                eye_centers.append((x + ex + ew // 2, y + ey + eh // 2))
            if len(eye_centers) >= 2:
                break
        eye_centers.sort(key=lambda pt: pt[0])
        unique: list[tuple[int, int]] = []
        for pt in eye_centers:
            if not unique or abs(unique[-1][0] - pt[0]) > 12 or abs(unique[-1][1] - pt[1]) > 12:
                unique.append(pt)
            if len(unique) >= 2:
                break
        return unique

    def _score_face_candidate(self, box: tuple[int, int, int, int], width: int, height: int, eye_count: int) -> float:
        x, y, w, h = box
        area = float(w * h)
        cx = x + w / 2
        cy = y + h / 2
        center_bias = 1.0 - min(1.0, (abs(cx - width / 2) / max(1, width)) * 1.15 + (abs(cy - height * 0.42) / max(1, height)) * 0.55)
        aspect = w / max(1, h)
        aspect_penalty = 1.0 - min(0.38, abs(aspect - 0.78) * 0.16)
        eye_bonus = 1.0 + (0.22 * min(2, eye_count))
        return area * max(0.2, center_bias) * aspect_penalty * eye_bonus

    def _detect_face_candidates(self, frame, mirrored: bool = False, offset: tuple[int, int] = (0, 0)) -> list[tuple[int, int, int, int]]:
        if frame.size == 0:
            return []
        height, width = frame.shape[:2]
        detect_frame = frame
        scale = 1.0
        if width > FACE_DETECT_MAX_WIDTH:
            scale = FACE_DETECT_MAX_WIDTH / width
            detect_frame = cv2.resize(frame, (FACE_DETECT_MAX_WIDTH, max(1, int(height * scale))), interpolation=cv2.INTER_AREA)
        gray = cv2.cvtColor(detect_frame, cv2.COLOR_BGR2GRAY)
        gray = cv2.equalizeHist(gray)
        candidates: list[tuple[int, int, int, int]] = []
        for cascade in self.face_cascades:
            if cascade.empty():
                continue
            for scale_factor, min_neighbors, min_size in (
                (1.03, 3, (28, 28)),
                (1.05, 4, (40, 40)),
                (1.08, 5, (60, 60)),
            ):
                faces = cascade.detectMultiScale(
                    gray,
                    scaleFactor=scale_factor,
                    minNeighbors=min_neighbors,
                    minSize=min_size,
                )
                for fx, fy, fw, fh in faces:
                    if scale != 1.0:
                        fx = int(fx / scale)
                        fy = int(fy / scale)
                        fw = int(fw / scale)
                        fh = int(fh / scale)
                    if mirrored:
                        fx = width - fx - fw
                    fx += offset[0]
                    fy += offset[1]
                    candidates.append((fx, fy, fw, fh))
        return candidates

    def _apply_transformations(self, frame):
        zoom = max(1.0, float(self.zoom_var.get()))
        if zoom > 1.01:
            height, width = frame.shape[:2]
            crop_w = int(width / zoom)
            crop_h = int(height / zoom)
            x1 = max(0, (width - crop_w) // 2)
            y1 = max(0, (height - crop_h) // 2)
            frame = cv2.resize(frame[y1:y1 + crop_h, x1:x1 + crop_w], (width, height), interpolation=cv2.INTER_LINEAR)

        if self.rotation_var.get() == "90 deg":
            frame = cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)
        elif self.rotation_var.get() == "180 deg":
            frame = cv2.rotate(frame, cv2.ROTATE_180)
        elif self.rotation_var.get() == "270 deg":
            frame = cv2.rotate(frame, cv2.ROTATE_90_COUNTERCLOCKWISE)

        return frame

    def _detect_primary_face(self, frame) -> Optional[tuple[int, int, int, int]]:
        if self.frame_counter % 2 != 0 and self.current_face_box is not None:
            return self.current_face_box
        if not self.face_cascades or all(cascade.empty() for cascade in self.face_cascades):
            return None
        height, width = frame.shape[:2]
        search_frames: list[tuple[np.ndarray, bool, tuple[int, int]]] = [(frame, False, (0, 0)), (cv2.flip(frame, 1), True, (0, 0))]
        if self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
            x1, y1, x2, y2 = self._expand_box(self.current_face_box, width, height, scale=2.15, pad_x=32, pad_y=32)
            search_frames.insert(0, (frame[y1:y2, x1:x2], False, (x1, y1)))

        candidates: list[tuple[int, int, int, int]] = []
        for search_frame, mirrored, offset in search_frames:
            if search_frame.size == 0:
                continue
            candidates.extend(self._detect_face_candidates(search_frame, mirrored=mirrored, offset=offset))
            if candidates:
                break

        if not candidates:
            if self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
                return self.current_face_box
            self.current_face_box = None
            self.current_eye_centers = []
            return None

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        gray = cv2.equalizeHist(gray)
        best_face: Optional[tuple[int, int, int, int]] = None
        best_eyes: list[tuple[int, int]] = []
        best_score = -1.0
        for box in candidates:
            x, y, w, h = self._clip_box(box, width, height)
            if w < 36 or h < 36:
                continue
            eyes = self._detect_eyes_in_face(gray, (x, y, w, h))
            score = self._score_face_candidate((x, y, w, h), width, height, len(eyes))
            if self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
                prev_x, prev_y, prev_w, prev_h = self.current_face_box
                prev_cx = prev_x + prev_w / 2
                prev_cy = prev_y + prev_h / 2
                cur_cx = x + w / 2
                cur_cy = y + h / 2
                distance = abs(cur_cx - prev_cx) + abs(cur_cy - prev_cy)
                score -= distance * 3.0
            if score > best_score:
                best_score = score
                best_face = (x, y, w, h)
                best_eyes = eyes

        if best_face is None:
            if self.current_face_box is not None and self.frame_counter - self.last_face_detect_frame <= FACE_HOLD_FRAMES:
                return self.current_face_box
            self.current_face_box = None
            self.current_eye_centers = []
            return None

        self.current_face_box = best_face
        self.current_eye_centers = best_eyes
        self.last_face_detect_frame = self.frame_counter
        return self.current_face_box

    def _compute_portrait_crop_box(
        self,
        width: int,
        height: int,
        face_box: Optional[tuple[int, int, int, int]],
        eye_centers: Optional[list[tuple[int, int]]] = None,
    ) -> tuple[int, int, int, int]:
        target_ratio = 3 / 4
        max_crop_h = min(height, int(width / target_ratio))
        max_crop_w = int(max_crop_h * target_ratio)

        if face_box is not None:
            fx, fy, fw, fh = face_box
            face_cx = fx + fw / 2
            face_cy = fy + fh / 2
            crop_h = max(int(fh * 2.70), int(height * 0.62))
            crop_h = min(crop_h, max_crop_h)
            crop_w = int(crop_h * target_ratio)
            x1 = int(face_cx - crop_w / 2)
            if eye_centers:
                eye_y = sum(pt[1] for pt in eye_centers) / len(eye_centers)
                y1 = int(eye_y - crop_h * 0.26)
            else:
                y1 = int(face_cy - crop_h * 0.34)
        else:
            crop_h = int(max_crop_h * 0.94)
            crop_h = max(CROP_MIN_HEIGHT, crop_h)
            crop_w = int(crop_h * target_ratio)
            x1 = (width - crop_w) // 2
            y1 = (height - crop_h) // 2

        x1 = max(0, min(x1, width - crop_w))
        y1 = max(0, min(y1, height - crop_h))
        x2 = x1 + crop_w
        y2 = y1 + crop_h
        return x1, y1, x2, y2

    def _smooth_crop_box(self, next_box: tuple[int, int, int, int]) -> tuple[int, int, int, int]:
        if self.stable_crop_box is None:
            self.stable_crop_box = next_box
            return next_box

        prev_x1, prev_y1, prev_x2, prev_y2 = self.stable_crop_box
        next_x1, next_y1, next_x2, next_y2 = next_box
        prev_w = prev_x2 - prev_x1
        prev_h = prev_y2 - prev_y1
        next_w = next_x2 - next_x1
        next_h = next_y2 - next_y1
        move_threshold_x = max(10, int(prev_w * 0.06))
        move_threshold_y = max(10, int(prev_h * 0.06))
        size_threshold_w = max(12, int(prev_w * 0.08))
        size_threshold_h = max(12, int(prev_h * 0.08))

        if (
            abs(next_x1 - prev_x1) < move_threshold_x
            and abs(next_y1 - prev_y1) < move_threshold_y
            and abs(next_w - prev_w) < size_threshold_w
            and abs(next_h - prev_h) < size_threshold_h
        ):
            return self.stable_crop_box

        alpha = 0.24
        prev_cx = (prev_x1 + prev_x2) / 2
        prev_cy = (prev_y1 + prev_y2) / 2
        next_cx = (next_x1 + next_x2) / 2
        next_cy = (next_y1 + next_y2) / 2

        blended_cx = prev_cx + (next_cx - prev_cx) * alpha
        blended_cy = prev_cy + (next_cy - prev_cy) * alpha
        blended_h = prev_h + (next_h - prev_h) * alpha
        blended_h = max(240, blended_h)
        blended_w = blended_h * (3 / 4)

        x1 = int(blended_cx - blended_w / 2)
        y1 = int(blended_cy - blended_h / 2)
        x2 = int(x1 + blended_w)
        y2 = int(y1 + blended_h)
        blended = (x1, y1, x2, y2)
        self.stable_crop_box = blended
        return blended

    def _crop_frame_with_box(self, frame, crop_box: tuple[int, int, int, int], output_size: tuple[int, int] = (1500, 2000)):
        x1, y1, x2, y2 = crop_box
        crop = frame[y1:y2, x1:x2]
        if crop.size == 0:
            return frame
        crop_h, crop_w = crop.shape[:2]
        target_w, target_h = output_size
        interpolation = cv2.INTER_AREA if target_w < crop_w or target_h < crop_h else cv2.INTER_CUBIC
        return cv2.resize(crop, output_size, interpolation=interpolation)

    def _constrain_crop_box(
        self,
        crop_box: tuple[int, int, int, int],
        width: int,
        height: int,
    ) -> tuple[int, int, int, int]:
        x1, y1, x2, y2 = crop_box
        crop_w = x2 - x1
        crop_h = y2 - y1
        crop_w = min(crop_w, width)
        crop_h = min(crop_h, height)
        crop_w = int(crop_h * (3 / 4))
        x1 = max(0, min(x1, width - crop_w))
        y1 = max(0, min(y1, height - crop_h))
        return x1, y1, x1 + crop_w, y1 + crop_h

    def _decorate_frame(self, frame):
        height, width = frame.shape[:2]
        overlay = frame.copy()
        cv2.rectangle(overlay, (0, 0), (width, 72), (35, 14, 51), -1)
        cv2.addWeighted(overlay, 0.52, frame, 0.48, 0, frame)
        is_portrait_preview = self.crop_portrait_var.get() and width < height

        if self.frame_guide_var.get():
            if is_portrait_preview:
                pad_x = int(width * 0.11)
                pad_top = int(height * 0.12)
                pad_bottom = int(height * 0.08)
                x1 = pad_x
                y1 = pad_top
                x2 = width - pad_x
                y2 = height - pad_bottom
                cx = (x1 + x2) // 2
                eye_y = y1 + int((y2 - y1) * 0.38)
                cv2.rectangle(frame, (x1, y1), (x2, y2), (93, 201, 244), 2)
                cv2.line(frame, (cx, y1), (cx, y2), (93, 201, 244), 1)
                cv2.line(frame, (x1, eye_y), (x2, eye_y), (93, 201, 244), 1)
                cv2.putText(
                    frame,
                    "Credencial auto 3:4",
                    (x1, max(88, y1 - 10)),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.48,
                    (93, 201, 244),
                    1,
                    cv2.LINE_AA,
                )
            else:
                guide_w = int(width * 0.22)
                guide_h = int(height * 0.42)
                cx = width // 2
                cy = height // 2 + 18
                x1 = max(30, cx - guide_w // 2)
                y1 = max(96, cy - guide_h // 2)
                x2 = min(width - 30, cx + guide_w // 2)
                y2 = min(height - 30, cy + guide_h // 2)
                cv2.rectangle(frame, (x1, y1), (x2, y2), (93, 201, 244), 2)
                cv2.line(frame, (cx, y1), (cx, y2), (93, 201, 244), 1)
                cv2.line(frame, (x1, cy), (x2, cy), (93, 201, 244), 1)
                cv2.putText(
                    frame,
                    "Guia credencial",
                    (x1, max(88, y1 - 10)),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.48,
                    (93, 201, 244),
                    1,
                    cv2.LINE_AA,
                )

        if self.face_guide_var.get():
            preview_face_box = self._detect_primary_face(frame)
            if preview_face_box is not None:
                fx, fy, fw, fh = preview_face_box
                cv2.rectangle(frame, (fx, fy), (fx + fw, fy + fh), (244, 201, 93), 2)
                cv2.putText(frame, "Rostro", (fx, max(88, fy - 8)), cv2.FONT_HERSHEY_SIMPLEX, 0.46, (244, 201, 93), 1, cv2.LINE_AA)

        course = self.session.course_display if self.session else "Sin sesion"
        photo_no = self.session.next_id if self.session else 1
        saved = len(self.session.records) if self.session else 0
        typed_name = self.student_var.get().strip() or "-"
        camera = self.current_camera_alias or "Sin camara"
        roster_text = "Sin lista cargada"
        if self.session is not None and self.session.has_roster:
            student = self.session.current_roster_student()
            if student is not None:
                roster_text = f"{student.display_name} | RUT {student.rut} | Restan {self.session.roster_remaining}"
            else:
                roster_text = "Lista completa"

        crop_mode = "ON" if self.crop_portrait_var.get() else "OFF"
        lines = [
            f"{course} | Foto {photo_no:03d} | Guardadas {saved}",
            f"Estudiante: {typed_name} | Cam: {camera}",
            f"Lista: {roster_text}",
            f"Enter captura | Auto credencial {'ON' if self.crop_portrait_var.get() else 'OFF'} | X cambia",
        ]

        y = 22
        for text in lines:
            cv2.putText(frame, text, (14, y), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (247, 241, 255), 1, cv2.LINE_AA)
            y += 17
        return frame

    def start_session(self) -> None:
        mode = self.mode_var.get()
        self.logger.info("Start session requested. mode=%s course=%s", mode, self.course_var.get().strip())
        photos_root = APP_ROOT / PHOTOS_DIRNAME
        photos_root.mkdir(parents=True, exist_ok=True)
        backup_root = APP_ROOT / BACKUP_PHOTOS_DIRNAME
        backup_root.mkdir(parents=True, exist_ok=True)

        if mode == "test":
            course_display = "PRUEBA"
            course_slug = "PRUEBA"
            session_dir = photos_root / TEST_FOLDER_NAME
            backup_dir = backup_root / TEST_FOLDER_NAME
        else:
            course_display = self.course_var.get().strip()
            if not course_display:
                self.logger.warning("Start session rejected: empty course in course mode.")
                messagebox.showwarning(APP_TITLE, "Escribe el nombre del curso antes de iniciar.")
                return
            course_slug = sanitize_folder_name(course_display)
            session_dir = photos_root / course_slug
            backup_dir = backup_course_dir(photos_root, course_slug)

        roster_students: list[RosterStudent] = []
        roster_course = self._resolve_roster_course(course_display)
        if self.roster_map:
            if not roster_course and len(self.roster_map) == 1:
                roster_course = next(iter(self.roster_map))
            if roster_course:
                roster_students = list(self.roster_map.get(roster_course, []))
                course_display = roster_course
                course_slug = sanitize_folder_name(course_display)
                session_dir = photos_root / course_slug
                backup_dir = backup_course_dir(photos_root, course_slug)
                self.course_var.set(course_display)
            elif len(self.roster_map) > 1:
                self.logger.warning(
                    "Start session blocked: multiple roster courses available and no match for %s",
                    course_display,
                )
                messagebox.showwarning(APP_TITLE, "Elegiste modo curso, pero la lista cargada tiene varios cursos. Selecciona uno en el campo Curso.")
                return

        session_dir.mkdir(parents=True, exist_ok=True)
        backup_dir.mkdir(parents=True, exist_ok=True)
        csv_path = session_dir / CSV_FILENAME
        if not csv_path.exists():
            csv_path.write_text("id,filename,student_name,course,rut,timestamp\n", encoding="utf-8")
        records = load_existing_records(csv_path)

        self.session = GuiSession(
            mode=mode,
            course_display=course_display,
            course_slug=course_slug,
            photos_root=photos_root,
            backup_root=backup_root,
            session_dir=session_dir,
            backup_dir=backup_dir,
            csv_path=csv_path,
            records=records,
            started_at=datetime.now(),
            roster_students=roster_students,
        )
        self.session_var.set(f"Sesion activa: {course_display} | Carpeta: {session_dir.name}")
        self.status_var.set(f"Sesion iniciada en {session_dir}")
        if roster_students:
            self.session.roster_index = self.roster_preview_index.get(course_display, 0)
            self._advance_roster_past_completed()
        self._sync_session_student_from_roster()
        self._refresh_student_card_mode()
        self._update_roster_session_label()
        self._refresh_course_view()
        self._refresh_recent()
        self.logger.info(
            "Session started. mode=%s course=%s session_dir=%s backup_dir=%s roster_students=%s",
            mode,
            course_display,
            session_dir,
            backup_dir,
            len(roster_students),
        )
        open_folder(photos_root)

    def capture_photo(self) -> None:
        if self.session is None:
            self.logger.warning("Capture requested without active session.")
            messagebox.showinfo(APP_TITLE, "Primero inicia una sesion.")
            return
        if self.current_frame is None:
            self.logger.warning("Capture requested without valid frame.")
            messagebox.showwarning(APP_TITLE, "Todavia no hay un frame valido de camara.")
            return

        roster_student = self.session.current_roster_student() if self.session.has_roster else None
        if self.session.has_roster and roster_student is None:
            self.logger.info("Capture blocked: roster completed for %s", self.session.course_display)
            messagebox.showinfo(APP_TITLE, "La lista de alumnos ya se terminó.")
            return
        if roster_student is not None:
            student_name = roster_student.display_name
        else:
            student_name = self.student_var.get().strip()
        if not student_name:
            self.logger.warning("Capture blocked: empty student name. roster=%s", bool(roster_student))
            messagebox.showwarning(APP_TITLE, "Escribe el nombre del estudiante.")
            return

        student_rut = roster_student.rut if roster_student is not None else ""
        if has_record_for_student(self.session.records, student_name, self.session.course_display, student_rut):
            self.logger.info(
                "Capture blocked: duplicate student already recorded. student=%s rut=%s course=%s",
                student_name,
                student_rut,
                self.session.course_display,
            )
            messagebox.showwarning(
                APP_TITLE,
                f"Ya existe una foto de {student_name} en este curso. No se guardo duplicado.",
            )
            if self.session.has_roster:
                self._advance_roster_past_completed()
                self._sync_session_student_from_roster()
            self._refresh_course_view(force=True)
            return
        similar_record = find_similar_record(self.current_frame, self.session.records, self.session.session_dir)
        if similar_record is not None:
            self.logger.info(
                "Capture blocked: duplicate image matches %s (%s). student=%s rut=%s course=%s",
                similar_record.student_name,
                similar_record.course,
                student_name,
                student_rut,
                self.session.course_display,
            )
            messagebox.showwarning(
                APP_TITLE,
                f"Esta foto ya coincide con {similar_record.student_name} en {similar_record.course}.\nNo se guardo otro registro.",
            )
            self._refresh_course_view(force=True)
            return

        countdown = int(self.countdown_var.get().split()[0])
        for remaining in range(countdown, 0, -1):
            self.status_var.set(f"Capturando en {remaining}...")
            self.root.update()
            self.root.after(1000)

        photo_id = self.session.next_id
        filename = self.session.filename_for(student_name, student_rut)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        image_path = self.session.session_dir / filename
        backup_path = self.session.backup_dir / filename

        if not cv2.imwrite(str(image_path), self.current_frame.copy()):
            self.logger.error("Image save failed: %s", image_path)
            messagebox.showerror(APP_TITLE, f"No se pudo guardar la imagen en {image_path}")
            return

        record = PhotoRecord(
            id=photo_id,
            filename=filename,
            student_name=student_name,
            course=self.session.course_display,
            rut=student_rut,
            timestamp=timestamp,
        )
        append_csv_record(self.session.csv_path, record)
        backup_status = "respaldo OK"
        try:
            ensure_photo_backup(image_path, backup_path)
            ensure_photo_backup(self.session.csv_path, self.session.backup_dir / CSV_FILENAME)
        except Exception as exc:
            self.logger.exception("Backup copy failed for %s: %s", filename, exc)
            backup_status = f"respaldo parcial: {exc}"
        self.session.records.append(record)
        if self.session.has_roster:
            self.session.advance_roster()
            self._advance_roster_past_completed()
            self._sync_session_student_from_roster()
        else:
            self.student_var.set("")
        self.status_var.set(f"Guardada: {filename} | {student_name} | {backup_status}")
        self._update_roster_session_label()
        self._refresh_course_view()
        self._refresh_recent()
        self.logger.info(
            "Capture saved. file=%s student=%s rut=%s course=%s records=%s backup=%s",
            filename,
            student_name,
            student_rut,
            self.session.course_display,
            len(self.session.records),
            backup_status,
        )

    def retake_last(self) -> None:
        if self.session is None or not self.session.records:
            self.logger.warning("Retake requested with no captures available.")
            messagebox.showinfo(APP_TITLE, "No hay capturas para rehacer.")
            return

        if self.session.has_roster:
            current = self.session.current_roster_student()
            current_name = current.display_name if current is not None else "el ultimo alumno"
            confirm = messagebox.askyesno(
                APP_TITLE,
                f"Se borrara la ultima captura y volveras a {current_name} para un nuevo intento.\n\nQuieres seguir?",
            )
            if not confirm:
                return

        record = self.session.records.pop()
        image_path = self.session.session_dir / record.filename
        if image_path.exists():
            image_path.unlink()
        append_retake_audit(self.session.backup_dir, record, note="reintento")
        rewrite_csv(self.session.csv_path, self.session.records)
        try:
            ensure_photo_backup(self.session.csv_path, self.session.backup_dir / CSV_FILENAME)
        except Exception:
            self.logger.exception("Failed to refresh backup CSV after retake.")
            pass
        if self.session.has_roster:
            self.session.retreat_roster()
            self._sync_session_student_from_roster()
            restored_student = self.session.current_roster_student()
            restored_name = restored_student.display_name if restored_student is not None else record.student_name
        else:
            self.student_var.set(record.student_name)
            restored_name = record.student_name
        self.status_var.set(f"Volviste atras. Listo para reintentar con: {restored_name}")
        self._update_roster_session_label()
        self._refresh_course_view()
        self._refresh_recent()
        self.logger.info("Retake completed. Removed=%s restored_student=%s", record.filename, restored_name)

    def close_session(self) -> None:
        if self.session is None:
            return
        self.status_var.set(f"Sesion cerrada: {self.session.course_display}")
        self.session_var.set("Sesion no iniciada")
        self.logger.info("Session closed. course=%s records=%s", self.session.course_display, len(self.session.records))
        self.session = None
        self.student_var.set("")
        self._refresh_student_card_mode()
        self._update_roster_preview(self.course_var.get())
        self._refresh_course_view()
        self._refresh_recent()

    def open_photos_root(self) -> None:
        photos_root = APP_ROOT / PHOTOS_DIRNAME
        photos_root.mkdir(parents=True, exist_ok=True)
        open_folder(photos_root)

    def _refresh_recent(self) -> None:
        if self.session is None or not self.session.records:
            self.recent_var.set("Sin capturas aun.")
            return
        def describe(record: PhotoRecord) -> str:
            rut = f"  {record.rut}" if getattr(record, "rut", "") else ""
            return f"{record.id:03d}  {record.filename}  {record.student_name}{rut}"

        self.recent_var.set("\n".join(describe(r) for r in self.session.records[-6:]))

    def _bind_shortcuts(self) -> None:
        self.root.bind("<Return>", self._handle_global_return)
        self.root.bind("<KP_Enter>", self._handle_global_return)
        self.root.bind("<KeyPress-c>", lambda _event: self.cycle_camera())
        self.root.bind("<KeyPress-C>", lambda _event: self.cycle_camera())
        self.root.bind("<Control-Right>", lambda _event: self.next_roster_student())
        self.root.bind("<Control-Left>", lambda _event: self.prev_roster_student())
        self.root.bind("<KeyPress-v>", lambda _event: self._toggle_mirror())
        self.root.bind("<KeyPress-V>", lambda _event: self._toggle_mirror())
        self.root.bind("<KeyPress-r>", lambda _event: self._toggle_face())
        self.root.bind("<KeyPress-R>", lambda _event: self._toggle_face())
        self.root.bind("<KeyPress-g>", lambda _event: self._toggle_guide())
        self.root.bind("<KeyPress-G>", lambda _event: self._toggle_guide())
        self.root.bind("<KeyPress-x>", lambda _event: self._toggle_crop())
        self.root.bind("<KeyPress-X>", lambda _event: self._toggle_crop())
        self.root.bind("<KeyPress-o>", lambda _event: self.open_photos_root())
        self.root.bind("<KeyPress-O>", lambda _event: self.open_photos_root())
        self.root.bind("<KeyPress-f>", lambda _event: self._focus_student())
        self.root.bind("<KeyPress-F>", lambda _event: self._focus_student())
        self.root.bind("<Escape>", lambda _event: self.student_var.set(""))

    def _toggle_mirror(self) -> None:
        self.mirror_var.set(not self.mirror_var.get())
        self.status_var.set(f"Volteo {'activado' if self.mirror_var.get() else 'desactivado'}")

    def _toggle_face(self) -> None:
        self.face_guide_var.set(not self.face_guide_var.get())
        self.status_var.set(f"Guia de rostro {'activada' if self.face_guide_var.get() else 'desactivada'}")

    def _toggle_guide(self) -> None:
        self.frame_guide_var.set(not self.frame_guide_var.get())
        self.status_var.set(f"Guia de encuadre {'activada' if self.frame_guide_var.get() else 'desactivada'}")

    def _toggle_crop(self) -> None:
        self.crop_portrait_var.set(not self.crop_portrait_var.get())
        self.status_var.set(f"Recorte automatico {'activado' if self.crop_portrait_var.get() else 'desactivado'}")

    def _focus_student(self) -> None:
        if self.student_entry is not None and self.student_entry.winfo_ismapped():
            self.student_entry.focus_set()

    def _report_tk_exception(self, exc_type, exc_value, exc_traceback) -> None:
        self.logger.critical(
            "Uncaught Tk exception: %s: %s",
            exc_type.__name__ if hasattr(exc_type, "__name__") else exc_type,
            exc_value,
            exc_info=(exc_type, exc_value, exc_traceback),
        )
        messagebox.showerror(APP_TITLE, f"Ocurrió un error inesperado.\n\n{exc_value}")

    def _sys_excepthook(self, exc_type, exc_value, exc_traceback) -> None:
        self.logger.critical(
            "Unhandled exception: %s: %s",
            exc_type.__name__ if hasattr(exc_type, "__name__") else exc_type,
            exc_value,
            exc_info=(exc_type, exc_value, exc_traceback),
        )

    def _release_capture(self) -> None:
        if self.capture is not None:
            self.capture.release()
            self.capture = None

    def on_close(self) -> None:
        self.logger.info("GUI close requested.")
        if self.preview_job is not None:
            self.root.after_cancel(self.preview_job)
        self._release_capture()
        self.logger.info("=== CastelCredCam GUI end ===")
        self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    app = CastelCredCamGUI()
    app.run()


if __name__ == "__main__":
    main()

